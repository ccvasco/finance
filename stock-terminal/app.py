#!/usr/bin/env python3
"""Stock Terminal — a self-contained stock analysis app powered by yfinance.

No web frameworks: just Python's stdlib http.server for a small JSON API plus
static file serving. Excel export uses openpyxl. All market data comes from
yfinance.

Run:
    python app.py            # serves on http://localhost:8765
    python app.py --port 9000
"""
import argparse
import datetime as _dt
import io
import json
import math
import os
import sys
import threading
import time
from functools import lru_cache
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Make the in-repo yfinance importable when run from the repo root.
sys.path.insert(0, _REPO_ROOT)


def _load_dotenv():
    """Populate os.environ from a .env file (repo root, then CWD) — e.g. for
    ANTHROPIC_API_KEY. No python-dotenv dependency; real env vars still win."""
    for base in (_REPO_ROOT, os.getcwd()):
        path = os.path.join(base, ".env")
        if not os.path.isfile(path):
            continue
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _, value = line.partition("=")
                os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


_load_dotenv()

import yfinance as yf  # noqa: E402
import pandas as pd  # noqa: E402
from concurrent.futures import ThreadPoolExecutor  # noqa: E402

from strategies import grade_row, flag_legend, _business_type  # noqa: E402  (strategy graders, same dir)
import chat as chat_agent  # noqa: E402  (chat agent for /api/chat, same dir)
import wiki  # noqa: E402  (Wikipedia company context, same dir)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    _HAS_OPENPYXL = True
except Exception:  # pragma: no cover
    _HAS_OPENPYXL = False

STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")

# --------------------------------------------------------------------------- #
# Tiny TTL cache so we don't hammer Yahoo on every request                     #
# --------------------------------------------------------------------------- #
_CACHE = {}
_CACHE_LOCK = threading.Lock()


def cached(key, ttl, producer, cache_none=True):
    """Memoize producer() under `key` for `ttl` seconds.

    cache_none=False skips storing a None result, so a transient upstream
    failure is retried on the next call rather than pinned for the whole TTL.
    Use it for shared inputs whose momentary absence poisons many rows at once
    — e.g. the risk-free rate, which gates WACC and therefore every DCF: one
    failed ^TNX fetch would otherwise blank the DCF column across the table for
    an hour."""
    now = time.time()
    with _CACHE_LOCK:
        hit = _CACHE.get(key)
        if hit and now - hit[0] < ttl:
            return hit[1]
    value = producer()  # produce outside the lock (network call)
    if value is None and not cache_none:
        return None
    with _CACHE_LOCK:
        _CACHE[key] = (now, value)
    return value


def clear_cache():
    with _CACHE_LOCK:
        _CACHE.clear()


def clear_ticker_cache(tickers):
    """Evict every cached entry for the given tickers (all cache keys embed the
    ticker as a ':'-separated segment, e.g. info:AAPL, fin:AAPL:income:annual)."""
    tset = {t.upper() for t in tickers}
    with _CACHE_LOCK:
        for k in [k for k in _CACHE if tset & set(k.split(":"))]:
            del _CACHE[k]


def clear_prefix_cache(prefix):
    """Evict cached entries whose key starts with `prefix` (e.g. 'mktcal:')."""
    with _CACHE_LOCK:
        for k in [k for k in _CACHE if k.startswith(prefix)]:
            del _CACHE[k]


# --------------------------------------------------------------------------- #
# Helpers                                                                      #
# --------------------------------------------------------------------------- #
def _num(v):
    """Coerce to a JSON-friendly float or None."""
    try:
        if v is None:
            return None
        f = float(v)
        if f != f or f in (float("inf"), float("-inf")):  # NaN / inf
            return None
        return f
    except (TypeError, ValueError):
        return None


def _epoch_to_iso(v):
    try:
        if not v:
            return None
        return _dt.datetime.fromtimestamp(int(v), _dt.timezone.utc).strftime("%Y-%m-%d")
    except (TypeError, ValueError, OSError):
        return None


# Company-profile fields (currency-independent) that a regional listing may be
# missing but that live on the primary Xetra listing — backfilled by _fetch_info.
_PROFILE_FIELDS = ("longBusinessSummary", "sector", "industry", "sectorKey",
                   "industryKey", "website", "city", "state", "country")

# German regional exchanges that return only a bare quote (name + price) with no
# company profile. The profile lives on the primary Xetra (.DE) listing, with
# Frankfurt (.F) as a secondary source.
_REGIONAL_DE_SUFFIXES = ("MU", "HM", "DU", "SG", "BE", "HA", "BM", "F")


def _fetch_info(ticker):
    info = yf.Ticker(ticker).info or {}
    # Regional German listings (.MU Munich, .HM Hamburg, …) carry no company
    # profile — no business summary, sector, or industry — so the ticker cell
    # would show no hover legend and blank sector/industry columns. Backfill
    # those descriptive fields from the primary Xetra (.DE) / Frankfurt (.F)
    # listing of the same symbol. Cached with the rest of the info payload.
    base, _, suffix = ticker.rpartition(".")
    if base and suffix.upper() in _REGIONAL_DE_SUFFIXES and not info.get("longBusinessSummary"):
        for alt_suffix in ("DE", "F"):
            if alt_suffix == suffix.upper():
                continue
            try:
                alt = yf.Ticker(f"{base}.{alt_suffix}").info or {}
            except Exception:
                continue
            if alt.get("longBusinessSummary"):
                for f in _PROFILE_FIELDS:
                    if not info.get(f) and alt.get(f):
                        info[f] = alt[f]
                break
    return info


def get_info(ticker):
    return cached(f"info:{ticker}", 300, lambda: _fetch_info(ticker))


def _brief_summary(text, limit=320):
    """First sentence(s) of a business summary, cut at a sentence boundary
    under `limit` chars — the screener's hover blurb for a ticker."""
    if not text:
        return None
    text = str(text).strip()
    if len(text) <= limit:
        return text
    cut = text[:limit]
    # prefer ending on a sentence; fall back to a word boundary + ellipsis
    dot = cut.rfind(". ")
    if dot > limit // 3:
        return cut[:dot + 1]
    return cut.rsplit(" ", 1)[0] + "…"


_ERP = 5.5  # Equity risk premium %, Damodaran US estimate

def get_risk_free_rate():
    """10-year US Treasury yield (%) from ^TNX, cached 1 hour."""
    def produce():
        try:
            info = yf.Ticker("^TNX").info
            return _num(info.get("regularMarketPrice") or info.get("previousClose"))
        except Exception:
            return None
    # Don't negatively-cache a failed fetch: the risk-free rate gates WACC for
    # every ticker, so pinning a None here would blank the whole DCF column for
    # an hour off a single transient ^TNX hiccup.
    return cached("rfr:^TNX", 3600, produce, cache_none=False)


def get_risk_free_history():
    """10-year US Treasury yield (%) daily-close series (^TNX), cached 1 hour.
    Used to look up the risk-free rate as of each historical fiscal year-end
    for the per-period WACC line (see _period_wacc in deepdive)."""
    def produce():
        try:
            df = yf.Ticker("^TNX").history(period="10y", auto_adjust=False)
        except Exception:
            return None
        if df is None or df.empty or "Close" not in df:
            return None
        return df["Close"].dropna()
    return cached("rfrhist:^TNX", 3600, produce)


def _series_asof(series, date_str):
    """Most-recent value in a (tz-aware) price/yield series at or before
    `date_str` (an 'YYYY-MM-DD' fiscal period end). None if unavailable or the
    date predates the series. Reconciles the naive fiscal date with the
    series' timezone so pandas' asof doesn't raise on the mismatch."""
    if series is None or getattr(series, "empty", True):
        return None
    try:
        ts = pd.Timestamp(date_str)
        tz = getattr(series.index, "tz", None)
        if tz is not None and ts.tz is None:
            ts = ts.tz_localize(tz)
        return _num(series.asof(ts))
    except Exception:
        return None


def _fx_rate(base, quote):
    """1 unit of `base` currency in `quote` currency, cached 1 hour.
    Used only for foreign-reporting tickers (ADRs etc.) whose financial
    statements come back in a different currency than the one they trade in
    — e.g. WIT trades in USD but reports in INR (financialCurrency)."""
    if not base or not quote or base == quote:
        return 1.0
    def produce():
        try:
            info = yf.Ticker(f"{base}{quote}=X").info
            return _num(info.get("regularMarketPrice") or info.get("previousClose"))
        except Exception:
            return None
    # Same reasoning as the risk-free rate: a failed FX fetch blanks the DCF and
    # the recomputed valuation ratios for every foreign-reporting ticker, so
    # retry next call rather than pinning the None for an hour.
    return cached(f"fx:{base}{quote}", 3600, produce, cache_none=False)


def _native_market_cap(market_cap, mkt_ccy, fin_ccy):
    """Market cap (trading currency) converted into the company's reporting
    currency, so it can be combined with statement figures (debt, cash,
    revenue, EBITDA — all in `fin_ccy`) without a currency mismatch. Returns
    the original market_cap unchanged when the two currencies already match,
    or None if a conversion is needed but the FX rate isn't available."""
    if market_cap is None:
        return None
    if not fin_ccy or not mkt_ccy or fin_ccy == mkt_ccy:
        return market_cap
    fx = _fx_rate(mkt_ccy, fin_ccy)
    return market_cap * fx if fx else None


def _wacc_detail(beta, market_cap, total_debt, interest_exp, tax_rate, rfr):
    """WACC plus every intermediate component (capital weights, CAPM cost of
    equity, cost of debt, tax rate) as a dict — the DCF export's discount-rate
    breakdown. None under the same conditions _compute_wacc returns None."""
    if beta is None or market_cap is None or rfr is None:
        return None
    total_debt = total_debt or 0
    total_capital = market_cap + total_debt
    if total_capital <= 0:
        return None
    w_e = market_cap / total_capital
    w_d = total_debt / total_capital
    cost_of_equity = rfr + beta * _ERP
    if interest_exp and total_debt > 0:
        cost_of_debt = abs(interest_exp) / total_debt * 100
    else:
        cost_of_debt = rfr  # fallback: riskless floor when no interest data
    return {
        "beta": beta, "rfr": rfr, "erp": _ERP,
        "w_e": w_e, "w_d": w_d, "tax_rate": tax_rate,
        "cost_of_equity": cost_of_equity, "cost_of_debt": cost_of_debt,
        "wacc": w_e * cost_of_equity + w_d * cost_of_debt * (1 - tax_rate),
    }


def _compute_wacc(beta, market_cap, total_debt, interest_exp, tax_rate, rfr):
    """WACC in %. Cost of equity via CAPM; cost of debt from interest expense / total debt
    (falls back to rfr when interest expense is unavailable)."""
    d = _wacc_detail(beta, market_cap, total_debt, interest_exp, tax_rate, rfr)
    return None if d is None else d["wacc"]


def _get_stmt(ticker, attr):
    def produce():
        try:
            return getattr(yf.Ticker(ticker), attr)
        except Exception:
            return None
    return cached(f"{attr}:{ticker}", 1800, produce)


def get_dividends(ticker):
    def produce():
        try:
            return yf.Ticker(ticker).dividends
        except Exception:
            return None
    return cached(f"divs:{ticker}", 1800, produce)


def get_raw_close(ticker):
    """Dividend-UNadjusted (but split-adjusted) close series for price-only
    performance. period=10y so we can cover the 10Y window."""
    def produce():
        try:
            df = yf.Ticker(ticker).history(period="10y", auto_adjust=False)
        except Exception:
            return None
        if df is None or df.empty or "Close" not in df:
            return None
        return df["Close"].dropna()
    return cached(f"rawclose:{ticker}", 1800, produce)


def _stmt_val(df, *labels, col=0):
    """Value of the first matching row at the `col`-th most-recent period."""
    if df is None or getattr(df, "empty", True):
        return None
    cols = sorted(df.columns, reverse=True)  # newest first
    if col >= len(cols):
        return None
    c = cols[col]
    for lab in labels:
        if lab in df.index:
            return _num(df.loc[lab, c])
    return None


def performance(ticker):
    """Price-only return (excludes dividends) over several windows, in %."""
    out = {"ytd": None, "1y": None, "3y": None, "5y": None, "10y": None}
    s = get_raw_close(ticker)
    if s is None or len(s) == 0:
        return out
    latest = float(s.iloc[-1])
    last_date = s.index[-1]
    span_years = (last_date - s.index[0]).days / 365.25

    def at_or_before(ts):
        sub = s[s.index <= ts]
        return float(sub.iloc[-1]) if len(sub) else None

    def pct(base):
        return ((latest / base) - 1.0) * 100.0 if base else None

    try:
        jan1 = pd.Timestamp(year=last_date.year, month=1, day=1, tz=s.index.tz)
        prev = s[s.index < jan1]
        out["ytd"] = _num(pct(float(prev.iloc[-1]) if len(prev) else None))
    except Exception:
        pass
    for label, yrs in (("1y", 1), ("3y", 3), ("5y", 5), ("10y", 10)):
        try:
            base = at_or_before(last_date - pd.DateOffset(years=yrs))
            if base is None and span_years >= yrs * 0.95:
                base = float(s.iloc[0])
            out[label] = _num(pct(base))
        except Exception:
            pass
    return out


def _downsample(vals, target=50):
    """Evenly pick ~`target` points (always including the first and last) so a
    5-year daily series compresses to a compact sparkline payload."""
    n = len(vals)
    if n <= target:
        return list(vals)
    idx = [int(k * (n - 1) / (target - 1)) for k in range(target)]
    return [vals[j] for j in idx]


def sparklines(ticker):
    """Downsampled close-price series for the 6mo / 1y / 5y screener mini-charts.
    Reuses the dividend-unadjusted close history already cached by get_raw_close
    (also used for the performance columns), so it adds no extra network calls."""
    s = get_raw_close(ticker)
    out = {"6mo": [], "1y": [], "5y": []}
    if s is None or len(s) == 0:
        return out
    last_date = s.index[-1]
    for key, off in (("6mo", pd.DateOffset(months=6)),
                     ("1y", pd.DateOffset(years=1)),
                     ("5y", pd.DateOffset(years=5))):
        try:
            sub = s[s.index >= (last_date - off)]
            vals = [round(float(v), 4) for v in sub.values if v == v]
            out[key] = _downsample(vals)
        except Exception:
            pass
    return out


def consecutive_div_increases(divs):
    """Number of consecutive completed years of rising annual dividends."""
    if divs is None or len(divs) == 0:
        return None
    annual = divs.groupby(divs.index.year).sum()
    annual = annual[annual.index < _dt.date.today().year]
    vals = list(annual.items())  # ascending by year
    if len(vals) < 2:
        return None
    count = 0
    for i in range(len(vals) - 1, 0, -1):
        if vals[i][1] > vals[i - 1][1] * 1.0001:
            count += 1
        else:
            break
    return count


def screener_row(ticker):
    """One row of comparison metrics. Missing values come back as None -> N/A.
    Any per-ticker failure (Yahoo rate limit, delisted symbol, parse error)
    becomes an error row, so one bad ticker can never 500 its whole batch."""
    try:
        return _screener_row(ticker)
    except Exception as e:
        return {"ticker": ticker, "error": str(e) or type(e).__name__}


def _screener_row(ticker):
    try:
        info = get_info(ticker)
    except Exception as e:
        return {"ticker": ticker, "error": str(e)}

    if not info or info.get("regularMarketPrice") is None and info.get("currentPrice") is None and not info.get("shortName"):
        return {"ticker": ticker, "error": "No data"}

    price = _num(info.get("currentPrice")) or _num(info.get("regularMarketPrice"))
    market_cap = _num(info.get("marketCap"))
    total_cash = _num(info.get("totalCash"))
    div_rate = _num(info.get("dividendRate"))            # forward annual estimate
    div_yield = (div_rate / price * 100) if (div_rate and price) else None

    # Foreign-reporting tickers (ADRs etc.) trade in one currency but file
    # statements in another — e.g. WIT trades in USD, reports in INR. Yahoo's
    # own cross-currency fields (enterpriseValue, priceToSalesTrailing12Months,
    # enterpriseToEbitda) get this wrong, so those three are recomputed below
    # from a market cap converted into the statement (financial) currency.
    mkt_ccy = info.get("currency")
    fin_ccy = info.get("financialCurrency")
    market_cap_native = _native_market_cap(market_cap, mkt_ccy, fin_ccy)
    fx_mismatch = bool(fin_ccy and mkt_ccy and fin_ccy != mkt_ccy)
    total_revenue = _num(info.get("totalRevenue"))

    # Heavier pulls (cached): statements, dividends, raw price history.
    inc = _get_stmt(ticker, "income_stmt")
    q_inc = _get_stmt(ticker, "quarterly_income_stmt")
    bal = _get_stmt(ticker, "balance_sheet")
    cf = _get_stmt(ticker, "cash_flow")
    divs = get_dividends(ticker)

    equity = _stmt_val(bal, "Stockholders Equity", "Total Stockholder Equity", "Common Stock Equity")
    lt_debt = _stmt_val(bal, "Long Term Debt", "Long Term Debt And Capital Lease Obligation")
    total_debt = _num(info.get("totalDebt")) or _stmt_val(bal, "Total Debt")
    total_assets = _stmt_val(bal, "Total Assets")
    current_liab = _stmt_val(bal, "Current Liabilities", "Total Current Liabilities")
    ebit = _stmt_val(inc, "EBIT", "Operating Income", "Operating Income Or Loss")
    pretax = _stmt_val(inc, "Pretax Income", "Income Before Tax")
    tax = _stmt_val(inc, "Tax Provision", "Income Tax Expense")
    net_income = _num(info.get("netIncomeToCommon")) or _stmt_val(inc, "Net Income", "Net Income Common Stockholders")
    fcf = _stmt_val(cf, "Free Cash Flow") or _num(info.get("freeCashflow"))
    # Earnings/cash consistency across the statement history (see S3 Pillar D).
    # Same labels the scalars above and piotroski_f resolve, so the counts and
    # the latest-year figures can never disagree about which line they mean.
    ni_pos_yrs, ni_yrs = _positive_year_count(
        inc, "Net Income", "Net Income Common Stockholders")
    fcf_pos_yrs, fcf_yrs = _positive_year_count(cf, "Free Cash Flow")
    div_paid = _stmt_val(cf, "Cash Dividends Paid", "Common Stock Dividend Paid", "Common Stock Dividends Paid")
    bt = _business_type({"sector": info.get("sector"), "industry": info.get("industry")})
    # Approximate NAREIT FFO (Net Income + D&A − property-sale gains + impairments)
    # — equity REITs only. The two adjustments matter: without them a REIT flatters
    # its payout by selling buildings (the gain lands in net income) and overstates
    # stress in an impairment year. Both come off the cash flow statement's
    # operating reconciliation, where they are already signed to add — a backed-out
    # gain is negative, an added-back loss or impairment positive — so the line is
    # summed as reported. Absent lines are 0: the adjustment is optional, the D&A
    # add-back is not. Still an approximation, not filed NAREIT FFO — "Operating
    # Gains Losses" can carry non-property items, and NAREIT adds back only real
    # estate depreciation where D&A also covers intangibles. Treat as directional.
    # Mortgage REITs are excluded by type, not by absence of a D&A line: FFO's
    # whole purpose is the depreciation add-back, and they own securities rather
    # than depreciable buildings, so it degenerates to net income while the gain
    # backout would strip the securities gains that *are* their business. They're
    # graded on book value instead, and the whole FFO family stays None (REITs.md).
    d_and_a = _stmt_val(cf, "Depreciation And Amortization", "Depreciation Amortization Depletion")
    ffo_gains = _stmt_val(cf, "Operating Gains Losses") or 0.0
    ffo_impair = _stmt_val(cf, "Asset Impairment Charge") or 0.0
    ffo = ((net_income + d_and_a + ffo_gains + ffo_impair)
           if (bt == "reit" and net_income is not None and d_and_a is not None) else None)
    # Book value per share trend — the mortgage-REIT quality signal (see grader).
    bvps_growth = _bvps_growth(bal)
    # Revenue per share trend — S2's fundamental-compounding signal (see grader).
    rps_growth = _rps_growth(inc, bal)

    # Derived ratios -----------------------------------------------------
    # Uses market_cap_native (not market_cap) so a mismatched-currency ticker
    # (WIT: USD price, INR statements) doesn't divide a USD numerator by an
    # INR denominator — market_cap_native equals market_cap unchanged for the
    # (overwhelmingly common) case where the two currencies already match.
    p_c = (market_cap_native / total_cash) if (market_cap_native and total_cash) else None
    p_fcf = (market_cap_native / fcf) if (market_cap_native and fcf and fcf > 0) else None
    # Debt/Equity computed from the same Total Debt and Total Equity shown in the
    # row, so the three figures reconcile. Yahoo's own debtToEquity (most-recent
    # quarter) is surfaced separately as debt_to_equity_mrq.
    debt_eq = (total_debt / equity * 100) if (total_debt and equity and equity != 0) else None
    lt_debt_eq = (lt_debt / equity * 100) if (lt_debt and equity and equity != 0) else None
    fcf_coverage = (fcf / abs(div_paid)) if (fcf is not None and div_paid) else None
    ebitda = _num(info.get("ebitda"))
    debt_ebitda = (total_debt / ebitda) if (total_debt and ebitda and ebitda > 0) else None
    ebitda_fcf = (ebitda / fcf) if (ebitda and fcf) else None
    # Enterprise Value, EV/EBITDA and Price/Sales all combine market cap with
    # statement figures — Yahoo's own passthrough fields for these get the
    # arithmetic wrong for a mismatched-currency ticker, so recompute them
    # from market_cap_native instead of trusting info.get(...) in that case.
    if fx_mismatch and market_cap_native is not None:
        enterprise_value = market_cap_native + (total_debt or 0) - (total_cash or 0)
        ev_ebitda = (enterprise_value / ebitda) if ebitda else None
        ps = (market_cap_native / total_revenue) if total_revenue else None
    else:
        enterprise_value = _num(info.get("enterpriseValue"))
        ev_ebitda = _num(info.get("enterpriseToEbitda"))
        ps = _num(info.get("priceToSalesTrailing12Months"))

    _q_basic = _series_from_stmt(q_inc, "Basic EPS")
    _q_basic_vals = [x["value"] for x in _q_basic[-4:] if x.get("value") is not None]
    basic_eps = sum(_q_basic_vals) if len(_q_basic_vals) == 4 else None

    tax_rate = 0.21
    if pretax and tax and pretax > 0:
        tax_rate = min(max(tax / pretax, 0.0), 0.5)

    roic = None
    if ebit and equity and total_debt is not None:
        invested = (total_debt or 0) + equity
        if invested:
            roic = ebit * (1 - tax_rate) / invested * 100

    roce = None  # EBIT / (total assets - current liabilities)
    if ebit and total_assets and current_liab is not None:
        capital_employed = total_assets - current_liab
        if capital_employed:
            roce = ebit / capital_employed * 100

    interest_exp = (_stmt_val(inc, "Interest Expense", "Interest Expense Non Operating") or
                    _stmt_val(inc, "Interest Expense", "Interest Expense Non Operating", col=1))
    wacc = _compute_wacc(_num(info.get("beta")), market_cap_native, total_debt,
                         interest_exp, tax_rate, get_risk_free_rate())
    # Times the coupon is covered by operating earnings — the question a
    # balance-sheet snapshot never asks, and the one that decides whether a
    # levered company actually survives (S3 Pillar C). abs() because the sign
    # convention on the interest line varies by filer. Negative EBIT yields a
    # negative ratio, which fails every band on its own — no guard needed.
    interest_coverage = ((ebit / abs(interest_exp))
                         if (ebit is not None and interest_exp) else None)

    # --- DCF fair value (10y two-stage FCFF; see _dcf_equity_value) --------
    # Blanked for balance-sheet archetypes: FCFF is meaningless for banks
    # (financial) and depreciation/capex-distorted for property businesses
    # (reit, mreit — FFO/book value are their metrics, not FCF).
    shares = _num(info.get("sharesOutstanding"))
    if not shares and market_cap and price:
        shares = market_cap / price      # both trading-ccy -> currency cancels
    dcf_value = dcf_upside = dcf_value_native = None
    if bt not in ("financial", "reit", "mreit") and shares:
        dcf_equity = _dcf_equity_value(fcf, _fcf_cagr(cf), wacc, total_debt, total_cash)
        if dcf_equity is not None:
            # equity value is in the reporting currency; convert the per-share
            # figure into the trading currency so it compares against Price
            # (fx = 1.0 when the currencies match — the common case). Keep the
            # unconverted per-share figure too (dcf_value_native, in the
            # reporting currency) so a mismatched-currency ticker can show what
            # the model actually produced before the FX step, as a hover.
            fx = _fx_rate(fin_ccy, mkt_ccy) if fx_mismatch else 1.0
            if fx:
                dcf_value_native = dcf_equity / shares
                dcf_value = dcf_value_native * fx
                if price:
                    dcf_upside = (dcf_value / price - 1) * 100

    dg = dividend_growth(ticker)
    perf = performance(ticker)
    spark = sparklines(ticker)

    row = {
        "ticker": ticker,
        "name": info.get("shortName") or info.get("longName") or ticker,
        # short business description, shown as the ticker cell's hover tooltip
        "summary": _brief_summary(info.get("longBusinessSummary")),
        # mini price-chart series (dividend-unadjusted close), for the screener
        "spark_6mo": spark["6mo"],
        "spark_1y": spark["1y"],
        "spark_5y": spark["5y"],
        "sector": info.get("sector"),
        "industry": info.get("industry"),
        "currency": info.get("currency"),
        # currency the financial statements (revenue, cash, debt, EBITDA, net
        # income, FCF, book value, etc.) are actually reported in — usually
        # the same as `currency`, but foreign-reporting tickers like ADRs
        # (e.g. WIT trades in USD, reports in INR) diverge. Falls back to
        # `currency` on the frontend when absent.
        "financial_currency": fin_ccy,
        "price": price,
        "market_cap": market_cap,
        "enterprise_value": _num(enterprise_value),
        # valuation
        "pe": _num(info.get("trailingPE")),
        "forward_pe": _num(info.get("forwardPE")),
        "peg": _num(info.get("trailingPegRatio")) or _num(info.get("pegRatio")),
        "pb": _num(info.get("priceToBook")),
        "ps": _num(ps),
        "pc": _num(p_c),
        "p_fcf": _num(p_fcf),
        "ev_ebitda": _num(ev_ebitda),
        # DCF fair value per share (trading currency, comparable to price) and
        # its premium/discount to the current price in signed percentage
        # points. None for financials and REITs (FCFF doesn't fit), FCF <= 0,
        # or when WACC is unavailable / too close to the terminal growth rate.
        "dcf_value": _num(dcf_value),
        # same DCF fair value per share, but left in the reporting currency
        # (financial_currency) — equals dcf_value when the trading and reporting
        # currencies match. The frontend surfaces it as a hover on the DCF Value
        # cell for mismatched-currency tickers, so the FX-converted figure isn't
        # read as a native reporting-currency number.
        "dcf_value_native": _num(dcf_value_native),
        "dcf_upside": _num(dcf_upside),
        "eps": _num(info.get("trailingEps")),       # diluted EPS TTM
        "eps_basic": _num(basic_eps),               # basic EPS TTM
        # profitability / income
        "income": _num(net_income),
        "profit_margin": _num(info.get("profitMargins")),
        "gross_margin": _num(info.get("grossMargins")),
        "operating_margin": _num(info.get("operatingMargins")),
        "ebitda_margin": _num(info.get("ebitdaMargins")),
        "fcf": _num(fcf),
        # Positive-year counts and the history depth behind them. None (not 0)
        # when the line has no usable history at all — a young company has no
        # track record, which is not the same claim as "never profitable".
        "ni_positive_years": ni_pos_yrs if ni_yrs else None,
        "ni_years": ni_yrs or None,
        "fcf_positive_years": fcf_pos_yrs if fcf_yrs else None,
        "fcf_years": fcf_yrs or None,
        "roa": _num(info.get("returnOnAssets")),
        "roe": _num(info.get("returnOnEquity")),
        "roic": _num(roic),
        "roce": _num(roce),
        "wacc": _num(wacc),
        # Excess return over cost of capital: > 0 means the business earns more
        # than it pays for capital (value creation), < 0 means value destruction.
        "roic_wacc": _num(roic - wacc) if (roic is not None and wacc is not None) else None,
        "revenue_per_share": _num(info.get("revenuePerShare")),
        # financial health
        "debt_to_equity": _num(debt_eq),
        "debt_to_equity_mrq": _num(info.get("debtToEquity")),
        "debt_ebitda": _num(debt_ebitda),
        "lt_debt_to_equity": _num(lt_debt_eq),
        "current_ratio": _num(info.get("currentRatio")),
        "quick_ratio": _num(info.get("quickRatio")),
        "total_cash": _num(total_cash),
        "total_debt": _num(total_debt),
        "total_equity": _num(equity),
        "ebitda": _num(ebitda),
        "ebitda_fcf": _num(ebitda_fcf),
        "interest_expense": _num(interest_exp),
        "interest_coverage": _num(interest_coverage),
        # Equity-REIT-only (None everywhere else, mortgage REITs included):
        # approximate NAREIT FFO (see comment above), P/FFO, FFO payout
        # (fraction paid out, mirrors payout_ratio) and FFO coverage (multiple,
        # mirrors fcf_coverage).
        "ffo": _num(ffo),
        "p_ffo": _num(market_cap_native / ffo) if (market_cap_native and ffo and ffo > 0) else None,
        "ffo_payout": _num(abs(div_paid) / ffo) if (ffo and ffo > 0 and div_paid) else None,
        "ffo_coverage": _num(ffo / abs(div_paid)) if (ffo is not None and div_paid) else None,
        # book value per share trend (annualized %), the mortgage-REIT signal
        "bvps_growth": _num(bvps_growth),
        # revenue per share trend (annualized %), S2's fundamental-growth signal
        "rps_growth": _num(rps_growth),
        # dividend
        "div_yield": _num(div_yield),
        "five_year_avg_yield": _num(info.get("fiveYearAvgDividendYield")),
        "payout_ratio": _num(info.get("payoutRatio")),
        "div_growth_3y": dg.get("cagr_3y"),
        "div_growth_5y": dg.get("cagr_5y"),
        "dividend_estimate": _num(div_rate),
        "dividend_ttm": _num(info.get("trailingAnnualDividendRate")),
        "fcf_coverage": _num(fcf_coverage),
        "years_div_increase": consecutive_div_increases(divs),
        "ex_dividend_date": _epoch_to_iso(info.get("exDividendDate")),
        # risk
        "beta": _num(info.get("beta")),
        "short_interest": _num(info.get("shortPercentOfFloat")),
        "days_to_cover": _num(info.get("shortRatio")),
        "altman_z": altman_z(inc, bal, market_cap_native),
        # The price-free variant, on its own bands (2.9 / 1.23). S3 grades on
        # this one; S1's kill-switch stays on the classic Z. See altman_z_prime.
        "altman_z_prime": altman_z_prime(inc, bal),
        "piotroski_f": piotroski_f(inc, bal, cf),
        # performance (price only, excludes dividends), %
        "perf_ytd": perf["ytd"],
        "perf_1y": perf["1y"],
        "perf_3y": perf["3y"],
        "perf_5y": perf["5y"],
        "perf_10y": perf["10y"],
    }
    row.update(grade_row(row))
    return row


def dividend_growth(ticker):
    def produce():
        try:
            divs = yf.Ticker(ticker).dividends
        except Exception:
            return {"cagr_3y": None, "cagr_5y": None, "annual": []}
        out = {"cagr_3y": None, "cagr_5y": None, "annual": []}
        if divs is None or len(divs) == 0:
            return out
        annual = divs.groupby(divs.index.year).sum()
        cur_year = _dt.date.today().year
        annual = annual[annual.index < cur_year]  # only completed years
        out["annual"] = [{"year": int(y), "value": _num(v)} for y, v in annual.items()][-8:]

        def cagr(n):
            if len(annual) < n + 1:
                return None
            end = float(annual.iloc[-1])
            start = float(annual.iloc[-1 - n])
            if start <= 0 or end <= 0:
                return None
            return ((end / start) ** (1.0 / n) - 1.0) * 100.0

        out["cagr_3y"] = _num(cagr(3))
        out["cagr_5y"] = _num(cagr(5))
        return out

    return cached(f"divgrowth:{ticker}", 1800, produce)


# Yahoo's two share-count rows, preferred first. They are NOT synonyms:
# "Share Issued" includes treasury shares, "Ordinary Shares Number" does not
# (for KO the two differ by 64%, for XOM by 92%).
_SHARES_LABELS = ("Ordinary Shares Number", "Share Issued")


def _stmt_label(df, *labels):
    """Which of `labels` _series_from_stmt would actually match (None if none).

    For fallback chains whose labels mean *different* things — "Ordinary Shares
    Number" is shares outstanding, "Share Issued" is outstanding + treasury —
    the caller has to know which one it got, or it will silently report one as
    the other."""
    if df is None or getattr(df, "empty", True):
        return None
    for label in labels:
        if label in df.index:
            return label
    return None


def _series_from_stmt(df, *labels):
    """Pull the first matching row from a statement df -> oldest->newest series."""
    if df is None or getattr(df, "empty", True):
        return []
    for label in labels:
        if label in df.index:
            row = df.loc[label]
            cols = sorted(df.columns)  # period timestamps ascending
            return [
                {"period": str(getattr(c, "date", lambda: c)()), "value": _num(row[c])}
                for c in cols
            ]
    return []


def _positive_year_count(df, *labels):
    """(positive periods, usable periods) across a statement line's annual
    history — the earnings-stability signal S3's quality pillar grades on.
    Yahoo's free feed carries ~4 annual periods, so this is a shorter window
    than Graham's ten-year test but the same idea: consistency, not the sign of
    whichever year happens to be last. Periods where the line is blank count
    neither way; (0, 0) when the line is absent entirely, which callers must
    read as "no history available" rather than "never positive"."""
    usable = [s["value"] for s in _series_from_stmt(df, *labels)
              if s["value"] is not None]
    return sum(1 for v in usable if v > 0), len(usable)


def _shares_outstanding_series(bal):
    """Shares outstanding per period (oldest->newest) + the basis it came from.

    Every caller that wants a share count should come through here rather than
    falling back across _SHARES_LABELS itself: "Share Issued" counts treasury
    shares that "Ordinary Shares Number" excludes, so treating them as synonyms
    silently overstates the count (KO: 7.04B vs 4.30B).

    basis is one of:
      "outstanding" — Yahoo's own outstanding row.
      "derived"     — no outstanding row, so reconstructed as issued − treasury.
                      Exact, not an estimate: the identity holds to the share
                      across Yahoo's data, so callers can treat it as they would
                      "outstanding".
      "issued"      — only issued shares, with no treasury row to net off. The
                      count may include treasury shares; callers that surface it
                      should say so.
      None          — no share data at all.
    """
    label = _stmt_label(bal, *_SHARES_LABELS)
    if label is None:
        return [], None
    series = _series_from_stmt(bal, label)
    if label == "Ordinary Shares Number":
        return series, "outstanding"
    treasury = {t["period"]: t["value"] for t in _series_from_stmt(bal, "Treasury Shares Number")}
    if not treasury:
        # Can't verify what's in the issued count. Usually treasury is simply nil
        # (issued == outstanding for AAPL/MSFT/WMT), but "usually" isn't a basis
        # for silently relabelling — hand it back as issued and let the caller warn.
        return series, "issued"
    return [{"period": s["period"],
             "value": (None if s["value"] is None
                       else s["value"] - (treasury.get(s["period"]) or 0.0))}
            for s in series], "derived"


def _trend_growth_pct(values, min_periods):
    """Annualized growth (%) from a least-squares fit of log(value) against
    time across an annual series (oldest → newest, one value per period).

    The per-share trend fields (_bvps_growth, _rps_growth) used to be
    endpoint-to-endpoint CAGR — which reads only the first and last year, so
    with the ~4 annual periods Yahoo's free feed carries, one anomalous base
    year (a 2021 trough) controlled the entire number: a company that merely
    rebounded printed the same "growth" as one that compounded. The fitted
    slope uses every period, so an endpoint loses much of that leverage (and
    with 2–3 evenly spaced points it degenerates to exactly the endpoint CAGR
    it replaces — the estimators only diverge once there are ≥4 points for the
    middle years to pull on). Deliberately NOT used for the DCF's FCF growth
    input: FCF is lumpy enough that its endpoint guard plus the DCF's 0–20%
    clamp is the better regularizer there.

    None when fewer than min_periods values exist or any value is non-positive
    (a log-trend has no meaning through zero — and a per-share figure that went
    negative mid-window is not a growth story to extrapolate)."""
    if len(values) < min_periods or any(v is None or v <= 0 for v in values):
        return None
    n = len(values)
    ys = [math.log(v) for v in values]
    x_bar = (n - 1) / 2
    y_bar = sum(ys) / n
    sxx = sum((x - x_bar) ** 2 for x in range(n))
    sxy = sum((x - x_bar) * (y - y_bar) for x, y in enumerate(ys))
    return (math.exp(sxy / sxx) - 1) * 100    # annual statements ≈ 1 yr apart


def _bvps_growth(bal):
    """Annualized trend (%) of book value per share across the balance-sheet
    history (BVPS = shareholders' equity ÷ shares outstanding, per period) —
    a log-linear fit over all periods, see _trend_growth_pct.

    The headline quality signal for a mortgage REIT: a mREIT that grows or holds
    book value while paying its dividend is compounding; one that erodes book
    value to fund the payout is destroying capital, however high the yield.
    None when fewer than two comparable years exist or book value is/was
    non-positive (a ratio then has no meaningful sign)."""
    eq = {s["period"]: s["value"] for s in _series_from_stmt(
        bal, "Stockholders Equity", "Total Stockholder Equity", "Common Stock Equity")
        if s["value"]}
    # Book value *per share* needs the outstanding count — issued shares would
    # divide equity across treasury stock the company holds itself, understating
    # BVPS and hiding the accretion from buybacks (issued stays flat while
    # outstanding falls), which is exactly the signal this metric exists to catch.
    sh = {s["period"]: s["value"] for s in _shares_outstanding_series(bal)[0] if s["value"]}
    periods = sorted(p for p in eq if p in sh)
    return _trend_growth_pct([eq[p] / sh[p] for p in periods], 2)


def _rps_growth(inc, bal):
    """Annualized trend (%) of revenue per share across the income-statement
    history (revenue ÷ shares outstanding, per period) — a log-linear fit over
    all periods, see _trend_growth_pct.

    The fundamental-compounding signal S2's track-record pillar scores:
    per-share so dilution can't fake growth (and buybacks rightly show up),
    revenue so it works across business models where EPS/FCF swing on
    accounting or cycle. Unlike _bvps_growth this requires ≥3 comparable
    periods (2 intervals): the pillar reads it as a *trend*, and a single
    year-over-year interval is a data point, not a track record. None when the
    history is shorter (the grader then falls back to price CAGR)."""
    rev = {s["period"]: s["value"] for s in _series_from_stmt(
        inc, "Total Revenue", "TotalRevenue") if s["value"]}
    # Outstanding, not issued — see _bvps_growth on why treasury shares would
    # hide exactly the buyback accretion a per-share series exists to catch.
    sh = {s["period"]: s["value"] for s in _shares_outstanding_series(bal)[0] if s["value"]}
    periods = sorted(p for p in rev if p in sh)
    return _trend_growth_pct([rev[p] / sh[p] for p in periods], 3)


# --- DCF fair value ---------------------------------------------------------
_DCF_YEARS = 10          # explicit-forecast horizon (years)
_DCF_TERMINAL_G = 2.5    # % — Gordon terminal growth (~long-run nominal GDP)
_DCF_G_MIN = 0.0         # % — stage-1 growth clamp floor
_DCF_G_MAX = 20.0        # % — stage-1 growth clamp ceiling
_DCF_WACC_MARGIN = 0.5   # pct-points WACC must exceed terminal growth by


def _fcf_cagr(cf):
    """Annualized Free Cash Flow growth (%) across the cash-flow statement
    history (~4 annual periods on Yahoo's free feed). None when fewer than two
    usable periods exist or either endpoint is non-positive (a growth rate then
    has no meaningful sign) — mirrors _bvps_growth."""
    series = [s for s in _series_from_stmt(cf, "Free Cash Flow") if s["value"]]
    if len(series) < 2:
        return None
    first, last = series[0]["value"], series[-1]["value"]
    if first <= 0 or last <= 0:
        return None
    n = len(series) - 1              # annual statements ≈ one year apart
    return ((last / first) ** (1 / n) - 1) * 100


def _dcf_detail(fcf, g0, wacc, total_debt, total_cash):
    """Two-stage FCFF DCF with the full year-by-year breakdown, in the
    statement (reporting) currency. Stage 1: `fcf` grown for _DCF_YEARS years
    with the growth rate fading linearly from g0 (clamped to [_DCF_G_MIN,
    _DCF_G_MAX]; g0=None falls back to the terminal rate — a flat conservative
    projection) down to _DCF_TERMINAL_G, each year discounted at `wacc`. Then
    a Gordon terminal value. Fading (rather than holding g0 then cliffing to
    terminal) reflects growth mean-reversion and keeps the result from hinging
    on the terminal value. g0/wacc in percentage points.

    Returns a dict carrying every intermediate step (per-year growth/flow/
    discount factor/PV, terminal value, EV -> equity bridge) so the DCF Excel
    export can show its work; _dcf_equity_value wraps this for callers that
    only want the final number.

    None when fcf is missing/non-positive, wacc is missing, or wacc doesn't
    clear the terminal rate by _DCF_WACC_MARGIN (the Gordon denominator
    would be ~0 and the output absurd rather than merely rough)."""
    if fcf is None or fcf <= 0 or wacc is None:
        return None
    if wacc <= _DCF_TERMINAL_G + _DCF_WACC_MARGIN:
        return None
    g0 = _DCF_TERMINAL_G if g0 is None else min(max(g0, _DCF_G_MIN), _DCF_G_MAX)
    w, gt, g0f = wacc / 100.0, _DCF_TERMINAL_G / 100.0, g0 / 100.0
    years, pv, flow = [], 0.0, fcf
    for t in range(1, _DCF_YEARS + 1):
        g = g0f + (gt - g0f) * (t - 1) / (_DCF_YEARS - 1)  # year 1 = g0 … year N = gt
        flow *= 1 + g
        df = 1 / (1 + w) ** t
        years.append({"year": t, "growth": g * 100, "fcf": flow,
                      "discount_factor": df, "pv": flow * df})
        pv += flow * df
    tv = flow * (1 + gt) / (w - gt)
    tv_df = 1 / (1 + w) ** _DCF_YEARS
    ev = pv + tv * tv_df
    return {
        "g0_used": g0,                 # % — after clamping / fallback
        "years": years,
        "stage1_pv": pv,
        "terminal_value": tv, "terminal_df": tv_df, "terminal_pv": tv * tv_df,
        "enterprise_value": ev,
        "equity_value": ev - (total_debt or 0) + (total_cash or 0),
    }


def _dcf_equity_value(fcf, g0, wacc, total_debt, total_cash):
    """Two-stage FCFF DCF -> equity value, in the statement (reporting)
    currency. Thin wrapper over _dcf_detail — see there for the methodology
    and the None conditions."""
    d = _dcf_detail(fcf, g0, wacc, total_debt, total_cash)
    return None if d is None else d["equity_value"]


def _yoy(series):
    """Year-over-year growth (%) keyed by period, from an oldest->newest series
    of {period, value}. A period only appears when the prior year is positive."""
    out = {}
    for i in range(1, len(series)):
        prev = series[i - 1].get("value")
        cur = series[i].get("value")
        per = series[i].get("period")
        if prev is not None and cur is not None and prev > 0:
            out[per] = (cur / prev - 1.0) * 100.0
    return out


def _altman_ratios(income, balance):
    """(X1 working-capital/TA, X2 retained-earnings/TA, X3 EBIT/TA, X5
    sales/TA, total liabilities) — the four terms both Altman variants share,
    plus the denominator each one's X4 needs. None when any input is missing."""
    ta = _stmt_val(balance, "Total Assets")
    ca = _stmt_val(balance, "Current Assets", "Total Current Assets")
    cl = _stmt_val(balance, "Current Liabilities", "Total Current Liabilities")
    tl = _stmt_val(balance, "Total Liabilities Net Minority Interest", "Total Liabilities")
    re = _stmt_val(balance, "Retained Earnings")
    rev = _stmt_val(income, "Total Revenue")
    ebit = _stmt_val(income, "EBIT", "Operating Income", "Operating Income Or Loss")
    if not ta or not tl or None in (ca, cl, re, rev, ebit):
        return None
    return (ca - cl) / ta, re / ta, ebit / ta, rev / ta, tl


def altman_z(income, balance, market_cap):
    """Altman Z-Score (public-company form). Higher = safer: >2.99 'safe',
    1.81-2.99 'grey', <1.81 'distress'. None when inputs are unavailable.
        Z = 1.2*WC/TA + 1.4*RE/TA + 3.3*EBIT/TA + 0.6*MktCap/TL + 1.0*Sales/TA
    """
    r = _altman_ratios(income, balance)
    if r is None or not market_cap:
        return None
    x1, x2, x3, x5, tl = r
    return _num(1.2 * x1 + 1.4 * x2 + 3.3 * x3 + 0.6 * market_cap / tl + 1.0 * x5)


def altman_z_prime(income, balance):
    """Altman Z'-Score (private-firm form): book equity replaces market cap in
    the X4 term and every weight is refitted. Higher = safer, on its own bands
    (they are not Z's): >2.9 'safe', 1.23-2.9 'grey', <1.23 'distress'.
        Z' = 0.717*WC/TA + 0.847*RE/TA + 3.107*EBIT/TA + 0.420*Equity/TL
             + 0.998*Sales/TA

    Carries no price term, so unlike altman_z it cannot fall merely because the
    stock got cheaper. That is why S3 grades on this variant: it is a value
    strategy, and the classic Z would take points back in Pillar C for exactly
    the discount Pillars A and B award. S1 keeps the classic Z on purpose — for
    a health screen a collapsing market cap is real distress information rather
    than circular reasoning.

    Negative equity flows straight through as a negative X4 term, which is the
    intended reading: it is the most levered state a balance sheet can be in.
    """
    r = _altman_ratios(income, balance)
    eq = _stmt_val(balance, "Stockholders Equity", "Total Stockholder Equity",
                   "Common Stock Equity")
    if r is None or eq is None:
        return None
    x1, x2, x3, x5, tl = r
    return _num(0.717 * x1 + 0.847 * x2 + 3.107 * x3 + 0.420 * eq / tl + 0.998 * x5)


def piotroski_f(income, balance, cashflow):
    """Piotroski F-Score (0-9): higher = stronger fundamentals, scored across 9
    profitability, leverage/liquidity and efficiency tests (1 point each).
    Needs a prior year; None when current/prior core figures are missing."""
    def cur(df, *l):
        return _stmt_val(df, *l, col=0)

    def prev(df, *l):
        return _stmt_val(df, *l, col=1)

    ni = cur(income, "Net Income", "Net Income Common Stockholders")
    ni_p = prev(income, "Net Income", "Net Income Common Stockholders")
    ta = cur(balance, "Total Assets")
    ta_p = prev(balance, "Total Assets")
    if None in (ni, ni_p, ta, ta_p) or not ta or not ta_p:
        return None

    cfo = cur(cashflow, "Operating Cash Flow",
              "Cash Flow From Continuing Operating Activities",
              "Total Cash From Operating Activities")
    ca, ca_p = (cur(balance, "Current Assets", "Total Current Assets"),
                prev(balance, "Current Assets", "Total Current Assets"))
    cl, cl_p = (cur(balance, "Current Liabilities", "Total Current Liabilities"),
                prev(balance, "Current Liabilities", "Total Current Liabilities"))
    ltd, ltd_p = (cur(balance, "Long Term Debt", "Long Term Debt And Capital Lease Obligation"),
                  prev(balance, "Long Term Debt", "Long Term Debt And Capital Lease Obligation"))
    gp, gp_p = cur(income, "Gross Profit"), prev(income, "Gross Profit")
    rev, rev_p = cur(income, "Total Revenue"), prev(income, "Total Revenue")
    sh = cur(income, "Diluted Average Shares", "Basic Average Shares")
    sh_p = prev(income, "Diluted Average Shares", "Basic Average Shares")
    if sh is None:
        # No average-share row: fall back to the balance-sheet count. Outstanding
        # only — issued shares would hold flat through a buyback and mask the
        # very share-count drop this test is looking for.
        bal_sh = {s["period"]: s["value"] for s in _shares_outstanding_series(balance)[0]}
        bal_periods = sorted(p for p, v in bal_sh.items() if v is not None)
        sh = bal_sh[bal_periods[-1]] if bal_periods else None
        sh_p = bal_sh[bal_periods[-2]] if len(bal_periods) > 1 else None

    score = 0
    # Profitability
    if ni > 0:
        score += 1
    if cfo is not None and cfo > 0:
        score += 1
    if (ni / ta) > (ni_p / ta_p):                       # ROA improving
        score += 1
    if cfo is not None and cfo > ni:                    # earnings backed by cash
        score += 1
    # Leverage, liquidity, dilution
    if None not in (ltd, ltd_p) and (ltd / ta) < (ltd_p / ta_p):
        score += 1
    if None not in (ca, cl, ca_p, cl_p) and cl and cl_p and (ca / cl) > (ca_p / cl_p):
        score += 1
    if None not in (sh, sh_p) and sh <= sh_p:           # no share dilution
        score += 1
    # Operating efficiency
    if None not in (gp, gp_p, rev, rev_p) and rev and rev_p and (gp / rev) > (gp_p / rev_p):
        score += 1
    if None not in (rev, rev_p) and (rev / ta) > (rev_p / ta_p):
        score += 1
    return score


def _dividend_yields(divs, prices):
    """{calendar_year: dividend yield %} = calendar-year dividends / year-end
    close. Empty dict when dividends or prices are unavailable."""
    out = {}
    if divs is None or len(divs) == 0 or prices is None or len(prices) == 0:
        return out
    try:
        ann = divs.groupby(divs.index.year).sum()
        yc = {int(y): float(p)
              for y, p in prices.groupby(prices.index.year).last().items()}
    except Exception:
        return out
    for y, d in ann.items():
        px = yc.get(int(y))
        if px and d:
            out[int(y)] = float(d) / px * 100.0
    return out


def share_dilution(income, balance, cashflow, ticker):
    """Last-5y share-count breakdown (outstanding / treasury) plus the per-year
    EPS, dividend-yield and payout-ratio lines.

    Returns (rows, basis) — see _shares_outstanding_series. Only basis "issued"
    is a caveat worth showing the user; "derived" is an exact reconstruction and
    reads as ordinary outstanding shares."""
    series, basis = _shares_outstanding_series(balance)
    shares = [s for s in series if s["value"] is not None]
    if not shares:
        return [], basis
    treasury = {t["period"]: t["value"]
                for t in _series_from_stmt(balance, "Treasury Shares Number")}
    ni = {n["period"]: n["value"]
          for n in _series_from_stmt(income, "Net Income", "Net Income Common Stockholders")}
    eps = {e["period"]: e["value"]
           for e in _series_from_stmt(income, "Diluted EPS", "Basic EPS")}
    dp = {d["period"]: d["value"]
          for d in _series_from_stmt(cashflow, "Cash Dividends Paid",
                                     "Common Stock Dividend Paid", "Common Stock Dividends Paid")}
    yields = _dividend_yields(get_dividends(ticker), get_raw_close(ticker))

    sel = shares[-5:]
    rows = []
    for s in sel:
        p = s["period"]
        yr = int(p[:4]) if p[:4].isdigit() else None
        net, paid = ni.get(p), dp.get(p)
        payout = (abs(paid) / net * 100) if (paid and net and net > 0) else None
        rows.append({
            "period": p[:4],
            "shares_outstanding": _num(s["value"]),
            "treasury_shares": _num(treasury.get(p)),
            "eps": _num(eps.get(p)),
            "div_yield": _num(yields.get(yr)) if yr is not None else None,
            "payout_ratio": _num(payout),
        })
    return rows, basis


def _strategy_panel(srow):
    """Deep-dive 'Strategy Ratings' panel rows from a graded screener row.
    Falls back to N/A strings when the row is an error row (no grades)."""
    def fmt(score, verdict):
        if score is None:
            return verdict or None
        return f"{int(round(score))} / 100 — {verdict}" if verdict else f"{int(round(score))} / 100"
    smin = srow.get("strategy_min")
    return {
        "S1 · Triage": fmt(srow.get("strategy_1"), srow.get("strategy_1_verdict")),
        "S1 · Flags": srow.get("strategy_1_flags") or None,
        "S2 · Compounder": fmt(srow.get("strategy_2"), srow.get("strategy_2_verdict")),
        "S3 · Defensive Value": fmt(srow.get("strategy_3"), srow.get("strategy_3_verdict")),
        "Min · All Strategies": None if smin is None else f"{int(round(smin))} / 100",
    }


def deepdive(ticker):
    def produce():
        info = get_info(ticker)
        # Routed through the shared statement cache (_get_stmt), the same one
        # screener_row and financials() use for the same ticker — so a prior
        # Screener/Watchlist view (or a subsequent Excel export, which builds
        # its statement sheets from financials() right after this function
        # returns) reuses this fetch instead of hitting Yahoo again for
        # data already in hand.
        income = _get_stmt(ticker, "income_stmt")
        q_income = _get_stmt(ticker, "quarterly_income_stmt")
        cashflow = _get_stmt(ticker, "cash_flow")
        balance = _get_stmt(ticker, "balance_sheet")

        price = _num(info.get("currentPrice")) or _num(info.get("regularMarketPrice"))
        div_rate = _num(info.get("dividendRate"))
        div_yield = (div_rate / price * 100) if (div_rate and price) else None
        growth = dividend_growth(ticker)

        # See the matching comment in _screener_row: foreign-reporting tickers
        # (ADRs etc.) trade in one currency but file statements in another.
        mkt_ccy = info.get("currency")
        fin_ccy = info.get("financialCurrency")
        market_cap = _num(info.get("marketCap"))
        market_cap_native = _native_market_cap(market_cap, mkt_ccy, fin_ccy)
        fx_mismatch = bool(fin_ccy and mkt_ccy and fin_ccy != mkt_ccy)

        revenue = _series_from_stmt(income, "Total Revenue", "TotalRevenue")
        gross = _series_from_stmt(income, "Gross Profit", "GrossProfit")
        op_income = _series_from_stmt(income, "Operating Income", "Operating Income Or Loss",
                                      "Total Operating Income As Reported", "OperatingIncome")
        net_income = _series_from_stmt(income, "Net Income", "NetIncome",
                                       "Net Income Common Stockholders")
        fcf = _series_from_stmt(cashflow, "Free Cash Flow", "FreeCashFlow")
        # align all series on the income-statement periods
        periods = ([r["period"] for r in revenue]
                   or [n["period"] for n in net_income]
                   or [g["period"] for g in gross])
        rev_map = {r["period"]: r["value"] for r in revenue}
        gp_map = {g["period"]: g["value"] for g in gross}
        oi_map = {o["period"]: o["value"] for o in op_income}
        ni_map = {n["period"]: n["value"] for n in net_income}
        fcf_map = {f["period"]: f["value"] for f in fcf}
        # Fiscal-year share counts, to derive per-year Revenue/Share (revenue ÷
        # shares) — the line overlaid on the revenue chart. Outstanding only: a
        # per-share figure divided by issued shares understates itself by the
        # treasury holding.
        shares_map = {s["period"]: s["value"]
                      for s in _shares_outstanding_series(balance)[0] if s["value"]}

        def _margin(num, den):
            return (num / den * 100) if (num is not None and den) else None

        def _rev_per_share(p):
            rev, sh = rev_map.get(p), shares_map.get(p)
            return (rev / sh) if (rev is not None and sh) else None

        rev_ni = [
            {"period": p[:4], "revenue": rev_map.get(p), "gross_profit": gp_map.get(p),
             "operating_income": oi_map.get(p),
             "net_income": ni_map.get(p), "fcf": fcf_map.get(p),
             "revenue_per_share": _rev_per_share(p),
             "gross_margin": _margin(gp_map.get(p), rev_map.get(p)),
             "operating_margin": _margin(oi_map.get(p), rev_map.get(p)),
             "net_margin": _margin(ni_map.get(p), rev_map.get(p))}
            for p in periods[-5:]   # 5-year scope
        ]

        # FCF dividend coverage = free cash flow / dividends paid (latest year)
        fcf_latest = _stmt_val(cashflow, "Free Cash Flow") or _num(info.get("freeCashflow"))
        ocf_latest = _stmt_val(cashflow, "Operating Cash Flow") or _num(info.get("operatingCashflow"))
        # Capital Expenditure is reported as a negative (a cash outflow); Yahoo's
        # own `capitalExpenditures` info field is unreliable (frequently absent),
        # so fall back to deriving it from the two figures already computed
        # above: FCF = OCF + Capex (capex negative) => Capex = FCF − OCF.
        capex_latest = _stmt_val(cashflow, "Capital Expenditure")
        if capex_latest is None and ocf_latest is not None and fcf_latest is not None:
            capex_latest = fcf_latest - ocf_latest
        div_paid = _stmt_val(cashflow, "Cash Dividends Paid", "Common Stock Dividend Paid",
                             "Common Stock Dividends Paid")
        fcf_coverage = (fcf_latest / abs(div_paid)) if (fcf_latest is not None and div_paid) else None
        total_equity = _stmt_val(balance, "Stockholders Equity", "Total Stockholder Equity",
                                 "Common Stock Equity")

        # Profitability ratios that need statement figures ------------------
        ebit_latest = _stmt_val(income, "EBIT", "Operating Income", "Operating Income Or Loss")
        pretax = _stmt_val(income, "Pretax Income", "Income Before Tax")
        tax = _stmt_val(income, "Tax Provision", "Income Tax Expense")
        total_debt = _num(info.get("totalDebt")) or _stmt_val(balance, "Total Debt")
        total_assets = _stmt_val(balance, "Total Assets")
        current_liab = _stmt_val(balance, "Current Liabilities", "Total Current Liabilities")

        tax_rate = 0.21
        if pretax and tax and pretax > 0:
            tax_rate = min(max(tax / pretax, 0.0), 0.5)

        roic = None  # NOPAT / (debt + equity)
        if ebit_latest and total_equity and total_debt is not None:
            invested = (total_debt or 0) + total_equity
            if invested:
                roic = ebit_latest * (1 - tax_rate) / invested * 100

        roce = None  # EBIT / (total assets - current liabilities)
        if ebit_latest and total_assets and current_liab is not None:
            capital_employed = total_assets - current_liab
            if capital_employed:
                roce = ebit_latest / capital_employed * 100

        interest_exp = (
            _stmt_val(income, "Interest Expense", "Interest Expense Non Operating") or
            _stmt_val(income, "Interest Expense", "Interest Expense Non Operating", col=1)
        )
        wacc = _compute_wacc(_num(info.get("beta")), market_cap_native, total_debt,
                             interest_exp, tax_rate, get_risk_free_rate())

        # Debt/Equity from the same Total Debt and Total Equity shown in the
        # panel, so the three figures reconcile. Yahoo's debtToEquity (computed
        # from the most-recent quarter) is shown alongside as "Debt/Equity (MRQ)".
        debt_eq = (total_debt / total_equity * 100) \
            if (total_debt and total_equity) else None

        ebitda_margin = _num(info.get("ebitdaMargins") and info["ebitdaMargins"] * 100)
        # Revenue and Operating Income, trailing-twelve-month — matching the
        # rest of the Profitability panel (margins, Net Income, Revenue/Share
        # are all Yahoo TTM figures). Yahoo doesn't expose a plain "operating
        # income TTM" field, so it's derived from the same revenue and
        # operatingMargins already backing the "Operating Margin %" row above,
        # keeping the two internally consistent (margin × revenue = income).
        revenue_ttm = _num(info.get("totalRevenue"))
        op_margin_frac = info.get("operatingMargins")
        operating_income_ttm = (_num(revenue_ttm * op_margin_frac)
                                if (revenue_ttm is not None and op_margin_frac is not None) else None)

        # Leverage / cash-conversion ratios built on EBITDA.
        ebitda = _num(info.get("ebitda"))
        debt_ebitda = (total_debt / ebitda) if (total_debt and ebitda and ebitda > 0) else None
        ebitda_fcf = (ebitda / fcf_latest) if (ebitda and fcf_latest) else None

        # Valuation / leverage / dividend metrics also shown in the screener (market_cap already set above).
        # market_cap_native (not market_cap) — see the matching comment in
        # _screener_row on why a mismatched-currency ticker needs this.
        total_cash = _num(info.get("totalCash"))
        p_c = (market_cap_native / total_cash) if (market_cap_native and total_cash) else None
        p_fcf = (market_cap_native / fcf_latest) if (market_cap_native and fcf_latest and fcf_latest > 0) else None
        lt_debt = _stmt_val(balance, "Long Term Debt",
                            "Long Term Debt And Capital Lease Obligation")
        lt_debt_eq = (lt_debt / total_equity * 100) if (lt_debt and total_equity) else None
        # Enterprise Value, EV/EBITDA and Price/Sales all combine market cap
        # with statement figures — Yahoo's own passthrough fields for these
        # get the arithmetic wrong for a mismatched-currency ticker.
        if fx_mismatch and market_cap_native is not None:
            enterprise_value = market_cap_native + (total_debt or 0) - (total_cash or 0)
            ev_ebitda = (enterprise_value / ebitda) if ebitda else None
            ps = (market_cap_native / revenue_ttm) if revenue_ttm else None
        else:
            enterprise_value = _num(info.get("enterpriseValue"))
            ev_ebitda = _num(info.get("enterpriseToEbitda"))
            ps = _num(info.get("priceToSalesTrailing12Months"))
        years_div_inc = consecutive_div_increases(get_dividends(ticker))

        # Year-over-year growth per period (bars for the last 5 years) --------
        eps_series = _series_from_stmt(income, "Diluted EPS", "Basic EPS")
        ebitda_series = _series_from_stmt(income, "EBITDA", "Normalized EBITDA")
        ebitda_map = {e["period"]: e["value"] for e in ebitda_series}
        rev_g = _yoy(revenue)
        eps_g = _yoy(eps_series)
        ebitda_g = _yoy(ebitda_series)
        growth_periods = sorted(set(rev_g) | set(eps_g) | set(ebitda_g))
        growth_rows = [
            {"period": p[:4],
             "revenue_growth": _num(rev_g.get(p)),
             "eps_growth": _num(eps_g.get(p)),
             "ebitda_growth": _num(ebitda_g.get(p)),
             # EBITDA margin per fiscal year, drawn as a line on the growth panel
             "ebitda_margin": _margin(ebitda_map.get(p), rev_map.get(p))}
            for p in growth_periods[-5:]
        ]

        dilution_rows, shares_basis = share_dilution(income, balance, cashflow, ticker)

        # Historical ROIC and cost of capital (last 5 fiscal years) — for the
        # "ROIC vs Cost of Capital" chart. Both are reconstructed per fiscal
        # year from that year's own statement figures. For WACC, every input
        # except beta is genuinely historical: the risk-free rate is read from
        # the 10Y Treasury as of each year-end, the capital-structure weights
        # from that year's market cap (period-end price × shares) and total
        # debt, the cost of debt from that year's interest expense ÷ debt, and
        # the tax rate from that year's pretax/tax. Beta is the one input Yahoo
        # exposes only as a single current value (a true historical beta would
        # need a per-period market regression), so it is held constant — it is
        # the least time-variable WACC input, and freezing it still lets the
        # line track the drivers that actually move (rates, leverage).
        ebit_series = _series_from_stmt(income, "EBIT", "Operating Income", "Operating Income Or Loss")
        pretax_series = _series_from_stmt(income, "Pretax Income", "Income Before Tax")
        taxprov_series = _series_from_stmt(income, "Tax Provision", "Income Tax Expense")
        tdebt_series = _series_from_stmt(balance, "Total Debt")
        tequity_series = _series_from_stmt(balance, "Stockholders Equity",
                                           "Total Stockholder Equity", "Common Stock Equity")
        intexp_series = _series_from_stmt(income, "Interest Expense",
                                          "Interest Expense Non Operating")
        ebit_map = {s["period"]: s["value"] for s in ebit_series}
        pretax_map = {s["period"]: s["value"] for s in pretax_series}
        taxprov_map = {s["period"]: s["value"] for s in taxprov_series}
        tdebt_map = {s["period"]: s["value"] for s in tdebt_series}
        tequity_map = {s["period"]: s["value"] for s in tequity_series}
        intexp_map = {s["period"]: s["value"] for s in intexp_series}
        # The "current" ROIC above prefers info.totalDebt (Yahoo's freshest
        # figure, often a more recent quarter) over the fiscal-year statement
        # value; override just the latest period here too, so the chart's
        # rightmost bar reconciles exactly with the ROIC % shown in this panel
        # instead of silently disagreeing on debt vintage.
        if periods:
            tdebt_map[periods[-1]] = total_debt

        def _period_tax_rate(p):
            pretax_p, tax_p = pretax_map.get(p), taxprov_map.get(p)
            if pretax_p and tax_p and pretax_p > 0:
                return min(max(tax_p / pretax_p, 0.0), 0.5)
            return 0.21

        def _period_roic(p):
            ebit_p, equity_p = ebit_map.get(p), tequity_map.get(p)
            if ebit_p is None or not equity_p:
                return None
            invested_p = (tdebt_map.get(p) or 0) + equity_p
            if not invested_p:
                return None
            return ebit_p * (1 - _period_tax_rate(p)) / invested_p * 100

        rfr_hist = get_risk_free_history()
        raw_close = get_raw_close(ticker)
        cur_beta = _num(info.get("beta"))

        def _period_wacc(p):
            """Per-fiscal-year WACC (%), beta held at today's value; see the
            block comment above. Mirrors _compute_wacc's structure so the
            latest year lines up with the panel's current WACC."""
            rfr_p = _series_asof(rfr_hist, p)
            if cur_beta is None or rfr_p is None:
                return None
            debt_p = tdebt_map.get(p) or 0
            # Historical market cap = period-end price × shares, converted into
            # the statement currency (via the current FX rate) so it weighs
            # against same-currency debt — matters only for mismatched-currency
            # tickers (ADRs), a no-op otherwise. See _native_market_cap.
            price_p = _series_asof(raw_close, p)
            shares_p = shares_map.get(p)
            mc_p = (price_p * shares_p) if (price_p and shares_p) else None
            mc_p = _native_market_cap(mc_p, mkt_ccy, fin_ccy)
            if mc_p is None:
                return None
            total_cap = mc_p + debt_p
            if total_cap <= 0:
                return None
            w_e, w_d = mc_p / total_cap, debt_p / total_cap
            cost_e = rfr_p + cur_beta * _ERP
            ie_p = intexp_map.get(p)
            cost_d = (abs(ie_p) / debt_p * 100) if (ie_p and debt_p > 0) else rfr_p
            return w_e * cost_e + w_d * cost_d * (1 - _period_tax_rate(p))

        sel_periods = periods[-5:]
        wacc_by_period = {p: _period_wacc(p) for p in sel_periods}
        # Reconcile the rightmost point with the panel's current WACC (which
        # uses Yahoo's freshest market cap / rates), mirroring the latest-debt
        # override applied to ROIC just above.
        if sel_periods and wacc is not None:
            wacc_by_period[sel_periods[-1]] = wacc

        roic_history = [{"period": p[:4], "roic": _num(_period_roic(p)),
                         "wacc": _num(wacc_by_period.get(p))} for p in sel_periods]

        # Basic EPS TTM = sum of last 4 quarters; Diluted EPS TTM from yfinance info
        _q_basic = _series_from_stmt(q_income, "Basic EPS")
        _q_basic_vals = [x["value"] for x in _q_basic[-4:] if x.get("value") is not None]
        basic_eps = sum(_q_basic_vals) if len(_q_basic_vals) == 4 else None

        # Cached (screener_row has its own TTL caches on every field it pulls),
        # so computing it here doesn't add a second round-trip to Yahoo.
        srow = screener_row(ticker)

        # REIT-specific metrics panel — only for REITs (industry names them
        # "REIT — …"). Equity REITs own buildings, so FFO (Net Income + property
        # depreciation, less sale gains) is their headline metric and P/FFO their
        # P/E; the panel shows the FFO family. Mortgage REITs own securities, not
        # buildings — FFO adds nothing for them (screener_row leaves it None by
        # type), so the panel shows what actually matters instead: book value per
        # share, P/B, and how well net income covers the (typically over-100%)
        # dividend. Routed on the business type rather than on whether FFO came
        # out non-None, so an mREIT that happens to report a D&A line can't land
        # on the equity branch. See REITs.md.
        reit_panel = reit_kind = None
        if "reit" in (info.get("industry") or "").lower():
            ffo = srow.get("ffo")
            shares = _num(info.get("sharesOutstanding"))
            book_ps = _num(info.get("bookValue"))
            pb = _num(info.get("priceToBook"))
            # Named for the frontend so the panel can say *which* REIT rubric it
            # is showing — the FFO family's absence is a deliberate answer for a
            # mortgage REIT, not missing data, and the title is where a reader
            # finds that out.
            reit_kind = "equity" if (_business_type(info) == "reit"
                                     and ffo is not None) else "mortgage"
            if reit_kind == "equity":
                payout = srow.get("ffo_payout")
                reit_panel = {
                    "FFO": _num(ffo),
                    "FFO/Share": _num(ffo / shares) if (ffo and shares) else None,
                    "P/FFO": srow.get("p_ffo"),
                    "FFO Payout %": _num(payout * 100) if payout is not None else None,
                    "FFO Coverage": srow.get("ffo_coverage"),
                    "Book Value/Share": book_ps,
                    "Price/Book": pb,
                }
            else:                                     # mortgage REIT — book value
                ni = _num(info.get("netIncomeToCommon"))
                reit_panel = {
                    "Book Value/Share": book_ps,
                    "Book Value Trend %": srow.get("bvps_growth"),
                    "Price/Book": pb,
                    "Net Income": ni,
                    "Div Coverage (NI)": _num(ni / abs(div_paid))
                    if (ni is not None and div_paid) else None,
                }

        return {
            "ticker": ticker,
            "name": info.get("shortName") or info.get("longName") or ticker,
            "sector": info.get("sector"),
            "industry": info.get("industry"),
            "currency": info.get("currency"),
            # currency the financial statements are actually reported in — see
            # the matching comment in _screener_row.
            "financial_currency": fin_ccy,
            "exchange": info.get("fullExchangeName") or info.get("exchange"),
            "summary": info.get("longBusinessSummary"),
            "website": info.get("website"),
            # company profile (rendered in the deep view's Company Profile panel)
            "hq": ", ".join(x for x in (info.get("city"), info.get("state"),
                                        info.get("country")) if x) or None,
            "country": info.get("country"),
            "employees": _num(info.get("fullTimeEmployees")),
            "price": price,
            "change_pct": _num(info.get("regularMarketChangePercent")),
            # DCF fair value from the graded screener row — the frontend shows
            # the "⭳ DCF" export button only when this is non-null (financials
            # and REITs are blanked by design; see _screener_row).
            "dcf_value": srow.get("dcf_value"),
            # DCF fair value left in the reporting currency (see _screener_row) —
            # the frontend uses it for the DCF Value hover when the trading and
            # reporting currencies differ.
            "dcf_value_native": srow.get("dcf_value_native"),
            # "equity" | "mortgage" | None — which REIT rubric panels.reit holds.
            "reit_kind": reit_kind,
            "panels": {
                "valuation": {
                    "Market Cap": _num(market_cap),
                    "Enterprise Value": _num(enterprise_value),
                    "Trailing P/E": _num(info.get("trailingPE")),
                    "Forward P/E": _num(info.get("forwardPE")),
                    "PEG Ratio": _num(info.get("trailingPegRatio")) or _num(info.get("pegRatio")),
                    "Price/Book": _num(info.get("priceToBook")),
                    "Price/Sales": _num(ps),
                    "Price/Cash": _num(p_c),
                    "Price/FCF": _num(p_fcf),
                    "EV/EBITDA": _num(ev_ebitda),
                    # DCF fair value + its premium/discount to price, from the
                    # graded row. N/A for financials/REITs (blanked by design).
                    "DCF Value": srow.get("dcf_value"),
                    "DCF Upside %": srow.get("dcf_upside"),
                    "Diluted EPS": _num(info.get("trailingEps")),
                    "Basic EPS": _num(basic_eps),
                },
                "dividend": {
                    "Dividend Rate": _num(div_rate),
                    "Dividend TTM": _num(info.get("trailingAnnualDividendRate")),
                    "Dividend Yield %": _num(div_yield),
                    "Payout Ratio %": _num(info.get("payoutRatio") and info["payoutRatio"] * 100),
                    "FCF Coverage": _num(fcf_coverage),
                    "Ex-Dividend Date": _epoch_to_iso(info.get("exDividendDate")),
                    "5Y Avg Yield %": _num(info.get("fiveYearAvgDividendYield")),
                    "Div Growth 3Y %": growth["cagr_3y"],
                    "Div Growth 5Y %": growth["cagr_5y"],
                    "Years ▲ Dividend": years_div_inc,
                },
                "profitability": {
                    "Revenue": _num(revenue_ttm),
                    "Gross Margin %": _num(info.get("grossMargins") and info["grossMargins"] * 100),
                    "Operating Margin %": _num(info.get("operatingMargins") and info["operatingMargins"] * 100),
                    "EBITDA Margin %": ebitda_margin,
                    "Profit Margin %": _num(info.get("profitMargins") and info["profitMargins"] * 100),
                    "ROE %": _num(info.get("returnOnEquity") and info["returnOnEquity"] * 100),
                    "ROA %": _num(info.get("returnOnAssets") and info["returnOnAssets"] * 100),
                    "ROIC %": _num(roic),
                    "ROCE %": _num(roce),
                    "WACC %": _num(wacc),
                    "ROIC − WACC %": _num(roic - wacc) if (roic is not None and wacc is not None) else None,
                    "Revenue/Share": _num(info.get("revenuePerShare")),
                    "Operating Income": _num(operating_income_ttm),
                    "Net Income": _num(info.get("netIncomeToCommon")),
                },
                "health": {
                    "Total Cash": _num(info.get("totalCash")),
                    "Total Debt": _num(info.get("totalDebt")),
                    "Total Equity": _num(total_equity),
                    "Debt/Equity": _num(debt_eq),
                    "Debt/Equity (MRQ)": _num(info.get("debtToEquity")),
                    "Debt/EBITDA": _num(debt_ebitda),
                    "LT Debt/Equity": _num(lt_debt_eq),
                    "Current Ratio": _num(info.get("currentRatio")),
                    "Quick Ratio": _num(info.get("quickRatio")),
                    "Operating Cash Flow": _num(ocf_latest),
                    "Capital Expenditure": _num(capex_latest),
                    "Free Cash Flow": _num(fcf_latest),
                    "EBITDA": _num(ebitda),
                    "EBITDA/FCF": _num(ebitda_fcf),
                },
                **({"reit": reit_panel} if reit_panel else {}),
                "risk": {
                    "Beta": _num(info.get("beta")),
                    "Short Interest %": _num(info.get("shortPercentOfFloat")
                                             and info["shortPercentOfFloat"] * 100),
                    "Days to Cover": _num(info.get("shortRatio")),
                    "Altman Z-Score": altman_z(income, balance, _num(info.get("marketCap"))),
                    "Piotroski F-Score": piotroski_f(income, balance, cashflow),
                },
                # Strategy grades come off the (cached) screener row so the
                # deep-dive and screener always show identical scores.
                "strategies": _strategy_panel(srow),
            },
            "dividend_growth": growth,
            "revenue_net_income": rev_ni,
            "growth": growth_rows,
            "share_dilution": dilution_rows,
            # "outstanding" | "issued" — which balance-sheet row the share bars
            # came from, so the chart and export can name what they're showing.
            "share_dilution_basis": shares_basis,
            "roic_history": roic_history,
            "wacc_current": _num(wacc),
            # Raw grade fields (same shape as a screener row) so the frontend
            # can drive the Strategy Ratings panel's hover with the identical
            # per-pillar derivation tooltip shown on the Screener/Watchlist
            # grade cells, instead of duplicating that logic server-side.
            "strategy_raw": {k: srow.get(k) for k in (
                "strategy_1", "strategy_1_verdict", "strategy_1_detail",
                "strategy_2", "strategy_2_verdict", "strategy_2_detail",
                "strategy_3", "strategy_3_verdict", "strategy_3_detail",
                "strategy_min",
            )},
        }

    return cached(f"deepdive:{ticker}", 600, produce)


# Price-chart range -> (fetch period, bar interval, visible window in days).
# The fetch period is deliberately longer than the visible window so the
# SMA-200 has enough lookback to be drawn from the very first visible bar
# (200 trading days ≈ 10 calendar months; 200 weeks ≈ 3.8 years). Longer
# ranges switch to weekly/monthly bars so a candle stays readable — a 5y
# range at daily bars would be ~1,250 sub-pixel candles.
_HIST_CFG = {
    "1mo": ("2y", "1d", 31),
    "6mo": ("2y", "1d", 183),
    "1y": ("2y", "1d", 366),
    "2y": ("10y", "1wk", 731),
    "5y": ("max", "1wk", 1827),
    "max": ("max", "1mo", None),
}
_SMA_WINDOWS = (20, 50, 200)


def history(ticker, rng="1y"):
    def produce():
        period, interval, window = _HIST_CFG.get(rng, (rng, "1d", None))
        try:
            df = yf.Ticker(ticker).history(period=period, interval=interval,
                                           auto_adjust=True)
        except Exception:
            return {"ticker": ticker, "points": []}
        if df is None or df.empty:
            return {"ticker": ticker, "points": []}
        # SMAs over the bar interval's closes (i.e. SMA-50 on a weekly chart
        # is a 50-week average), computed on the full fetch before trimming
        # so they are complete across the visible window.
        for w in _SMA_WINDOWS:
            df[f"sma{w}"] = df["Close"].rolling(w).mean()
        if window is not None:
            df = df[df.index >= df.index[-1] - pd.Timedelta(days=window)]
        pts = [
            {"date": idx.strftime("%Y-%m-%d"),
             "open": _num(row.get("Open")), "high": _num(row.get("High")),
             "low": _num(row.get("Low")), "close": _num(row.get("Close")),
             "volume": _num(row.get("Volume")),
             **{f"sma{w}": _num(row.get(f"sma{w}")) for w in _SMA_WINDOWS}}
            for idx, row in df.iterrows()
        ]
        return {"ticker": ticker, "range": rng, "interval": interval, "points": pts}

    return cached(f"hist:{ticker}:{rng}", 600, produce)


_STMT_ATTR = {
    "income": ("income_stmt", "quarterly_income_stmt"),
    "balance": ("balance_sheet", "quarterly_balance_sheet"),
    "cashflow": ("cash_flow", "quarterly_cash_flow"),
}


_QSTMT_ATTR = {
    "income": "quarterly_income_stmt",
    "balance": "quarterly_balance_sheet",
    "cashflow": "quarterly_cash_flow",
}


def _ttm_column(ticker, stmt, labels):
    """Trailing-twelve-month column for a statement table.

    Income statement and cash flow are flows, so TTM = sum of the 4 most-recent
    quarters (only when a full year of quarters exists and the row is present in
    all four). The balance sheet is a point-in-time snapshot, so a sum is
    meaningless — its "trailing" column is the most-recent quarter, labelled MRQ.
    Returns (column_label, {row_label: value}); ("", {}) when unavailable.
    """
    attr = _QSTMT_ATTR.get(stmt, "quarterly_income_stmt")
    # Routed through the shared statement cache (_get_stmt) rather than a
    # direct yf.Ticker fetch: screener_row/deepdive already pull this same
    # quarterly statement for the same ticker, so this is a cache hit far
    # more often than not (notably during Excel export, which otherwise
    # re-fetched every statement from scratch right after screener_row/
    # deepdive had just fetched the identical data moments earlier).
    q = _get_stmt(ticker, attr)
    if q is None or getattr(q, "empty", True):
        return "", {}
    cols = sorted(q.columns, reverse=True)  # newest first
    if stmt == "balance":
        sel, col_label = cols[:1], "MRQ"
    else:
        if len(cols) < 4:
            return "", {}
        sel, col_label = cols[:4], "TTM"
    out = {}
    for label in labels:
        if label not in q.index:
            continue
        vals = [_num(q.loc[label, c]) for c in sel]
        if stmt == "balance":
            out[label] = vals[0]
        elif all(v is not None for v in vals):
            out[label] = sum(vals)
    return col_label, out


def financials(ticker, stmt="income", freq="annual"):
    def produce():
        attr_annual, attr_quarter = _STMT_ATTR.get(stmt, _STMT_ATTR["income"])
        attr = attr_annual if freq == "annual" else attr_quarter
        # Routed through the shared statement cache (_get_stmt) rather than a
        # direct yf.Ticker fetch: screener_row/deepdive already pull this same
        # statement for the same ticker, so this is usually a cache hit — most
        # visibly during Excel export, which previously re-fetched every
        # statement from scratch right after screener_row/deepdive had just
        # fetched the identical data moments earlier in the same request.
        df = _get_stmt(ticker, attr)
        if df is None or df.empty:
            return {"ticker": ticker, "stmt": stmt, "freq": freq, "periods": [], "rows": []}
        cols = sorted(df.columns, reverse=True)  # newest first
        periods = [str(getattr(c, "date", lambda: c)()) for c in cols]
        rows = []
        for label in df.index:
            rows.append({
                "label": str(label),
                "values": [_num(df.loc[label, c]) for c in cols],
            })
        # Prepend a trailing-twelve-month (or MRQ for the balance sheet) column.
        ttm_label, ttm = _ttm_column(ticker, stmt, list(df.index))
        if ttm_label and ttm:
            periods = [ttm_label] + periods
            for row in rows:
                row["values"] = [ttm.get(row["label"])] + row["values"]
        # Surface Yahoo's own TTM EBITDA (the figure behind EBITDA Margin) next to
        # the statement's EBITDA line, so the two are visible side by side. It's a
        # single TTM scalar, so it populates the TTM column only.
        if stmt == "income" and ttm_label == "TTM":
            try:
                y_ebitda = _num(get_info(ticker).get("ebitda"))
            except Exception:
                y_ebitda = None
            if y_ebitda is not None:
                yahoo_row = {"label": "EBITDA (Yahoo TTM)",
                             "values": [y_ebitda] + [None] * (len(periods) - 1)}
                idx = next((i for i, r in enumerate(rows)
                            if "EBITDA" in r["label"]), None)
                rows.insert(idx + 1, yahoo_row) if idx is not None else rows.append(yahoo_row)
        return {"ticker": ticker, "stmt": stmt, "freq": freq,
                "periods": periods, "rows": rows}

    return cached(f"fin:{ticker}:{stmt}:{freq}", 1800, produce)


# --------------------------------------------------------------------------- #
# Calendars — earnings & stock splits                                          #
# --------------------------------------------------------------------------- #
def _ts_to_iso(ts):
    """pandas Timestamp / datetime -> 'YYYY-MM-DD' (None when missing/NaT)."""
    try:
        if ts is None or pd.isna(ts):
            return None
    except (TypeError, ValueError):
        pass
    try:
        return pd.Timestamp(ts).strftime("%Y-%m-%d")
    except Exception:
        try:
            return ts.isoformat()
        except Exception:
            return None


def market_calendar(start=None, end=None, limit=80):
    """Market-wide upcoming earnings + stock splits over a date window."""
    def produce():
        out = {"start": start, "end": end, "earnings": [], "splits": []}
        try:
            cals = yf.Calendars(start=start or None, end=end or None)
        except Exception as e:
            out["error"] = str(e)
            return out

        try:
            edf = cals.get_earnings_calendar(limit=limit)
            if edf is not None and not edf.empty:
                for sym, row in edf.iterrows():
                    out["earnings"].append({
                        "ticker": str(sym),
                        "name": row.get("Company"),
                        "market_cap": _num(row.get("Marketcap")),
                        "event": row.get("Event Name"),
                        "date": _ts_to_iso(row.get("Event Start Date")),
                        "timing": row.get("Timing"),
                        "eps_estimate": _num(row.get("EPS Estimate")),
                        "eps_actual": _num(row.get("Reported EPS")),
                        "surprise_pct": _num(row.get("Surprise(%)")),
                    })
        except Exception as e:
            out["earnings_error"] = str(e)

        try:
            sdf = cals.get_splits_calendar(limit=limit)
            if sdf is not None and not sdf.empty:
                for sym, row in sdf.iterrows():
                    out["splits"].append({
                        "ticker": str(sym),
                        "name": row.get("Company"),
                        "date": _ts_to_iso(row.get("Payable On")),
                        "old_share_worth": _num(row.get("Old Share Worth")),
                        "share_worth": _num(row.get("Share Worth")),
                        "optionable": bool(row.get("Optionable")),
                    })
        except Exception as e:
            out["splits_error"] = str(e)

        return out

    return cached(f"mktcal:{start}:{end}:{limit}", 900, produce)


def stock_calendar(ticker):
    """Per-ticker calendar: next earnings + estimates, ex-dividend / dividend
    dates, past earnings (with surprises) and the full stock-split history."""
    def produce():
        tk = yf.Ticker(ticker)
        out = {
            "ticker": ticker,
            "upcoming": {},
            "earnings_history": [],
            "splits": [],
        }

        try:
            cal = tk.calendar or {}
        except Exception:
            cal = {}
        ed = cal.get("Earnings Date") or []
        out["upcoming"] = {
            "earnings_dates": [_ts_to_iso(d) for d in ed if d is not None],
            "eps_low": _num(cal.get("Earnings Low")),
            "eps_high": _num(cal.get("Earnings High")),
            "eps_avg": _num(cal.get("Earnings Average")),
            "revenue_low": _num(cal.get("Revenue Low")),
            "revenue_high": _num(cal.get("Revenue High")),
            "revenue_avg": _num(cal.get("Revenue Average")),
            "ex_dividend_date": _ts_to_iso(cal.get("Ex-Dividend Date")),
            "dividend_date": _ts_to_iso(cal.get("Dividend Date")),
        }

        # Past earnings with estimate/actual/surprise (needs lxml; degrade ok).
        try:
            df = tk.get_earnings_dates(limit=12)
            if df is not None and not df.empty:
                for idx, row in df.iterrows():
                    out["earnings_history"].append({
                        "date": _ts_to_iso(idx),
                        "eps_estimate": _num(row.get("EPS Estimate")),
                        "eps_actual": _num(row.get("Reported EPS")),
                        "surprise_pct": _num(row.get("Surprise(%)")),
                    })
        except Exception as e:
            out["earnings_history_error"] = str(e)

        # Stock-split history (newest first).
        try:
            sp = tk.splits
            if sp is not None and len(sp):
                rows = [{"date": _ts_to_iso(idx), "ratio": _num(r)}
                        for idx, r in sp.items()]
                out["splits"] = list(reversed(rows))
        except Exception as e:
            out["splits_error"] = str(e)

        return out

    return cached(f"stockcal:{ticker}", 600, produce)


def company_profile(ticker):
    """Wikipedia ethics/controversy context for a ticker, cached for a day —
    the article changes rarely and the lookup costs several API round-trips.
    The cache key embeds the ticker, so ↻ Refresh (clear_ticker_cache) evicts
    it along with everything else for the symbol."""
    def produce():
        info = get_info(ticker)
        name = (info.get("shortName") or info.get("longName") or ticker)
        out = wiki.company_context(name)
        out["ticker"] = ticker
        return out
    return cached(f"wiki:{ticker}", 86400, produce)


# --------------------------------------------------------------------------- #
# Excel export                                                                 #
# --------------------------------------------------------------------------- #
# Metrics-sheet columns, mirroring the screener table's COLS (views.js) key for
# key and in its order, so an exported row reads like the row on screen. The
# screener's star and mini-chart columns have nothing to write to a cell and are
# skipped; DCF Upside, which the screener folds into the DCF Value tooltip
# rather than giving a column, is kept here beside the fair value it qualifies.
# Labels are the unabbreviated forms — a sheet has no narrow-column budget to
# spend, so "Market Cap" over the table's "Mkt Cap".
_METRIC_COLS = [
    ("ticker", "Ticker"),
    ("price", "Price"),
    ("dcf_value", "DCF Value"), ("dcf_upside", "DCF Upside"),
    ("market_cap", "Market Cap"), ("enterprise_value", "Enterprise Value"),
    ("industry", "Industry"),
    # strategy grades (0-100; see the strategy .md docs in this directory)
    ("strategy_1", "S1 Triage"), ("strategy_2", "S2 Compounder"),
    ("strategy_3", "S3 Defensive"), ("strategy_min", "Strat Min"),
    ("strategy_1_flags", "S1 Flags"),
    # valuation
    ("pe", "P/E"), ("forward_pe", "Forward P/E"), ("peg", "PEG"),
    ("ps", "P/S"), ("pb", "P/B"), ("ev_ebitda", "EV/EBITDA"),
    ("p_fcf", "P/FCF"), ("pc", "P/C"),
    ("eps", "Diluted EPS"), ("eps_basic", "Basic EPS"),
    # profitability
    ("profit_margin", "Net Margin"), ("gross_margin", "Gross Margin"),
    ("operating_margin", "Operating Margin"), ("ebitda_margin", "EBITDA Margin"),
    ("roe", "ROE"), ("roa", "ROA"), ("roic", "ROIC"), ("roce", "ROCE"), ("wacc", "WACC"),
    ("roic_wacc", "ROIC−WACC"),
    ("revenue_per_share", "Revenue/Share"), ("income", "Net Income"), ("fcf", "FCF"),
    # financial health
    ("debt_to_equity", "Debt/Eq"), ("debt_to_equity_mrq", "Debt/Eq (MRQ)"),
    ("debt_ebitda", "Debt/EBITDA"), ("lt_debt_to_equity", "LT Debt/Eq"),
    ("current_ratio", "Current Ratio"), ("quick_ratio", "Quick Ratio"),
    ("total_cash", "Total Cash"), ("total_debt", "Total Debt"),
    ("total_equity", "Total Equity"), ("ebitda", "EBITDA"),
    ("ebitda_fcf", "EBITDA/FCF"),
    # dividend
    ("div_yield", "Yield"), ("five_year_avg_yield", "5Y Avg Yield"),
    ("payout_ratio", "Payout Ratio"), ("div_growth_3y", "Div Growth 3Y"),
    ("div_growth_5y", "Div Growth 5Y"), ("dividend_estimate", "Div Estimate"),
    ("dividend_ttm", "Div TTM"), ("fcf_coverage", "FCF Coverage"),
    ("years_div_increase", "Yrs Div Increase"), ("ex_dividend_date", "Ex-Dividend Date"),
    # risk
    ("beta", "Beta"), ("short_interest", "Short Interest"),
    ("days_to_cover", "Days to Cover"), ("altman_z", "Altman Z-Score"),
    ("piotroski_f", "Piotroski F-Score"),
    # performance (price only), YTD out to 10Y
    ("perf_ytd", "Perf YTD"), ("perf_1y", "Perf 1Y"), ("perf_3y", "Perf 3Y"),
    ("perf_5y", "Perf 5Y"), ("perf_10y", "Perf 10Y"),
]

# screener_row keys whose value is in percentage points. Excel exports normalize
# every rate/margin/return to a decimal fraction (value / 100) so a single
# convention holds across both workbooks; the already-fraction keys (margins,
# ROE/ROA, payout ratio, short interest) pass through unchanged.
_PCT_KEYS = {
    "roic", "roce", "wacc", "roic_wacc", "dcf_upside", "div_yield",
    "five_year_avg_yield",
    "div_growth_3y", "div_growth_5y", "debt_to_equity", "debt_to_equity_mrq",
    "lt_debt_to_equity", "perf_ytd", "perf_1y", "perf_3y", "perf_5y", "perf_10y",
}

# Export columns rendered with an Excel percent number format, so they display as
# a percentage (e.g. 41.72% / -25.00%) rather than the stored decimal fraction.
_PCT_FORMAT_KEYS = {"perf_ytd", "perf_1y", "perf_3y", "perf_5y", "perf_10y"}


def _export_val(key, value):
    """Normalize a screener_row value for export: percentage-point fields become
    decimal fractions; everything else passes through."""
    if value is not None and key in _PCT_KEYS:
        return value / 100.0
    return value


def build_workbook(tickers):
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2933")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")

    # Sheet 1: Metrics
    ws = wb.active
    ws.title = "Metrics"
    ws.append([label for _, label in _METRIC_COLS])
    for tk in tickers:
        row = screener_row(tk)
        ws.append([_export_val(key, row.get(key)) for key, _ in _METRIC_COLS])
    style_header(ws, len(_METRIC_COLS))
    ws.freeze_panes = "A2"
    # Show the performance columns as percentages (values are stored as fractions).
    pct_cols = [i + 1 for i, (key, _) in enumerate(_METRIC_COLS) if key in _PCT_FORMAT_KEYS]
    for col in pct_cols:
        for r in range(2, len(tickers) + 2):
            ws.cell(row=r, column=col).number_format = "0.00%"

    # Sheet 2: Price History (aligned by date across tickers)
    ws2 = wb.create_sheet("Price History")
    series = {tk: {p["date"]: p["close"] for p in history(tk, "1y")["points"]}
              for tk in tickers}
    all_dates = sorted({d for s in series.values() for d in s})
    ws2.append(["Date"] + list(tickers))
    for d in all_dates:
        ws2.append([d] + [series[tk].get(d) for tk in tickers])
    style_header(ws2, len(tickers) + 1)
    ws2.freeze_panes = "B2"

    # Sheet 3: Financials (income statement, annual, stacked per ticker)
    ws3 = wb.create_sheet("Financials")
    bold = Font(bold=True)
    r = 1
    for tk in tickers:
        fin = financials(tk, "income", "annual")
        cell = ws3.cell(row=r, column=1, value=f"{tk} — Income Statement (Annual)")
        cell.font = bold
        r += 1
        ws3.cell(row=r, column=1, value="Line Item").font = header_font
        ws3.cell(row=r, column=1).fill = header_fill
        for j, p in enumerate(fin["periods"]):
            c = ws3.cell(row=r, column=2 + j, value=p)
            c.font = header_font
            c.fill = header_fill
        r += 1
        for row in fin["rows"]:
            ws3.cell(row=r, column=1, value=row["label"])
            for j, v in enumerate(row["values"]):
                ws3.cell(row=r, column=2 + j, value=v)
            r += 1
        r += 2  # gap between tickers

    for sheet in wb.worksheets:
        sheet.column_dimensions["A"].width = 34
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Panels rendered on the deep-dive view, in display order, for the export.
# "reit" is present only for REITs (build_deepdive_workbook skips empty panels).
_PANEL_ORDER = [
    ("valuation", "Valuation"), ("profitability", "Profitability"),
    ("health", "Financial Health"), ("dividend", "Dividend"),
    ("reit", "REIT Metrics"), ("risk", "Risk"),
    ("strategies", "Strategy Ratings"),
]

# Deep-dive panel metrics stored in percentage points but without a trailing
# "%" in their label; the export divides them by 100 like the "… %" metrics so
# the workbook is uniformly fraction-based (matching the screener export).
_DD_PCT_LABELS = {"Debt/Equity", "Debt/Equity (MRQ)", "LT Debt/Equity"}


def _dd_metric(label, value):
    """Map an Overview (label, value) pair for export: percentage-point metrics
    become decimal fractions and shed any trailing ' %' from the label."""
    is_pct = label.endswith(" %") or label in _DD_PCT_LABELS
    out_label = label[:-2] if label.endswith(" %") else label
    if is_pct and isinstance(value, (int, float)) and not isinstance(value, bool):
        return out_label, value / 100.0
    return out_label, value


def build_deepdive_workbook(ticker):
    """Single-company workbook mirroring the deep-dive view: an Overview sheet
    with every metric panel, a Charts Data sheet with the revenue/growth/share-
    dilution series, and one sheet per financial statement (annual)."""
    d = deepdive(ticker)
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2933")
    section_fill = PatternFill("solid", fgColor="3E4C59")
    bold = Font(bold=True)

    def header_row(ws, row, labels):
        for j, lab in enumerate(labels, start=1):
            cell = ws.cell(row=row, column=j, value=lab)
            cell.font = header_font
            cell.fill = header_fill

    # ---- Sheet 1: Overview -------------------------------------------------- #
    ws = wb.active
    ws.title = "Overview"
    ws.cell(row=1, column=1, value=f"{d['ticker']} — {d.get('name') or ''}").font = \
        Font(bold=True, size=13)
    chg = d.get("change_pct")
    r = 2
    fin_ccy = d.get("financial_currency")
    overview_rows = [("Sector", d.get("sector")), ("Industry", d.get("industry")),
                      ("Exchange", d.get("exchange")), ("Currency", d.get("currency")),
                      ("Price", d.get("price")),
                      ("Change", chg / 100.0 if isinstance(chg, (int, float)) else chg)]
    if fin_ccy and fin_ccy != d.get("currency"):
        # Foreign-reporting ticker (ADR etc.): price/market cap trade in
        # `Currency`, but revenue, cash, debt, EBITDA, net income, FCF, book
        # value and similar Overview/statement figures are all reported in
        # this currency instead — see the matching comment in _screener_row.
        overview_rows.insert(4, ("Financial Statement Currency", fin_ccy))
    for k, v in overview_rows:
        ws.cell(row=r, column=1, value=k).font = bold
        ws.cell(row=r, column=2, value=v)
        r += 1
    r += 1
    for key, label in _PANEL_ORDER:
        metrics = d["panels"].get(key) or {}
        if not metrics:                     # e.g. "reit" on a non-REIT — no header
            continue
        sc = ws.cell(row=r, column=1, value=label)
        sc.font = header_font
        sc.fill = section_fill
        ws.cell(row=r, column=2).fill = section_fill
        r += 1
        for metric, value in metrics.items():
            out_label, out_value = _dd_metric(metric, value)
            ws.cell(row=r, column=1, value=out_label)
            ws.cell(row=r, column=2, value=out_value)
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    # ---- Sheet 2: Charts Data ----------------------------------------------- #
    ws2 = wb.create_sheet("Charts Data")
    r = 1

    def write_table(title, rows, cols, pct=()):
        nonlocal r
        ws2.cell(row=r, column=1, value=title).font = bold
        r += 1
        header_row(ws2, r, [lab for _, lab in cols])
        r += 1
        for row in rows:
            for j, (k, _) in enumerate(cols, start=1):
                v = row.get(k)
                if k in pct and isinstance(v, (int, float)) and not isinstance(v, bool):
                    v = v / 100.0  # percentage points -> decimal fraction
                ws2.cell(row=r, column=j, value=v)
            r += 1
        r += 2

    write_table("Revenue · Profit · Net Income · FCF", d.get("revenue_net_income", []), [
        ("period", "Year"), ("revenue", "Revenue"), ("gross_profit", "Gross Profit"),
        ("operating_income", "Operating Income"), ("net_income", "Net Income"),
        ("fcf", "FCF"), ("gross_margin", "Gross Margin"),
        ("operating_margin", "Operating Margin"), ("net_margin", "Net Margin")],
        pct={"gross_margin", "operating_margin", "net_margin"})
    write_table("Growth · YoY", d.get("growth", []), [
        ("period", "Year"), ("revenue_growth", "Revenue Growth"),
        ("eps_growth", "EPS Growth"), ("ebitda_growth", "EBITDA Growth"),
        ("ebitda_margin", "EBITDA Margin")],
        pct={"revenue_growth", "eps_growth", "ebitda_growth", "ebitda_margin"})
    # Name the share column for the basis Yahoo actually gave us — "Shares
    # Issued" counts treasury shares, so calling it "Outstanding" would overstate
    # the count (see share_dilution).
    shares_col = ("Shares Issued (incl. treasury)"
                  if d.get("share_dilution_basis") == "issued" else "Shares Outstanding")
    write_table("Share Dilution", d.get("share_dilution", []), [
        ("period", "Year"), ("shares_outstanding", shares_col),
        ("treasury_shares", "Treasury Shares"), ("eps", "Diluted EPS"),
        ("div_yield", "Dividend Yield"), ("payout_ratio", "Payout Ratio")],
        pct={"div_yield", "payout_ratio"})
    ws2.column_dimensions["A"].width = 30

    # ---- Sheets 3-5: Financial statements (annual) -------------------------- #
    for stmt_key, sheet_name in (("income", "Income Statement"),
                                 ("balance", "Balance Sheet"),
                                 ("cashflow", "Cash Flow")):
        fin = financials(ticker, stmt_key, "annual")
        wss = wb.create_sheet(sheet_name)
        header_row(wss, 1, ["Line Item"] + list(fin["periods"]))
        rr = 2
        for row in fin["rows"]:
            wss.cell(row=rr, column=1, value=row["label"])
            for j, v in enumerate(row["values"]):
                wss.cell(row=rr, column=2 + j, value=v)
            rr += 1
        wss.freeze_panes = "B2"
        wss.column_dimensions["A"].width = 34

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_dcf_workbook(ticker):
    """Single-sheet workbook showing every input and intermediate step behind
    the screener's DCF Value for one ticker: the inputs (base FCF, growth,
    WACC and its full CAPM breakdown, debt/cash/shares, currencies), the
    historical FCF the growth rate came from, the 10-year projection table
    (per-year growth, flow, discount factor, PV), the terminal value, and the
    enterprise -> equity -> per-share bridge. When the DCF is N/A for this
    ticker the sheet says why instead. Methodology: METRICS.md §2.

    Raises ValueError with a reason when the row itself can't be built."""
    row = screener_row(ticker)
    if row.get("error"):
        raise ValueError(row["error"])
    info = get_info(ticker)
    cf = _get_stmt(ticker, "cash_flow")
    inc = _get_stmt(ticker, "income_stmt")

    # Reconstruct the same DCF inputs _screener_row used (all fetches cached,
    # so this adds no network round-trips and cannot drift from the row).
    price, market_cap = row.get("price"), row.get("market_cap")
    fcf, wacc = row.get("fcf"), row.get("wacc")
    total_debt, total_cash = row.get("total_debt"), row.get("total_cash")
    mkt_ccy, fin_ccy = row.get("currency"), row.get("financial_currency")
    fx_mismatch = bool(fin_ccy and mkt_ccy and fin_ccy != mkt_ccy)
    fx = _fx_rate(fin_ccy, mkt_ccy) if fx_mismatch else 1.0
    shares = _num(info.get("sharesOutstanding"))
    shares_derived = False
    if not shares and market_cap and price:
        shares, shares_derived = market_cap / price, True
    fcf_hist = [s for s in _series_from_stmt(cf, "Free Cash Flow")
                if s["value"] is not None]
    cagr_raw = _fcf_cagr(cf)
    bt = _business_type({"sector": row.get("sector"), "industry": row.get("industry")})
    detail = (_dcf_detail(fcf, cagr_raw, wacc, total_debt, total_cash)
              if bt not in ("financial", "reit", "mreit") else None)

    # WACC breakdown — same inputs _screener_row feeds _compute_wacc.
    pretax = _stmt_val(inc, "Pretax Income", "Income Before Tax")
    tax = _stmt_val(inc, "Tax Provision", "Income Tax Expense")
    tax_rate = 0.21
    if pretax and tax and pretax > 0:
        tax_rate = min(max(tax / pretax, 0.0), 0.5)
    interest_exp = (_stmt_val(inc, "Interest Expense", "Interest Expense Non Operating") or
                    _stmt_val(inc, "Interest Expense", "Interest Expense Non Operating", col=1))
    mcap_native = _native_market_cap(market_cap, mkt_ccy, fin_ccy)
    wd = _wacc_detail(_num(info.get("beta")), mcap_native,
                      total_debt, interest_exp, tax_rate, get_risk_free_rate())

    # Reason the DCF is N/A (only meaningful when detail is None) — shown in
    # the Result block so a blank sheet still explains itself.
    reason = None
    if row.get("dcf_value") is None:
        if bt in ("financial", "reit", "mreit"):
            reason = (f"Not computed for this business type ({bt}): an FCFF DCF is "
                      "meaningless for balance-sheet businesses and "
                      "depreciation/capex-distorted for property businesses.")
        elif fcf is None or fcf <= 0:
            reason = "Free cash flow is missing or non-positive — no cash stream to discount."
        elif wacc is None:
            reason = "WACC unavailable (beta, market cap, or the risk-free rate is missing)."
        elif wacc <= _DCF_TERMINAL_G + _DCF_WACC_MARGIN:
            reason = (f"WACC ({wacc:.2f}%) does not clear the terminal growth rate "
                      f"({_DCF_TERMINAL_G}%) by {_DCF_WACC_MARGIN} pt — the Gordon "
                      "denominator would be ~0 and the output absurd.")
        elif not shares:
            reason = "Shares outstanding unavailable (and not derivable from market cap ÷ price)."
        else:
            reason = "FX rate unavailable to convert the reporting-currency value."

    wb = Workbook()
    bold = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2933")
    section_fill = PatternFill("solid", fgColor="3E4C59")
    ws = wb.active
    ws.title = "DCF Valuation"
    ws.cell(row=1, column=1,
            value=f"{ticker} — {row.get('name') or ''} — DCF Valuation "
                  f"({_DCF_YEARS}y two-stage FCFF)").font = Font(bold=True, size=13)

    # The whole sheet is a live spreadsheet model: input cells hold values,
    # everything downstream holds Excel *formulas* referencing those cells, so
    # the reader can trace every number and even tweak an assumption and watch
    # the valuation recompute. `A` collects the absolute address ($B$n) of each
    # cell later formulas point at. Rows 3–7 are reserved for the Result block
    # (which references the bridge below); the body is written first from row 9
    # so those addresses exist, then the Result block is back-filled.
    A = {}

    def absref(coord):
        col = "".join(ch for ch in coord if ch.isalpha())
        num = "".join(ch for ch in coord if ch.isdigit())
        return f"${col}${num}"

    def lbl(s):
        """Label text safe to write to a cell: openpyxl stores any string
        starting with '=' as a formula (which then errors), so a leading '='
        becomes a fullwidth '＝'. Defensive — current labels avoid a leading '='."""
        return "＝" + s[1:] if isinstance(s, str) and s.startswith("=") else s

    r = 9

    def section(title):
        nonlocal r
        sc = ws.cell(row=r, column=1, value=title)
        sc.font = header_font
        sc.fill = section_fill
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = section_fill
        r += 1

    def kv(label, value=None, formula=None, pct=False, num=None, bold_row=False):
        """One label/value(-or-formula) row. Returns the value cell's absolute
        address so later formulas can reference it."""
        nonlocal r
        ws.cell(row=r, column=1, value=lbl(label)).font = bold
        c = ws.cell(row=r, column=2)
        c.value = formula if formula is not None else value
        if pct:
            c.number_format = "0.00%"
        elif num:
            c.number_format = num
        if bold_row:
            c.font = bold
        coord = absref(c.coordinate)
        r += 1
        return coord

    CCY = "#,##0"          # currency amounts
    PS = "0.00"            # per-share / price
    FXF = "0.000000"       # fx rate

    # --- Assumptions & inputs (editable values) --------------------------- #
    section("Assumptions & inputs")
    kv("Business type (archetype)", value=bt)
    A["fcf"] = kv(f"Base FCF — latest annual ({fin_ccy or mkt_ccy or 'reporting'})",
                  value=fcf, num=CCY)
    A["terminal"] = kv("Terminal growth", value=_DCF_TERMINAL_G / 100.0, pct=True)
    kv("Projection horizon (years)", value=_DCF_YEARS)
    A["debt"] = kv(f"Total Debt ({fin_ccy or mkt_ccy or 'reporting'})", value=total_debt, num=CCY)
    A["cash"] = kv(f"Total Cash ({fin_ccy or mkt_ccy or 'reporting'})", value=total_cash, num=CCY)
    A["shares"] = kv("Shares Outstanding" + (" (= market cap ÷ price)" if shares_derived else ""),
                     value=shares, num=CCY)
    A["price"] = kv(f"Current Price ({mkt_ccy or 'trading'})", value=price, num=PS)
    # Always present (1.0 when the currencies match) so the bridge's × FX step
    # is uniform whether or not this is a foreign-reporting ticker.
    A["fx"] = kv(f"FX rate (1 {fin_ccy or mkt_ccy or 'reporting'} → {mkt_ccy or 'trading'})",
                 value=fx, num=FXF)
    kv("Trading currency", value=mkt_ccy)
    kv("Reporting currency", value=fin_ccy or mkt_ccy)
    r += 1

    # --- WACC breakdown (the discount rate) ------------------------------- #
    if wd:
        section("WACC breakdown (CAPM) — the discount rate")
        rccy = fin_ccy or mkt_ccy or "reporting"
        A["beta"] = kv("Beta", value=wd["beta"], num="0.0000")
        A["rfr"] = kv("Risk-free rate (10Y Treasury)", value=wd["rfr"] / 100.0, pct=True)
        A["erp"] = kv("Equity risk premium (fixed US estimate)", value=wd["erp"] / 100.0, pct=True)
        A["coste"] = kv("Cost of equity = rfr + β × ERP",
                        formula=f"={A['rfr']}+{A['beta']}*{A['erp']}", pct=True)
        # cost of debt = interest expense ÷ total debt (formula when the
        # interest-expense line exists; otherwise it falls back to the
        # risk-free rate, which has no cell to reference, so it's a value).
        if interest_exp and total_debt and total_debt > 0:
            A["intexp"] = kv(f"Interest expense ({rccy})", value=abs(interest_exp), num=CCY)
            A["costd"] = kv("Cost of debt = interest ÷ total debt",
                            formula=f"={A['intexp']}/{A['debt']}", pct=True)
        else:
            A["costd"] = kv("Cost of debt (= risk-free; no interest-expense data)",
                            value=wd["cost_of_debt"] / 100.0, pct=True)
        # capital-structure weights from market cap (in the reporting currency,
        # so it weighs against same-currency debt) and total debt.
        A["mcap"] = kv(f"Market cap ({rccy}, for weights)", value=mcap_native, num=CCY)
        A["we"] = kv("Equity weight = mktcap ÷ (mktcap + debt)",
                     formula=f"={A['mcap']}/({A['mcap']}+{A['debt']})", pct=True)
        A["wd"] = kv("Debt weight = debt ÷ (mktcap + debt)",
                     formula=f"={A['debt']}/({A['mcap']}+{A['debt']})", pct=True)
        # effective tax rate = tax ÷ pretax, capped to [0, 50%] (formula when
        # both statement lines exist and pretax is positive; else the 21% default).
        if pretax and tax and pretax > 0:
            A["pretax"] = kv(f"Pretax income ({rccy})", value=pretax, num=CCY)
            A["taxprov"] = kv(f"Tax provision ({rccy})", value=tax, num=CCY)
            A["tax"] = kv("Effective tax = min(max(tax ÷ pretax, 0), 50%)",
                          formula=f"=MIN(MAX({A['taxprov']}/{A['pretax']},0),0.5)", pct=True)
        else:
            A["tax"] = kv("Effective tax rate (default 21%)", value=wd["tax_rate"], pct=True)
        A["wacc"] = kv("WACC = wE·costE + wD·costD·(1−tax)",
                       formula=f"={A['we']}*{A['coste']}+{A['wd']}*{A['costd']}*(1-{A['tax']})",
                       pct=True, bold_row=True)
        r += 1

    # --- Historical FCF (source of the growth rate) ----------------------- #
    nz_cells = []            # addresses of non-zero FCF cells (the CAGR endpoints)
    if fcf_hist:
        section("Historical FCF (source of the growth rate)")
        for j, lab in enumerate(("Fiscal year", f"FCF ({fin_ccy or mkt_ccy or ''})",
                                 "YoY growth"), start=1):
            c = ws.cell(row=r, column=j, value=lab)
            c.font = header_font
            c.fill = header_fill
        r += 1
        for i, s in enumerate(fcf_hist):
            ws.cell(row=r, column=1, value=s["period"][:10])
            vc = ws.cell(row=r, column=2, value=s["value"])
            vc.number_format = CCY
            if s["value"]:
                nz_cells.append(absref(vc.coordinate))
            prev = fcf_hist[i - 1]["value"] if i else None
            if prev and prev > 0 and s["value"] is not None:
                yc = ws.cell(row=r, column=3, value=f"=B{r}/B{r - 1}-1")
                yc.number_format = "0.00%"
            r += 1
        r += 1

    # --- Growth rate used (formula: CAGR of the history, then clamped) ---- #
    if detail:
        section("Growth rate used")
        if cagr_raw is not None and len(nz_cells) >= 2:
            n = len(nz_cells) - 1        # compounding periods between endpoints
            A["cagr"] = kv("Historical FCF CAGR (raw)",
                           formula=f"=({nz_cells[-1]}/{nz_cells[0]})^(1/{n})-1", pct=True)
            A["g0"] = kv(f"Stage-1 growth g0 (clamped {_DCF_G_MIN:g}–{_DCF_G_MAX:g}%)",
                         formula=f"=MEDIAN({_DCF_G_MIN / 100:g},{A['cagr']},{_DCF_G_MAX / 100:g})",
                         pct=True, bold_row=True)
        else:
            kv("Historical FCF CAGR (raw)", value="N/A (insufficient history)")
            A["g0"] = kv("Stage-1 growth g0 (flat fallback = terminal growth)",
                         formula=f"={A['terminal']}", pct=True, bold_row=True)
        r += 1

    # --- Projection (all formulas referencing the cells above) ------------ #
    if detail:
        section("Projection (reporting currency)")
        for j, lab in enumerate(("Year", "Growth", "Projected FCF",
                                 "Discount factor", "Present value"), start=1):
            c = ws.cell(row=r, column=j, value=lab)
            c.font = header_font
            c.fill = header_fill
        r += 1
        first_pv_row = r
        for t in range(1, _DCF_YEARS + 1):
            rr = r
            ws.cell(row=rr, column=1, value=t)
            # growth fades linearly: year 1 = g0, year N = terminal
            gc = ws.cell(row=rr, column=2,
                         value=f"={A['g0']}+({A['terminal']}-{A['g0']})*{t - 1}/{_DCF_YEARS - 1}")
            gc.number_format = "0.00%"
            if t == 1:
                fc = ws.cell(row=rr, column=3, value=f"={A['fcf']}*(1+B{rr})")
            else:
                fc = ws.cell(row=rr, column=3, value=f"=C{rr - 1}*(1+B{rr})")
            fc.number_format = CCY
            dc = ws.cell(row=rr, column=4, value=f"=1/(1+{A['wacc']})^{t}")
            dc.number_format = "0.0000"
            pc = ws.cell(row=rr, column=5, value=f"=C{rr}*D{rr}")
            pc.number_format = CCY
            r += 1
        last_fcf_coord = f"C{r - 1}"
        last_pv_row = r - 1
        tr = r
        ws.cell(row=tr, column=1, value="Terminal").font = bold
        tg = ws.cell(row=tr, column=2, value=f"={A['terminal']}")
        tg.number_format = "0.00%"
        tv = ws.cell(row=tr, column=3,
                     value=f"={last_fcf_coord}*(1+{A['terminal']})/({A['wacc']}-{A['terminal']})")
        tv.number_format = CCY
        td = ws.cell(row=tr, column=4, value=f"=1/(1+{A['wacc']})^{_DCF_YEARS}")
        td.number_format = "0.0000"
        tp = ws.cell(row=tr, column=5, value=f"=C{tr}*D{tr}")
        tp.number_format = CCY
        A["tvpv"] = absref(tp.coordinate)
        r += 2

        # --- Valuation bridge --------------------------------------------- #
        # Subtotal rows are bold nouns (no leading "=" — a cell whose text
        # starts with "=" is stored as a formula and errors); the operation
        # rows keep their + − ÷ × prefixes.
        section("Valuation bridge (reporting currency → trading currency)")
        stage1 = kv(f"PV of years 1–{_DCF_YEARS}",
                    formula=f"=SUM(E{first_pv_row}:E{last_pv_row})", num=CCY)
        kv("+ PV of terminal value", formula=f"={A['tvpv']}", num=CCY)
        ev = kv("Enterprise value (DCF)", formula=f"={stage1}+{A['tvpv']}", num=CCY, bold_row=True)
        kv("− Total debt", formula=f"={A['debt']}", num=CCY)
        kv("+ Total cash", formula=f"={A['cash']}", num=CCY)
        equity = kv("Equity value", formula=f"={ev}-{A['debt']}+{A['cash']}", num=CCY, bold_row=True)
        kv("÷ Shares outstanding", formula=f"={A['shares']}", num=CCY)
        pershare = kv(f"Per share ({fin_ccy or mkt_ccy or 'reporting'})",
                      formula=f"={equity}/{A['shares']}", num=PS)
        kv(f"× FX (1 {fin_ccy or mkt_ccy or 'reporting'} → {mkt_ccy or 'trading'})",
           formula=f"={A['fx']}", num=FXF)
        A["dcf"] = kv(f"DCF Value ({mkt_ccy or 'trading'})",
                      formula=f"={pershare}*{A['fx']}", num=PS, bold_row=True)
        kv("Current price", formula=f"={A['price']}", num=PS)
        A["upside"] = kv("Upside vs price", formula=f"={A['dcf']}/{A['price']}-1",
                         pct=True, bold_row=True)
        r += 1

    ws.cell(row=r, column=1,
            value="Every downstream cell is a live formula referencing the inputs above — "
                  "change an assumption and the valuation recomputes. Assumption-heavy by "
                  "construction: growth is extrapolated from a handful of annual statements "
                  "and the discount rate hinges on beta and a fixed 5.5% US ERP. Treat as a "
                  "screen, not a target price. Full methodology: METRICS.md §2.").font = \
        Font(italic=True)

    # --- Result block (rows 3–7), back-filled now the bridge cells exist -- #
    rc = ws.cell(row=3, column=1, value="Result")
    rc.font = header_font
    rc.fill = section_fill
    for c in range(2, 6):
        ws.cell(row=3, column=c).fill = section_fill

    def result_row(rr, label, value=None, formula=None, pct=False, num=None, bold_row=True):
        ws.cell(row=rr, column=1, value=lbl(label)).font = bold
        c = ws.cell(row=rr, column=2)
        c.value = formula if formula is not None else value
        if pct:
            c.number_format = "0.00%"
        elif num:
            c.number_format = num
        if bold_row:
            c.font = bold

    if detail:
        result_row(4, f"DCF Value ({mkt_ccy or 'trading'} / share)", formula=f"={A['dcf']}", num=PS)
        result_row(5, f"Current Price ({mkt_ccy or 'trading'})", formula=f"={A['price']}",
                   num=PS, bold_row=False)
        result_row(6, "DCF Upside vs Price", formula=f"={A['upside']}", pct=True)
    else:
        result_row(4, f"DCF Value ({mkt_ccy or 'trading'} / share)", value=None)
        result_row(5, f"Current Price ({mkt_ccy or 'trading'})", value=price, num=PS, bold_row=False)
        result_row(6, "DCF Upside vs Price", value=None)
        result_row(7, "Why N/A", value=reason, bold_row=False)

    ws.column_dimensions["A"].width = 46
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# HTTP handler                                                                 #
# --------------------------------------------------------------------------- #
def _parse_tickers(qs):
    raw = (qs.get("tickers", [""])[0] or qs.get("ticker", [""])[0] or "")
    out, seen = [], set()
    for t in raw.replace(" ", ",").split(","):
        t = t.strip().upper()
        if t and t not in seen:
            seen.add(t)
            out.append(t)
    return out


class Handler(BaseHTTPRequestHandler):
    server_version = "StockTerminal/1.0"

    def log_message(self, fmt, *args):  # quieter logs
        sys.stderr.write("%s - %s\n" % (self.address_string(), fmt % args))

    # -- helpers ----------------------------------------------------------- #
    def _send_json(self, obj, status=200):
        body = json.dumps(obj).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def _send_bytes(self, data, content_type, filename=None, status=200,
                    no_cache=False):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        if no_cache:
            # Without an explicit Cache-Control, browsers heuristically cache
            # static files and can keep running stale JS/CSS after an update.
            # no-cache = may store, but must revalidate before reuse.
            self.send_header("Cache-Control", "no-cache")
        if filename:
            self.send_header("Content-Disposition",
                             f'attachment; filename="{filename}"')
        self.end_headers()
        self.wfile.write(data)

    def _serve_static(self, path):
        if path in ("/", ""):
            path = "/index.html"
        rel = path.lstrip("/")
        full = os.path.normpath(os.path.join(STATIC_DIR, rel))
        if not full.startswith(STATIC_DIR) or not os.path.isfile(full):
            self.send_error(404, "Not found")
            return
        ctype = {
            ".html": "text/html", ".js": "application/javascript",
            ".css": "text/css", ".svg": "image/svg+xml",
            ".json": "application/json", ".ico": "image/x-icon",
        }.get(os.path.splitext(full)[1], "application/octet-stream")
        with open(full, "rb") as f:
            data = f.read()
        self._send_bytes(data, ctype + ("; charset=utf-8" if ctype.startswith("text") or ctype.endswith("javascript") else ""),
                         no_cache=True)

    def _stream_chat(self, payload):
        """/api/chat: relay the agent's reply as Server-Sent Events. Headers
        go out before the first model token, so failures after that point
        must arrive as in-stream error events (chat_agent never raises)."""
        self.send_response(200)
        self.send_header("Content-Type", "text/event-stream; charset=utf-8")
        self.send_header("Cache-Control", "no-cache")
        self.end_headers()
        try:
            for event in chat_agent.stream_reply(payload.get("messages"),
                                                 payload.get("rows"),
                                                 payload.get("context_label")):
                self.wfile.write(b"data: " + json.dumps(event).encode("utf-8") + b"\n\n")
                self.wfile.flush()
        except (BrokenPipeError, ConnectionResetError):
            pass  # user closed the tab mid-reply

    # -- routing ----------------------------------------------------------- #
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        qs = parse_qs(parsed.query)
        # refresh=1 evicts the server cache for the request's scope first, so
        # the response is rebuilt from fresh Yahoo data.
        refresh = qs.get("refresh", ["0"])[0] == "1"
        try:
            if path == "/api/health":
                return self._send_json({"ok": True})
            if path == "/api/meta":
                # Static reference data (S1 flag legend) for the UI — generated
                # from the grader's own flag catalogue so the two can't drift.
                return self._send_json({"flags": flag_legend()})
            if path == "/api/screener":
                tickers = _parse_tickers(qs)
                if not tickers:
                    return self._send_json({"rows": []})
                if refresh:
                    clear_ticker_cache(tickers)
                with ThreadPoolExecutor(max_workers=min(8, len(tickers))) as ex:
                    rows = list(ex.map(screener_row, tickers))
                return self._send_json({"rows": rows})
            if path == "/api/deepdive":
                t = _parse_tickers(qs)
                if not t:
                    return self._send_json({"error": "ticker required"}, 400)
                if refresh:
                    clear_ticker_cache(t[:1])
                return self._send_json(deepdive(t[0]))
            if path == "/api/history":
                t = _parse_tickers(qs)
                rng = qs.get("range", ["1y"])[0]
                if not t:
                    return self._send_json({"error": "ticker required"}, 400)
                return self._send_json(history(t[0], rng))
            if path == "/api/financials":
                t = _parse_tickers(qs)
                stmt = qs.get("stmt", ["income"])[0]
                freq = qs.get("freq", ["annual"])[0]
                if not t:
                    return self._send_json({"error": "ticker required"}, 400)
                return self._send_json(financials(t[0], stmt, freq))
            if path == "/api/calendar":
                start = qs.get("start", [None])[0] or None
                end = qs.get("end", [None])[0] or None
                try:
                    limit = max(1, min(int(qs.get("limit", ["80"])[0]), 100))
                except ValueError:
                    limit = 80
                if refresh:
                    clear_prefix_cache("mktcal:")
                return self._send_json(market_calendar(start, end, limit))
            if path == "/api/stock_calendar":
                t = _parse_tickers(qs)
                if not t:
                    return self._send_json({"error": "ticker required"}, 400)
                return self._send_json(stock_calendar(t[0]))
            if path == "/api/company_profile":
                t = _parse_tickers(qs)
                if not t:
                    return self._send_json({"error": "ticker required"}, 400)
                if refresh:
                    clear_ticker_cache(t[:1])
                return self._send_json(company_profile(t[0]))
            return self._serve_static(path)
        except BrokenPipeError:
            pass
        except Exception as e:  # never crash the server on one bad request
            self._send_json({"error": str(e)}, 500)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        length = int(self.headers.get("Content-Length", 0) or 0)
        raw = self.rfile.read(length) if length else b"{}"
        try:
            payload = json.loads(raw or b"{}")
        except json.JSONDecodeError:
            payload = {}
        try:
            if path == "/api/export":
                if not _HAS_OPENPYXL:
                    return self._send_json(
                        {"error": "openpyxl not installed. Run: pip install openpyxl"}, 500)
                tickers = [t.strip().upper() for t in payload.get("tickers", []) if t.strip()]
                if not tickers:
                    return self._send_json({"error": "no tickers"}, 400)
                data = build_workbook(tickers)
                stamp = _dt.date.today().strftime("%Y%m%d")
                return self._send_bytes(
                    data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename=f"stock-terminal-{stamp}.xlsx")
            if path == "/api/export_deepdive":
                if not _HAS_OPENPYXL:
                    return self._send_json(
                        {"error": "openpyxl not installed. Run: pip install openpyxl"}, 500)
                raw_t = payload.get("ticker") or (payload.get("tickers") or [None])[0]
                tk = (raw_t or "").strip().upper()
                if not tk:
                    return self._send_json({"error": "ticker required"}, 400)
                data = build_deepdive_workbook(tk)
                stamp = _dt.date.today().strftime("%Y%m%d")
                return self._send_bytes(
                    data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename=f"{tk}-{stamp}.xlsx")
            if path == "/api/export_dcf":
                if not _HAS_OPENPYXL:
                    return self._send_json(
                        {"error": "openpyxl not installed. Run: pip install openpyxl"}, 500)
                raw_t = payload.get("ticker") or (payload.get("tickers") or [None])[0]
                tk = (raw_t or "").strip().upper()
                if not tk:
                    return self._send_json({"error": "ticker required"}, 400)
                try:
                    data = build_dcf_workbook(tk)
                except ValueError as e:   # error row (bad ticker, Yahoo failure)
                    return self._send_json({"error": str(e)}, 400)
                stamp = _dt.date.today().strftime("%Y%m%d")
                return self._send_bytes(
                    data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename=f"{tk}-DCF-{stamp}.xlsx")
            if path == "/api/cache/clear":
                clear_cache()
                return self._send_json({"ok": True})
            if path == "/api/chat":
                return self._stream_chat(payload)
            return self._send_json({"error": "unknown endpoint"}, 404)
        except BrokenPipeError:
            pass
        except Exception as e:
            self._send_json({"error": str(e)}, 500)


def main():
    ap = argparse.ArgumentParser(description="Stock Terminal server")
    ap.add_argument("--port", type=int, default=8765)
    ap.add_argument("--host", default="127.0.0.1")
    args = ap.parse_args()
    httpd = ThreadingHTTPServer((args.host, args.port), Handler)
    url = f"http://{args.host}:{args.port}"
    print(f"\n  📈  Bibes Terminal running at  {url}")
    print(f"      openpyxl export: {'enabled' if _HAS_OPENPYXL else 'DISABLED (pip install openpyxl)'}")
    print("      Press Ctrl+C to stop.\n")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n  Bye.")
        httpd.shutdown()


if __name__ == "__main__":
    main()
