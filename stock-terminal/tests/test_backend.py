"""
Comprehensive tests for stock-terminal/app.py.

Covers:
  - Pure helpers: _num, _epoch_to_iso, _parse_tickers, _stmt_val, _series_from_stmt
  - Business logic: consecutive_div_increases, performance, ROIC, derived ratios
  - TTL cache: hit/miss/expiry/clear, thread safety
  - HTTP API: health, screener, deepdive, history, financials, export (all via a
    real ThreadingHTTPServer started in a thread), 400/404/path-traversal guards
  - screener_row: valid ticker, invalid ticker, missing-field graceful None
  - deepdive: panel keys present, FCF coverage position, Total Equity present
  - financials: period ordering, empty result
  - Excel workbook: correct sheet names, header row, expected column count

All network calls to Yahoo are mocked with unittest.mock to keep tests fast,
deterministic, and offline-safe.
"""
import datetime
import importlib
import io
import json
import math
import os
import sys
import tempfile
import threading
import time
import unittest
from http.server import ThreadingHTTPServer
from unittest.mock import MagicMock, patch
from urllib.request import urlopen
from urllib.error import HTTPError
from urllib.parse import urlencode

import pandas as pd
import numpy as np

# ---------------------------------------------------------------------------
# Make the stock-terminal package importable from the tests/ sub-directory.
# ---------------------------------------------------------------------------
TERMINAL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(TERMINAL_DIR))   # repo root → yfinance importable
sys.path.insert(0, TERMINAL_DIR)                     # stock-terminal/ → app importable

import app  # noqa: E402  (the stock-terminal app module)
import strategies  # noqa: E402  (strategy graders)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _make_df(index_labels, col_dates, data):
    """Build a DataFrame that mimics a yfinance statement.

    yfinance orientation: rows = line items (index), cols = period Timestamps.
    data = {row_label: [val_col0, val_col1, ...]}
    pd.DataFrame(data, index=timestamps).T gives us rows=labels, cols=timestamps.
    """
    cols = [pd.Timestamp(d) for d in col_dates]
    df = pd.DataFrame(data, index=cols).T
    return df


def _make_div_series(year_amounts: dict):
    """Build a yfinance-style dividends Series keyed by monthly Timestamps."""
    idx, vals = [], []
    for year, total in sorted(year_amounts.items()):
        idx.append(pd.Timestamp(f"{year}-06-15"))
        vals.append(total)
    return pd.Series(vals, index=pd.DatetimeIndex(idx))


# ---------------------------------------------------------------------------
# 1. Pure helpers
# ---------------------------------------------------------------------------
class TestNum(unittest.TestCase):

    def test_normal_float(self):
        self.assertAlmostEqual(app._num(3.14), 3.14)

    def test_integer(self):
        self.assertEqual(app._num(42), 42.0)

    def test_none(self):
        self.assertIsNone(app._num(None))

    def test_nan(self):
        self.assertIsNone(app._num(float("nan")))

    def test_pos_inf(self):
        self.assertIsNone(app._num(float("inf")))

    def test_neg_inf(self):
        self.assertIsNone(app._num(float("-inf")))

    def test_string_number(self):
        self.assertAlmostEqual(app._num("2.5"), 2.5)

    def test_invalid_string(self):
        self.assertIsNone(app._num("abc"))

    def test_zero(self):
        self.assertEqual(app._num(0), 0.0)

    def test_negative(self):
        self.assertAlmostEqual(app._num(-100.5), -100.5)

    def test_numpy_nan(self):
        self.assertIsNone(app._num(np.nan))

    def test_large_int(self):
        self.assertEqual(app._num(4_000_000_000_000), 4e12)


class TestEpochToIso(unittest.TestCase):

    def test_known_epoch(self):
        # 2024-05-15 00:00:00 UTC
        epoch = int(datetime.datetime(2024, 5, 15, tzinfo=datetime.timezone.utc).timestamp())
        self.assertEqual(app._epoch_to_iso(epoch), "2024-05-15")

    def test_none(self):
        self.assertIsNone(app._epoch_to_iso(None))

    def test_zero(self):
        # falsy → None
        self.assertIsNone(app._epoch_to_iso(0))

    def test_string_epoch(self):
        epoch = int(datetime.datetime(2023, 1, 1, tzinfo=datetime.timezone.utc).timestamp())
        self.assertEqual(app._epoch_to_iso(str(epoch)), "2023-01-01")

    def test_invalid(self):
        self.assertIsNone(app._epoch_to_iso("not-a-number"))

    def test_float_epoch(self):
        epoch = datetime.datetime(2025, 3, 10, tzinfo=datetime.timezone.utc).timestamp()
        self.assertEqual(app._epoch_to_iso(epoch), "2025-03-10")


class TestParseTickers(unittest.TestCase):

    def test_comma_separated(self):
        qs = {"tickers": ["AAPL,MSFT,NVDA"]}
        self.assertEqual(app._parse_tickers(qs), ["AAPL", "MSFT", "NVDA"])

    def test_space_separated(self):
        qs = {"tickers": ["AAPL MSFT"]}
        self.assertEqual(app._parse_tickers(qs), ["AAPL", "MSFT"])

    def test_lowercased_input(self):
        qs = {"tickers": ["aapl,msft"]}
        self.assertEqual(app._parse_tickers(qs), ["AAPL", "MSFT"])

    def test_deduplication(self):
        qs = {"tickers": ["AAPL,AAPL,MSFT"]}
        self.assertEqual(app._parse_tickers(qs), ["AAPL", "MSFT"])

    def test_empty(self):
        self.assertEqual(app._parse_tickers({}), [])

    def test_ticker_fallback_key(self):
        qs = {"ticker": ["GOOGL"]}
        self.assertEqual(app._parse_tickers(qs), ["GOOGL"])

    def test_whitespace_stripped(self):
        qs = {"tickers": [" TSLA , AMZN "]}
        self.assertEqual(app._parse_tickers(qs), ["TSLA", "AMZN"])


# ---------------------------------------------------------------------------
# 2. Statement helpers
# ---------------------------------------------------------------------------
class TestStmtVal(unittest.TestCase):

    def setUp(self):
        self.df = _make_df(
            ["Total Revenue", "Net Income", "Gross Profit"],
            ["2024-12-31", "2023-12-31", "2022-12-31"],
            {
                "Total Revenue": [400e9, 380e9, 370e9],
                "Net Income":    [100e9,  90e9,  80e9],
                "Gross Profit":  [180e9, 170e9, 160e9],
            },
        )

    def test_first_label_newest_col(self):
        # col=0 → newest = 2024
        v = app._stmt_val(self.df, "Total Revenue")
        self.assertAlmostEqual(v, 400e9)

    def test_fallback_label(self):
        # "Revenue" not in df, but "Net Income" is
        v = app._stmt_val(self.df, "Revenue", "Net Income")
        self.assertAlmostEqual(v, 100e9)

    def test_older_period(self):
        v = app._stmt_val(self.df, "Total Revenue", col=1)
        self.assertAlmostEqual(v, 380e9)

    def test_oldest_period(self):
        v = app._stmt_val(self.df, "Total Revenue", col=2)
        self.assertAlmostEqual(v, 370e9)

    def test_col_out_of_range(self):
        self.assertIsNone(app._stmt_val(self.df, "Total Revenue", col=99))

    def test_label_not_found(self):
        self.assertIsNone(app._stmt_val(self.df, "EBITDA"))

    def test_none_df(self):
        self.assertIsNone(app._stmt_val(None, "Total Revenue"))

    def test_empty_df(self):
        self.assertIsNone(app._stmt_val(pd.DataFrame(), "Total Revenue"))

    def test_nan_value_returns_none(self):
        df = _make_df(["Total Revenue"], ["2024-12-31"], {"Total Revenue": [float("nan")]})
        self.assertIsNone(app._stmt_val(df, "Total Revenue"))


class TestSeriesFromStmt(unittest.TestCase):

    def setUp(self):
        self.df = _make_df(
            ["Total Revenue", "Net Income"],
            ["2022-12-31", "2023-12-31", "2024-12-31"],
            {
                "Total Revenue": [370e9, 380e9, 400e9],
                "Net Income":    [ 80e9,  90e9, 100e9],
            },
        )

    def test_returns_ascending_periods(self):
        series = app._series_from_stmt(self.df, "Total Revenue")
        periods = [s["period"] for s in series]
        self.assertEqual(periods, sorted(periods))

    def test_values_match_df(self):
        series = app._series_from_stmt(self.df, "Total Revenue")
        self.assertEqual(len(series), 3)
        self.assertAlmostEqual(series[-1]["value"], 400e9)

    def test_fallback_label(self):
        series = app._series_from_stmt(self.df, "EBIT", "Net Income")
        self.assertEqual(len(series), 3)
        self.assertAlmostEqual(series[0]["value"], 80e9)

    def test_label_missing_returns_empty(self):
        self.assertEqual(app._series_from_stmt(self.df, "EBITDA"), [])

    def test_none_df(self):
        self.assertEqual(app._series_from_stmt(None, "Total Revenue"), [])

    def test_nan_value_in_series_returns_none(self):
        df = _make_df(["Total Revenue"], ["2024-12-31"], {"Total Revenue": [float("nan")]})
        series = app._series_from_stmt(df, "Total Revenue")
        self.assertIsNone(series[0]["value"])


class TestPositiveYearCount(unittest.TestCase):
    """_positive_year_count — the earnings-stability signal behind S3 Pillar D."""

    def _df(self, values):
        return _make_df(["Net Income"],
                        ["2022-12-31", "2023-12-31", "2024-12-31", "2025-12-31"],
                        {"Net Income": values})

    def test_all_positive(self):
        self.assertEqual(
            app._positive_year_count(self._df([5e9, 6e9, 7e9, 9e9]), "Net Income"),
            (4, 4))

    def test_counts_only_positive_periods(self):
        self.assertEqual(
            app._positive_year_count(self._df([5e9, -1e9, 7e9, 9e9]), "Net Income"),
            (3, 4))

    def test_zero_is_not_positive(self):
        self.assertEqual(
            app._positive_year_count(self._df([5e9, 0.0, 7e9, 9e9]), "Net Income"),
            (3, 4))

    def test_blank_period_counts_neither_way(self):
        # A missing figure shrinks the denominator rather than reading as a loss.
        self.assertEqual(
            app._positive_year_count(
                self._df([5e9, float("nan"), 7e9, 9e9]), "Net Income"),
            (3, 3))

    def test_absent_line_reports_no_history(self):
        # (0, 0) means "no history", which the grader must not read as "never
        # positive" — it falls back to the latest value's sign instead.
        self.assertEqual(
            app._positive_year_count(self._df([5e9] * 4), "Nonexistent"), (0, 0))

    def test_none_df(self):
        self.assertEqual(app._positive_year_count(None, "Net Income"), (0, 0))

    def test_fallback_label(self):
        df = _make_df(["Net Income Common Stockholders"], ["2024-12-31"],
                      {"Net Income Common Stockholders": [5e9]})
        self.assertEqual(
            app._positive_year_count(df, "Net Income",
                                     "Net Income Common Stockholders"), (1, 1))


# ---------------------------------------------------------------------------
# 2b. _yoy — year-over-year growth keyed by period
# ---------------------------------------------------------------------------
class TestYoY(unittest.TestCase):

    @staticmethod
    def _series(*pairs):
        return [{"period": p, "value": v} for p, v in pairs]

    def test_basic_growth(self):
        s = self._series(("2022", 100.0), ("2023", 110.0), ("2024", 121.0))
        out = app._yoy(s)
        # first period has no prior year -> absent
        self.assertNotIn("2022", out)
        self.assertAlmostEqual(out["2023"], 10.0)
        self.assertAlmostEqual(out["2024"], 10.0)

    def test_negative_growth(self):
        # share count shrinking (buybacks) -> negative growth
        s = self._series(("2022", 1000.0), ("2023", 950.0))
        self.assertAlmostEqual(app._yoy(s)["2023"], -5.0)

    def test_skips_when_prior_not_positive(self):
        # prior year zero or negative -> growth undefined for that step
        s = self._series(("2022", 0.0), ("2023", 50.0), ("2024", 60.0))
        out = app._yoy(s)
        self.assertNotIn("2023", out)          # prev == 0 skipped
        self.assertAlmostEqual(out["2024"], 20.0)

    def test_skips_none_values(self):
        s = self._series(("2022", 100.0), ("2023", None), ("2024", 150.0))
        out = app._yoy(s)
        self.assertNotIn("2023", out)          # cur is None
        self.assertNotIn("2024", out)          # prev is None

    def test_empty_and_single(self):
        self.assertEqual(app._yoy([]), {})
        self.assertEqual(app._yoy(self._series(("2024", 100.0))), {})


# ---------------------------------------------------------------------------
# 3. Dividend helpers
# ---------------------------------------------------------------------------
class TestConsecutiveDivIncreases(unittest.TestCase):

    def test_no_dividends(self):
        self.assertIsNone(app.consecutive_div_increases(None))
        self.assertIsNone(app.consecutive_div_increases(pd.Series([], dtype=float)))

    def test_single_year(self):
        divs = _make_div_series({2020: 1.0})
        # Need ≥ 2 completed years to count
        with patch("app._dt") as mock_dt:
            mock_dt.date.today.return_value = datetime.date(2026, 6, 1)
            self.assertIsNone(app.consecutive_div_increases(divs))

    def test_five_consecutive_increases(self):
        divs = _make_div_series({2019: 1.0, 2020: 1.1, 2021: 1.2, 2022: 1.3, 2023: 1.4, 2024: 1.5})
        with patch("app._dt") as mock_dt:
            mock_dt.date.today.return_value = datetime.date(2026, 1, 1)
            count = app.consecutive_div_increases(divs)
        self.assertEqual(count, 5)

    def test_streak_broken(self):
        # 2021 < 2020 → streak resets; 2022, 2023 are increases = streak of 2
        divs = _make_div_series({2019: 1.0, 2020: 1.2, 2021: 1.0, 2022: 1.1, 2023: 1.2})
        with patch("app._dt") as mock_dt:
            mock_dt.date.today.return_value = datetime.date(2026, 1, 1)
            count = app.consecutive_div_increases(divs)
        self.assertEqual(count, 2)

    def test_zero_streak(self):
        # Last year was a cut
        divs = _make_div_series({2022: 1.5, 2023: 1.4, 2024: 1.3})
        with patch("app._dt") as mock_dt:
            mock_dt.date.today.return_value = datetime.date(2026, 1, 1)
            count = app.consecutive_div_increases(divs)
        self.assertEqual(count, 0)

    def test_flat_dividend_not_counted(self):
        # Exactly equal (within 0.01% threshold) → not an increase
        divs = _make_div_series({2022: 1.0, 2023: 1.0, 2024: 1.0})
        with patch("app._dt") as mock_dt:
            mock_dt.date.today.return_value = datetime.date(2026, 1, 1)
            count = app.consecutive_div_increases(divs)
        self.assertEqual(count, 0)

    def test_current_year_excluded(self):
        # 2025 is the current year → should not be counted
        divs = _make_div_series({2023: 1.0, 2024: 1.1, 2025: 1.2})
        with patch("app._dt") as mock_dt:
            mock_dt.date.today.return_value = datetime.date(2025, 6, 1)
            count = app.consecutive_div_increases(divs)
        # Only 2023→2024 increase is countable
        self.assertEqual(count, 1)


# ---------------------------------------------------------------------------
# 4. Performance calculation
# ---------------------------------------------------------------------------
class TestPerformance(unittest.TestCase):

    def _price_series(self, start="2015-01-02", end="2025-06-25", start_val=100.0, end_val=300.0):
        dates = pd.date_range(start, end, freq="B", tz="America/New_York")
        vals = np.linspace(start_val, end_val, len(dates))
        return pd.Series(vals, index=dates)

    def test_all_nones_on_empty(self):
        with patch("app.get_raw_close", return_value=None):
            p = app.performance("FAKE")
        for k in ("ytd", "1y", "3y", "5y", "10y"):
            self.assertIsNone(p[k])

    def test_ytd_positive(self):
        s = self._price_series(start="2024-01-02", end="2025-06-25",
                               start_val=100.0, end_val=130.0)
        with patch("app.get_raw_close", return_value=s):
            p = app.performance("X")
        # Last close is 130, last-of-2024 close is somewhere ~115, so YTD should be positive
        self.assertIsNotNone(p["ytd"])
        self.assertGreater(p["ytd"], 0)

    def test_1y_return(self):
        # Build series where price exactly doubled over 2 years
        dates = pd.date_range("2023-01-02", "2025-01-02", freq="B", tz="UTC")
        vals = np.linspace(100.0, 200.0, len(dates))
        s = pd.Series(vals, index=dates)
        with patch("app.get_raw_close", return_value=s):
            p = app.performance("X")
        # 1Y return: ~100% over 2 years → ~50% in 1 year (linear)
        self.assertIsNotNone(p["1y"])
        self.assertGreater(p["1y"], 0)

    def test_10y_fallback_uses_earliest_close(self):
        # Series spans ~10 years but doesn't reach exactly 10 years before today
        s = self._price_series(start="2016-01-04", end="2025-12-31",
                               start_val=50.0, end_val=200.0)
        with patch("app.get_raw_close", return_value=s):
            p = app.performance("X")
        # span ≈ 9.99y ≥ 10*0.95 → should use first close as fallback
        self.assertIsNotNone(p["10y"])
        self.assertGreater(p["10y"], 0)

    def test_short_series_no_10y(self):
        # Only 2 years of data → 10Y should be None
        s = self._price_series(start="2023-01-02", end="2025-01-02",
                               start_val=100.0, end_val=120.0)
        with patch("app.get_raw_close", return_value=s):
            p = app.performance("X")
        self.assertIsNone(p["10y"])

    def test_pct_calculation(self):
        # 100 → 150 = exactly 50% return
        dates = pd.date_range("2024-01-02", "2025-06-25", freq="B", tz="UTC")
        vals = np.concatenate([np.full(len(dates) - 1, 100.0), [150.0]])
        s = pd.Series(vals, index=dates)
        with patch("app.get_raw_close", return_value=s):
            p = app.performance("X")
        self.assertIsNotNone(p["1y"])
        # Within a tolerance (not exactly 50% since base isn't 100)
        self.assertGreater(p["1y"], 0)


# ---------------------------------------------------------------------------
# 5. Derived-ratio logic (inline, no network)
# ---------------------------------------------------------------------------
class TestDerivedRatios(unittest.TestCase):

    def test_roic_normal(self):
        """ROIC = EBIT*(1-tax) / (debt+equity)*100"""
        ebit, tax_rate, debt, equity = 10e9, 0.21, 20e9, 30e9
        roic = ebit * (1 - tax_rate) / (debt + equity) * 100
        self.assertAlmostEqual(roic, 15.8, places=1)

    def test_roic_zero_invested(self):
        # equity=0, debt=0 → division by zero guard → None
        # Replicate the guard: invested = 0+0=0 → skip
        ebit, equity, total_debt = 10e9, 0, 0
        invested = total_debt + equity
        result = None if not invested else ebit / invested
        self.assertIsNone(result)

    def test_fcf_coverage_above_threshold(self):
        coverage = 10e9 / abs(-5e9)   # FCF / dividends_paid
        self.assertAlmostEqual(coverage, 2.0)
        self.assertGreater(coverage, 1.2)   # green band

    def test_fcf_coverage_yellow_band(self):
        coverage = 9e9 / abs(-9e9)
        self.assertAlmostEqual(coverage, 1.0)
        self.assertGreaterEqual(coverage, 0.8)
        self.assertLess(coverage, 1.2)

    def test_fcf_coverage_red_band(self):
        coverage = 5e9 / abs(-10e9)
        self.assertAlmostEqual(coverage, 0.5)
        self.assertLess(coverage, 0.8)

    def test_fcf_coverage_no_dividend(self):
        # div_paid = 0 or None → coverage = None
        div_paid = 0
        result = (100e9 / abs(div_paid)) if div_paid else None
        self.assertIsNone(result)

    def test_lt_debt_eq(self):
        lt_debt, equity = 30e9, 40e9
        lt_debt_eq = lt_debt / equity * 100
        self.assertAlmostEqual(lt_debt_eq, 75.0)

    def test_lt_debt_eq_negative_equity(self):
        # Negative equity → result is negative (leverage ratio can be negative)
        lt_debt, equity = 30e9, -10e9
        lt_debt_eq = lt_debt / equity * 100
        self.assertLess(lt_debt_eq, 0)

    def test_p_fcf_positive_fcf(self):
        mktcap, fcf = 100e9, 5e9
        self.assertAlmostEqual(mktcap / fcf, 20.0)

    def test_p_fcf_negative_fcf_returns_none(self):
        fcf = -1e9
        result = (100e9 / fcf) if (fcf and fcf > 0) else None
        self.assertIsNone(result)


# ---------------------------------------------------------------------------
# 6. TTL cache
# ---------------------------------------------------------------------------
class TestCache(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def tearDown(self):
        app.clear_cache()

    def test_cache_miss_calls_producer(self):
        calls = []
        def prod():
            calls.append(1)
            return "result"
        app.cached("key1", 60, prod)
        self.assertEqual(len(calls), 1)

    def test_cache_hit_skips_producer(self):
        calls = []
        def prod():
            calls.append(1)
            return "result"
        app.cached("key2", 60, prod)
        app.cached("key2", 60, prod)
        self.assertEqual(len(calls), 1)

    def test_cache_expiry(self):
        calls = []
        def prod():
            calls.append(1)
            return "result"
        app.cached("key3", 0.05, prod)   # 50ms TTL
        time.sleep(0.1)
        app.cached("key3", 0.05, prod)
        self.assertEqual(len(calls), 2)

    def test_clear_cache_forces_refetch(self):
        calls = []
        def prod():
            calls.append(1)
            return "value"
        app.cached("key4", 60, prod)
        app.clear_cache()
        app.cached("key4", 60, prod)
        self.assertEqual(len(calls), 2)

    def test_different_keys_independent(self):
        a_calls, b_calls = [], []
        app.cached("ka", 60, lambda: a_calls.append(1) or "a")
        app.cached("kb", 60, lambda: b_calls.append(1) or "b")
        app.cached("ka", 60, lambda: a_calls.append(1) or "a")
        self.assertEqual(len(a_calls), 1)
        self.assertEqual(len(b_calls), 1)

    def test_cache_returns_value(self):
        val = app.cached("kv", 60, lambda: {"x": 42})
        self.assertEqual(val, {"x": 42})

    def test_cache_thread_safety(self):
        results = []
        def produce():
            time.sleep(0.01)
            return 99
        def worker():
            results.append(app.cached("tkey", 60, produce))
        threads = [threading.Thread(target=worker) for _ in range(10)]
        for t in threads: t.start()
        for t in threads: t.join()
        self.assertEqual(len(results), 10)
        self.assertTrue(all(r == 99 for r in results))

    def test_clear_ticker_cache_evicts_all_ticker_keys(self):
        for key in ("info:AAPL", "deepdive:AAPL", "hist:AAPL:1y",
                    "fin:AAPL:income:annual", "divs:AAPL", "stockcal:AAPL"):
            app.cached(key, 60, lambda: "v")
        app.cached("info:MSFT", 60, lambda: "v")
        app.clear_ticker_cache(["AAPL"])
        with app._CACHE_LOCK:
            keys = set(app._CACHE)
        self.assertEqual(keys, {"info:MSFT"})

    def test_clear_ticker_cache_multiple_tickers(self):
        app.cached("info:AAPL", 60, lambda: "v")
        app.cached("info:MSFT", 60, lambda: "v")
        app.cached("info:NVDA", 60, lambda: "v")
        app.clear_ticker_cache(["aapl", "msft"])   # case-insensitive
        with app._CACHE_LOCK:
            keys = set(app._CACHE)
        self.assertEqual(keys, {"info:NVDA"})

    def test_clear_ticker_cache_spares_unrelated_keys(self):
        app.cached("rfr:^TNX", 60, lambda: 4.2)
        app.cached("mktcal:2026-07-03:2026-07-10:80", 60, lambda: {})
        app.clear_ticker_cache(["AAPL"])
        with app._CACHE_LOCK:
            self.assertEqual(len(app._CACHE), 2)

    def test_clear_prefix_cache(self):
        app.cached("mktcal:a:b:80", 60, lambda: 1)
        app.cached("mktcal:c:d:80", 60, lambda: 2)
        app.cached("info:AAPL", 60, lambda: 3)
        app.clear_prefix_cache("mktcal:")
        with app._CACHE_LOCK:
            keys = set(app._CACHE)
        self.assertEqual(keys, {"info:AAPL"})

    def test_clear_ticker_cache_then_refetch(self):
        calls = []
        app.cached("deepdive:AAPL", 60, lambda: calls.append(1) or "v")
        app.clear_ticker_cache(["AAPL"])
        app.cached("deepdive:AAPL", 60, lambda: calls.append(1) or "v")
        self.assertEqual(len(calls), 2)


# ---------------------------------------------------------------------------
# 7. screener_row (with mocked yfinance)
# ---------------------------------------------------------------------------
_BASE_INFO = {
    "shortName": "Test Corp",
    "currentPrice": 100.0,
    "regularMarketChangePercent": 1.5,
    "marketCap": 1e12,
    "trailingPE": 20.0,
    "forwardPE": 18.0,
    "trailingPegRatio": 1.5,
    "priceToBook": 3.0,
    "priceToSalesTrailing12Months": 5.0,
    "totalCash": 10e9,
    "enterpriseToEbitda": 12.0,
    "trailingEps": 5.0,
    "totalRevenue": 400e9,
    "operatingMargins": 0.20,
    "netIncomeToCommon": 50e9,
    "profitMargins": 0.25,
    "ebitda": 120e9,
    "freeCashflow": 40e9,
    "returnOnAssets": 0.10,
    "returnOnEquity": 0.20,
    "debtToEquity": 50.0,
    "totalDebt": 30e9,
    "dividendRate": 2.0,
    "trailingAnnualDividendRate": 1.9,
    "payoutRatio": 0.40,
    "fiveYearAvgDividendYield": 2.5,
    "sector": "Technology",
    "industry": "Software",
    "currency": "USD",
}

_INCOME_DF = _make_df(
    ["Total Revenue", "Net Income", "Gross Profit", "EBIT", "Pretax Income", "Tax Provision"],
    ["2024-12-31", "2023-12-31"],
    {
        "Total Revenue":  [400e9, 370e9],
        "Net Income":     [100e9,  90e9],
        "Gross Profit":   [180e9, 165e9],
        "EBIT":           [ 80e9,  75e9],
        "Pretax Income":  [ 78e9,  73e9],
        "Tax Provision":  [ 16e9,  15e9],
    },
)

_BAL_DF = _make_df(
    ["Stockholders Equity", "Long Term Debt", "Total Debt"],
    ["2024-12-31", "2023-12-31"],
    {
        "Stockholders Equity": [50e9, 45e9],
        "Long Term Debt":      [25e9, 22e9],
        "Total Debt":          [30e9, 27e9],
    },
)

_CF_DF = _make_df(
    ["Free Cash Flow", "Operating Cash Flow", "Cash Dividends Paid"],
    ["2024-12-31", "2023-12-31"],
    {
        "Free Cash Flow":      [40e9, 38e9],
        # No "Capital Expenditure" row on purpose — exercises the FCF−OCF
        # derivation fallback (see deepdive()'s capex_latest).
        "Operating Cash Flow": [48e9, 44e9],
        "Cash Dividends Paid": [-8e9, -7e9],
    },
)

_DIVS = _make_div_series({2019: 1.5, 2020: 1.6, 2021: 1.7, 2022: 1.8, 2023: 1.9, 2024: 2.0})

_CLOSE_SERIES = pd.Series(
    np.linspace(50.0, 100.0, 2500),
    index=pd.date_range("2015-06-01", periods=2500, freq="B", tz="UTC"),
)


def _patch_all(ticker="TST"):
    """Return a context-manager stack that mocks all external data calls."""
    import contextlib
    @contextlib.contextmanager
    def ctx():
        app.clear_cache()
        with patch("app.get_info", return_value=_BASE_INFO), \
             patch("app._get_stmt", side_effect=lambda t, attr: {
                 "income_stmt": _INCOME_DF, "balance_sheet": _BAL_DF,
                 "cash_flow": _CF_DF,
             }.get(attr, None)), \
             patch("app.get_dividends", return_value=_DIVS), \
             patch("app.get_raw_close", return_value=_CLOSE_SERIES), \
             patch("app.performance", return_value={
                 "ytd": 5.0, "1y": 12.0, "3y": 30.0, "5y": 60.0, "10y": 100.0}):
            yield
    return ctx()


class TestScreenerRow(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def test_required_keys_present(self):
        with _patch_all():
            row = app.screener_row("TST")
        required = ["ticker", "name", "price", "market_cap", "pe", "forward_pe", "peg",
                    "pb", "ps", "pc", "p_fcf", "ev_ebitda", "dcf_value", "dcf_upside",
                    "eps",
                    "income", "profit_margin", "fcf", "roa", "roe", "roic",
                    "debt_to_equity", "lt_debt_to_equity",
                    "div_yield", "five_year_avg_yield", "dividend_estimate", "dividend_ttm",
                    "payout_ratio", "fcf_coverage", "years_div_increase",
                    "perf_ytd", "perf_1y", "perf_3y", "perf_5y", "perf_10y"]
        for k in required:
            self.assertIn(k, row, f"Missing key: {k}")

    def test_price_resolved(self):
        with _patch_all():
            row = app.screener_row("TST")
        self.assertAlmostEqual(row["price"], 100.0)

    def test_unexpected_exception_becomes_error_row(self):
        # A failure anywhere past get_info (rate limit, parse error) must come
        # back as an error row, never raise — one bad ticker would otherwise
        # 500 its entire screener batch.
        with _patch_all(), \
             patch("app.performance", side_effect=RuntimeError("Too Many Requests")):
            row = app.screener_row("TST")
        self.assertEqual(row["ticker"], "TST")
        self.assertEqual(row["error"], "Too Many Requests")

    def test_exception_without_message_uses_type_name(self):
        with _patch_all(), \
             patch("app.get_dividends", side_effect=KeyError()):
            row = app.screener_row("TST")
        self.assertEqual(row["ticker"], "TST")
        self.assertTrue(row["error"])   # never an empty message

    def test_div_yield_computed(self):
        with _patch_all():
            row = app.screener_row("TST")
        # div_rate=2.0, price=100 → yield = 2.0%
        self.assertAlmostEqual(row["div_yield"], 2.0)

    def test_p_c_computed(self):
        with _patch_all():
            row = app.screener_row("TST")
        # marketCap=1e12, totalCash=10e9 → 100
        self.assertAlmostEqual(row["pc"], 100.0)

    def test_p_fcf_computed(self):
        with _patch_all():
            row = app.screener_row("TST")
        # marketCap=1e12, fcf=40e9 → 25
        self.assertAlmostEqual(row["p_fcf"], 25.0)

    def test_lt_debt_eq_computed(self):
        with _patch_all():
            row = app.screener_row("TST")
        # lt_debt=25e9, equity=50e9 → 50%
        self.assertAlmostEqual(row["lt_debt_to_equity"], 50.0)

    def test_fcf_coverage_computed(self):
        with _patch_all():
            row = app.screener_row("TST")
        # fcf=40e9, div_paid=-8e9 → 5.0
        self.assertAlmostEqual(row["fcf_coverage"], 5.0)

    def test_roic_computed(self):
        with _patch_all():
            row = app.screener_row("TST")
        # ebit=80e9, tax_rate=16/78≈0.205, debt=30e9, equity=50e9 → invested=80e9
        self.assertIsNotNone(row["roic"])
        self.assertGreater(row["roic"], 0)

    def test_roic_effective_tax_rate_capped(self):
        # Tax rate > 50% → capped at 0.5
        high_tax_info = {**_BASE_INFO}
        income = _make_df(
            ["EBIT", "Pretax Income", "Tax Provision", "Total Revenue", "Net Income", "Gross Profit"],
            ["2024-12-31"],
            {
                "EBIT": [10e9], "Pretax Income": [10e9], "Tax Provision": [8e9],  # 80% → capped at 50%
                "Total Revenue": [100e9], "Net Income": [2e9], "Gross Profit": [40e9],
            },
        )
        app.clear_cache()
        with patch("app.get_info", return_value=high_tax_info), \
             patch("app._get_stmt", side_effect=lambda t, attr: {
                 "income_stmt": income, "balance_sheet": _BAL_DF, "cash_flow": _CF_DF}.get(attr)), \
             patch("app.get_dividends", return_value=_DIVS), \
             patch("app.get_raw_close", return_value=_CLOSE_SERIES), \
             patch("app.performance", return_value={"ytd": 1.0, "1y": 1.0, "3y": 1.0, "5y": 1.0, "10y": 1.0}):
            row = app.screener_row("TST")
        roic_capped = 10e9 * (1 - 0.5) / (30e9 + 50e9) * 100
        self.assertAlmostEqual(row["roic"], roic_capped, places=0)

    def test_perf_fields_from_performance(self):
        with _patch_all():
            row = app.screener_row("TST")
        self.assertAlmostEqual(row["perf_ytd"], 5.0)
        self.assertAlmostEqual(row["perf_10y"], 100.0)

    def test_invalid_ticker_returns_error(self):
        app.clear_cache()
        with patch("app.get_info", return_value={}):
            row = app.screener_row("ZZZINVALID")
        self.assertIn("error", row)
        self.assertEqual(row["ticker"], "ZZZINVALID")

    def test_missing_optional_fields_return_none(self):
        sparse_info = {"shortName": "Sparse", "currentPrice": 50.0}
        app.clear_cache()
        with patch("app.get_info", return_value=sparse_info), \
             patch("app._get_stmt", return_value=None), \
             patch("app.get_dividends", return_value=None), \
             patch("app.get_raw_close", return_value=None), \
             patch("app.performance", return_value={"ytd": None, "1y": None, "3y": None, "5y": None, "10y": None}):
            row = app.screener_row("SPARSE")
        for k in ("pe", "forward_pe", "peg", "div_yield", "roic", "fcf_coverage",
                  "lt_debt_to_equity", "perf_10y"):
            self.assertIsNone(row[k], f"Expected None for {k}, got {row[k]}")


# ---------------------------------------------------------------------------
# 7b. Risk metrics — Altman Z-Score & Piotroski F-Score
# ---------------------------------------------------------------------------
class TestAltmanZ(unittest.TestCase):

    def _income(self):
        return _make_df(["Total Revenue", "EBIT"], ["2024-12-31"],
                        {"Total Revenue": [1200.0], "EBIT": [150.0]})

    def _balance(self, **overrides):
        base = {
            "Total Assets": [1000.0], "Current Assets": [400.0],
            "Current Liabilities": [200.0],
            "Total Liabilities Net Minority Interest": [500.0],
            "Retained Earnings": [300.0],
        }
        base.update(overrides)
        return _make_df(list(base), ["2024-12-31"], base)

    def test_known_value(self):
        # WC=200; Z = 1.2*.2 + 1.4*.3 + 3.3*.15 + 0.6*(2000/500) + 1.0*1.2 = 4.755
        z = app.altman_z(self._income(), self._balance(), 2000.0)
        self.assertAlmostEqual(z, 4.755, places=3)

    def test_none_without_market_cap(self):
        self.assertIsNone(app.altman_z(self._income(), self._balance(), None))

    def test_none_when_balance_missing_total_assets(self):
        bal = _make_df(["Current Assets"], ["2024-12-31"], {"Current Assets": [400.0]})
        self.assertIsNone(app.altman_z(self._income(), bal, 2000.0))


class TestAltmanZPrime(unittest.TestCase):
    """Z' — the private-firm variant S3 grades on. Book equity replaces market
    cap in X4, so the score carries no price term."""

    def _income(self):
        return _make_df(["Total Revenue", "EBIT"], ["2024-12-31"],
                        {"Total Revenue": [1200.0], "EBIT": [150.0]})

    def _balance(self, **overrides):
        base = {
            "Total Assets": [1000.0], "Current Assets": [400.0],
            "Current Liabilities": [200.0],
            "Total Liabilities Net Minority Interest": [500.0],
            "Retained Earnings": [300.0], "Stockholders Equity": [500.0],
        }
        base.update(overrides)
        return _make_df(list(base), ["2024-12-31"], base)

    def test_known_value(self):
        # X1=.2 X2=.3 X3=.15 X4=500/500=1.0 X5=1.2
        # Z' = 0.717*.2 + 0.847*.3 + 3.107*.15 + 0.420*1.0 + 0.998*1.2 = 2.48115
        zp = app.altman_z_prime(self._income(), self._balance())
        self.assertAlmostEqual(zp, 2.48115, places=5)

    def test_carries_no_price_term(self):
        # The property the whole S3 Pillar C change rests on. The classic Z
        # drops 1.2 points when the stock halves (0.6 * 1000/500) — in a value
        # strategy that means Pillar C takes back what Pillars A/B award for
        # the discount. Z' does not move at all.
        inc, bal = self._income(), self._balance()
        z_rich, z_cheap = app.altman_z(inc, bal, 2000.0), app.altman_z(inc, bal, 1000.0)
        self.assertAlmostEqual(z_rich - z_cheap, 1.2, places=6)
        self.assertAlmostEqual(app.altman_z_prime(inc, bal), 2.48115, places=5)

    def test_negative_equity_drags_score_down(self):
        pos = app.altman_z_prime(self._income(), self._balance())
        neg = app.altman_z_prime(
            self._income(), self._balance(**{"Stockholders Equity": [-500.0]}))
        self.assertLess(neg, pos)
        # X4 swings 1.0 -> -1.0, weighted 0.420
        self.assertAlmostEqual(pos - neg, 0.840, places=6)

    def test_none_without_equity(self):
        bal = self._balance()
        bal = bal.drop(index="Stockholders Equity")
        self.assertIsNone(app.altman_z_prime(self._income(), bal))

    def test_none_when_balance_missing_total_assets(self):
        bal = _make_df(["Current Assets"], ["2024-12-31"], {"Current Assets": [400.0]})
        self.assertIsNone(app.altman_z_prime(self._income(), bal))

    def test_equity_label_fallback(self):
        # Yahoo's label varies by filer; all three spellings must resolve.
        for label in ("Total Stockholder Equity", "Common Stock Equity"):
            base = {
                "Total Assets": [1000.0], "Current Assets": [400.0],
                "Current Liabilities": [200.0],
                "Total Liabilities Net Minority Interest": [500.0],
                "Retained Earnings": [300.0], label: [500.0],
            }
            bal = _make_df(list(base), ["2024-12-31"], base)
            self.assertAlmostEqual(
                app.altman_z_prime(self._income(), bal), 2.48115, places=5,
                msg=f"equity label {label!r} did not resolve")


class TestPiotroskiF(unittest.TestCase):

    # Two-year fixtures engineered so all 9 criteria pass -> score 9.
    _INC = _make_df(
        ["Net Income", "Total Revenue", "Gross Profit", "Diluted Average Shares"],
        ["2024-12-31", "2023-12-31"],
        {
            "Net Income":             [100.0, 80.0],
            "Total Revenue":          [1200.0, 1100.0],
            "Gross Profit":           [600.0, 520.0],
            "Diluted Average Shares": [1000.0, 1010.0],
        },
    )
    _BAL = _make_df(
        ["Total Assets", "Current Assets", "Current Liabilities", "Long Term Debt"],
        ["2024-12-31", "2023-12-31"],
        {
            "Total Assets":        [1000.0, 1000.0],
            "Current Assets":      [400.0, 360.0],
            "Current Liabilities": [200.0, 200.0],
            "Long Term Debt":      [200.0, 250.0],
        },
    )
    _CF = _make_df(["Operating Cash Flow"], ["2024-12-31", "2023-12-31"],
                   {"Operating Cash Flow": [120.0, 90.0]})

    def test_perfect_score(self):
        self.assertEqual(app.piotroski_f(self._INC, self._BAL, self._CF), 9)

    def test_score_in_range(self):
        s = app.piotroski_f(self._INC, self._BAL, self._CF)
        self.assertIsInstance(s, int)
        self.assertTrue(0 <= s <= 9)

    def test_none_without_prior_year(self):
        one_yr_inc = _make_df(["Net Income"], ["2024-12-31"], {"Net Income": [100.0]})
        one_yr_bal = _make_df(["Total Assets"], ["2024-12-31"], {"Total Assets": [1000.0]})
        self.assertIsNone(app.piotroski_f(one_yr_inc, one_yr_bal, self._CF))

    def test_loss_making_lowers_score(self):
        # Flip net income negative in both years -> several criteria fail
        inc = _make_df(
            ["Net Income", "Total Revenue", "Gross Profit"],
            ["2024-12-31", "2023-12-31"],
            {"Net Income": [-50.0, -30.0], "Total Revenue": [1000.0, 1100.0],
             "Gross Profit": [400.0, 480.0]},
        )
        s = app.piotroski_f(inc, self._BAL, self._CF)
        self.assertLess(s, 9)


class TestBvpsGrowth(unittest.TestCase):
    """Annualized book-value-per-share trend (mortgage-REIT quality signal)."""

    def _bal(self, equity, shares, dates):
        return _make_df(["Stockholders Equity", "Ordinary Shares Number"], dates,
                        {"Stockholders Equity": equity, "Ordinary Shares Number": shares})

    def test_growing_book_value_is_positive(self):
        # BVPS 8 -> 10 over 2 year-gaps: (10/8)^(1/2)-1 ≈ +11.8%/yr
        bal = self._bal([1000.0, 900.0, 800.0], [100.0, 100.0, 100.0],
                        ["2024-12-31", "2023-12-31", "2022-12-31"])
        g = app._bvps_growth(bal)
        self.assertAlmostEqual(g, ((10.0 / 8.0) ** 0.5 - 1) * 100, places=4)

    def test_eroding_book_value_is_negative(self):
        # newest year (2024) has lower BVPS than the prior year (2023): equity
        # shrank AND shares grew, so BVPS falls 1000/100=10 -> 900/110≈8.2.
        bal = self._bal([900.0, 1000.0], [110.0, 100.0],
                        ["2024-12-31", "2023-12-31"])
        self.assertLess(app._bvps_growth(bal), 0)

    def test_none_without_two_years(self):
        bal = self._bal([1000.0], [100.0], ["2024-12-31"])
        self.assertIsNone(app._bvps_growth(bal))

    def test_none_on_nonpositive_equity(self):
        bal = self._bal([-500.0, 900.0], [100.0, 100.0],
                        ["2024-12-31", "2023-12-31"])
        self.assertIsNone(app._bvps_growth(bal))


class TestFcfCagr(unittest.TestCase):
    """Annualized FCF growth from the cash-flow statement (DCF stage-1 input)."""

    def _cf(self, fcf, dates):
        return _make_df(["Free Cash Flow"], dates, {"Free Cash Flow": fcf})

    def test_growth_from_fixture(self):
        # _CF_DF: 38e9 (2023) -> 40e9 (2024), one year gap
        self.assertAlmostEqual(app._fcf_cagr(_CF_DF),
                               (40.0 / 38.0 - 1) * 100, places=6)

    def test_multi_year_cagr(self):
        # 100 -> 121 over 2 year-gaps: exactly 10%/yr
        cf = self._cf([121.0, 110.0, 100.0],
                      ["2024-12-31", "2023-12-31", "2022-12-31"])
        self.assertAlmostEqual(app._fcf_cagr(cf), 10.0, places=6)

    def test_none_single_period(self):
        self.assertIsNone(app._fcf_cagr(self._cf([100.0], ["2024-12-31"])))

    def test_none_on_nonpositive_endpoint(self):
        cf = self._cf([100.0, -50.0], ["2024-12-31", "2023-12-31"])
        self.assertIsNone(app._fcf_cagr(cf))

    def test_none_without_fcf_row_or_df(self):
        df = _make_df(["Operating Cash Flow"], ["2024-12-31"],
                      {"Operating Cash Flow": [100.0]})
        self.assertIsNone(app._fcf_cagr(df))
        self.assertIsNone(app._fcf_cagr(None))


class TestDcfEquityValue(unittest.TestCase):
    """Two-stage FCFF DCF math (see app._dcf_equity_value)."""

    def test_flat_growth_matches_closed_form(self):
        # g0 = terminal g, so every year grows 2.5% and the whole model
        # collapses to a closed form: PV = fcf·x·(1−x^N)/(1−x) + TV/(1+w)^N
        # with x = 1.025/1.10.
        fcf, w, g = 100.0, 0.10, 0.025
        x = (1 + g) / (1 + w)
        pv_stage1 = fcf * x * (1 - x ** 10) / (1 - x)
        fcf10 = fcf * (1 + g) ** 10
        pv_tv = (fcf10 * (1 + g) / (w - g)) / (1 + w) ** 10
        expected = pv_stage1 + pv_tv
        got = app._dcf_equity_value(100.0, 2.5, 10.0, 0, 0)
        self.assertAlmostEqual(got, expected, places=6)

    def test_g0_none_falls_back_to_terminal(self):
        self.assertEqual(app._dcf_equity_value(100.0, None, 10.0, 0, 0),
                         app._dcf_equity_value(100.0, 2.5, 10.0, 0, 0))

    def test_g0_clamped_to_band(self):
        self.assertEqual(app._dcf_equity_value(100.0, 50.0, 10.0, 0, 0),
                         app._dcf_equity_value(100.0, 20.0, 10.0, 0, 0))
        self.assertEqual(app._dcf_equity_value(100.0, -5.0, 10.0, 0, 0),
                         app._dcf_equity_value(100.0, 0.0, 10.0, 0, 0))

    def test_higher_growth_is_worth_more(self):
        self.assertGreater(app._dcf_equity_value(100.0, 15.0, 10.0, 0, 0),
                           app._dcf_equity_value(100.0, 5.0, 10.0, 0, 0))

    def test_guards_return_none(self):
        self.assertIsNone(app._dcf_equity_value(None, 5.0, 10.0, 0, 0))
        self.assertIsNone(app._dcf_equity_value(0.0, 5.0, 10.0, 0, 0))
        self.assertIsNone(app._dcf_equity_value(-10.0, 5.0, 10.0, 0, 0))
        self.assertIsNone(app._dcf_equity_value(100.0, 5.0, None, 0, 0))
        # WACC must clear terminal growth (2.5) by the margin (0.5): 3.0 fails
        self.assertIsNone(app._dcf_equity_value(100.0, 5.0, 3.0, 0, 0))

    def test_net_debt_bridge(self):
        base = app._dcf_equity_value(100.0, 2.5, 10.0, 0, 0)
        self.assertAlmostEqual(
            app._dcf_equity_value(100.0, 2.5, 10.0, 100.0, 40.0), base - 60.0)
        # None debt/cash treated as zero
        self.assertAlmostEqual(
            app._dcf_equity_value(100.0, 2.5, 10.0, None, None), base)


class TestDcfDetail(unittest.TestCase):
    """_dcf_detail's breakdown must reconcile with _dcf_equity_value, and
    _wacc_detail with _compute_wacc (both pairs are wrapper + detail)."""

    def test_equity_value_matches_wrapper(self):
        d = app._dcf_detail(100.0, 8.0, 10.0, 30.0, 10.0)
        self.assertEqual(d["equity_value"],
                         app._dcf_equity_value(100.0, 8.0, 10.0, 30.0, 10.0))

    def test_none_conditions_match_wrapper(self):
        for args in ((None, 5.0, 10.0, 0, 0), (0.0, 5.0, 10.0, 0, 0),
                     (100.0, 5.0, None, 0, 0), (100.0, 5.0, 3.0, 0, 0)):
            self.assertIsNone(app._dcf_detail(*args))
            self.assertIsNone(app._dcf_equity_value(*args))

    def test_fade_endpoints_and_structure(self):
        d = app._dcf_detail(100.0, 12.0, 10.0, 0, 0)
        years = d["years"]
        self.assertEqual(len(years), app._DCF_YEARS)
        self.assertEqual([y["year"] for y in years], list(range(1, 11)))
        self.assertAlmostEqual(years[0]["growth"], 12.0)              # year 1 = g0
        self.assertAlmostEqual(years[-1]["growth"], app._DCF_TERMINAL_G)  # year N = terminal
        self.assertEqual(d["g0_used"], 12.0)

    def test_pv_components_reconcile(self):
        d = app._dcf_detail(100.0, 8.0, 10.0, 30.0, 10.0)
        self.assertAlmostEqual(d["stage1_pv"], sum(y["pv"] for y in d["years"]))
        self.assertAlmostEqual(d["terminal_pv"],
                               d["terminal_value"] * d["terminal_df"])
        self.assertAlmostEqual(d["enterprise_value"],
                               d["stage1_pv"] + d["terminal_pv"])
        self.assertAlmostEqual(d["equity_value"],
                               d["enterprise_value"] - 30.0 + 10.0)
        # per-year rows are internally consistent too
        for y in d["years"]:
            self.assertAlmostEqual(y["pv"], y["fcf"] * y["discount_factor"])

    def test_wacc_detail_matches_compute_wacc(self):
        args = (1.2, 1e12, 30e9, 1.5e9, 0.21, 4.0)
        d = app._wacc_detail(*args)
        self.assertEqual(d["wacc"], app._compute_wacc(*args))
        self.assertAlmostEqual(d["w_e"] + d["w_d"], 1.0)
        self.assertAlmostEqual(d["cost_of_equity"], 4.0 + 1.2 * app._ERP)
        self.assertIsNone(app._wacc_detail(None, 1e12, 0, None, 0.21, 4.0))
        self.assertIsNone(app._compute_wacc(None, 1e12, 0, None, 0.21, 4.0))


class TestScreenerRowDcf(unittest.TestCase):
    """dcf_value / dcf_upside wiring inside _screener_row."""

    def setUp(self):
        app.clear_cache()

    def _row(self, info_overrides=None, fx=None):
        info = {**_BASE_INFO, "beta": 1.0, "sharesOutstanding": 1e10,
                **(info_overrides or {})}
        with _patch_all(), \
             patch("app.get_info", return_value=info), \
             patch("app.get_risk_free_rate", return_value=4.0):
            if fx:
                with patch("app._fx_rate", side_effect=fx):
                    return app.screener_row("TST")
            return app.screener_row("TST")

    def test_happy_path_cross_checks_helper(self):
        row = self._row()
        self.assertIsNotNone(row["dcf_value"])
        expected = app._dcf_equity_value(
            40e9, app._fcf_cagr(_CF_DF), row["wacc"], 30e9, 10e9) / 1e10
        self.assertAlmostEqual(row["dcf_value"], expected, places=6)
        self.assertAlmostEqual(row["dcf_upside"],
                               (row["dcf_value"] / 100.0 - 1) * 100, places=6)

    def test_shares_fallback_from_market_cap(self):
        # No sharesOutstanding: falls back to market_cap/price = 1e12/100 = 1e10
        explicit = self._row()
        info = {**_BASE_INFO, "beta": 1.0}
        with _patch_all(), \
             patch("app.get_info", return_value=info), \
             patch("app.get_risk_free_rate", return_value=4.0):
            fallback = app.screener_row("TST")
        self.assertAlmostEqual(fallback["dcf_value"], explicit["dcf_value"])

    def test_none_when_wacc_unavailable(self):
        # _BASE_INFO has no beta -> WACC None -> DCF None (keys still present)
        with _patch_all():
            row = app.screener_row("TST")
        self.assertIsNone(row["dcf_value"])
        self.assertIsNone(row["dcf_upside"])

    def test_blanked_for_financials_and_reits(self):
        bank = self._row({"sector": "Financial Services",
                          "industry": "Banks — Diversified"})
        self.assertIsNone(bank["dcf_value"])
        reit = self._row({"sector": "Real Estate", "industry": "REIT — Retail"})
        self.assertIsNone(reit["dcf_value"])
        mreit = self._row({"sector": "Real Estate", "industry": "REIT — Mortgage"})
        self.assertIsNone(mreit["dcf_value"])

    def test_fx_mismatch_converts_to_trading_currency(self):
        # Trades USD, reports INR: the per-share DCF (reporting ccy) must come
        # back in USD via _fx_rate(INR, USD) so it compares against price.
        rates = {("USD", "INR"): 83.0, ("INR", "USD"): 1 / 83.0}
        row = self._row({"currency": "USD", "financialCurrency": "INR"},
                        fx=lambda b, q: rates.get((b, q), 1.0))
        self.assertIsNotNone(row["dcf_value"])
        expected = app._dcf_equity_value(
            40e9, app._fcf_cagr(_CF_DF), row["wacc"], 30e9, 10e9) / 1e10 / 83.0
        self.assertAlmostEqual(row["dcf_value"], expected, places=6)

    def test_dcf_keys_registered_as_non_metrics(self):
        # Blanked-by-design fields must not count toward Stage-0 quarantine
        self.assertIn("dcf_value", strategies._NON_METRIC_KEYS)
        self.assertIn("dcf_upside", strategies._NON_METRIC_KEYS)
        self.assertIn("financial_currency", strategies._NON_METRIC_KEYS)


@unittest.skipUnless(app._HAS_OPENPYXL, "openpyxl not installed")
class TestDcfWorkbook(unittest.TestCase):
    """build_dcf_workbook — the per-stock DCF-valuation Excel export."""

    def setUp(self):
        app.clear_cache()

    def _build(self, info_overrides=None):
        info = {**_BASE_INFO, "beta": 1.0, "sharesOutstanding": 1e10,
                **(info_overrides or {})}
        with _patch_all(), \
             patch("app.get_info", return_value=info), \
             patch("app.get_risk_free_rate", return_value=4.0):
            row = app.screener_row("TST")
            app.clear_cache()   # row is rebuilt inside the builder from the same mocks
            return row, app.build_dcf_workbook("TST")

    def _cells(self, data):
        """(worksheet, {label: colB-value}) from the DCF sheet. Column B now
        holds Excel *formulas* (strings starting with '=') for every derived
        cell — openpyxl does not evaluate them, so value assertions here target
        the input cells; formula wiring is asserted separately, and the numbers
        are covered by TestDcfDetail / the live LibreOffice recalc check."""
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data))
        ws = wb["DCF Valuation"]
        return ws, {ws.cell(row=i, column=1).value: ws.cell(row=i, column=2).value
                    for i in range(1, ws.max_row + 1)
                    if ws.cell(row=i, column=1).value is not None}

    def test_inputs_carry_real_values(self):
        row, data = self._build()
        _, cells = self._cells(data)
        # input cells hold plain values (not formulas)
        self.assertAlmostEqual(cells["Base FCF — latest annual (USD)"], row["fcf"])
        self.assertAlmostEqual(cells["Shares Outstanding"], 1e10)
        self.assertAlmostEqual(cells["Terminal growth"], app._DCF_TERMINAL_G / 100.0)
        self.assertEqual(cells["Projection horizon (years)"], app._DCF_YEARS)
        self.assertAlmostEqual(cells["Beta"], 1.0)
        self.assertAlmostEqual(cells["Risk-free rate (10Y Treasury)"], 0.04)
        self.assertAlmostEqual(cells["Equity risk premium (fixed US estimate)"],
                               app._ERP / 100.0)

    def test_derived_cells_are_live_formulas(self):
        _, data = self._build()
        _, cells = self._cells(data)
        # every downstream number is a formula, not a baked value
        for label in ("DCF Value (USD / share)", "DCF Upside vs Price",
                      "Cost of equity = rfr + β × ERP",
                      "Equity weight = mktcap ÷ (mktcap + debt)",
                      "Debt weight = debt ÷ (mktcap + debt)",
                      "Effective tax = min(max(tax ÷ pretax, 0), 50%)",
                      "WACC = wE·costE + wD·costD·(1−tax)",
                      "Enterprise value (DCF)", "Equity value",
                      "DCF Value (USD)", "Upside vs price"):
            v = cells[label]
            self.assertTrue(isinstance(v, str) and v.startswith("="),
                            f"{label!r} should be a formula, got {v!r}")

    def test_no_label_cell_is_a_formula(self):
        # Regression: labels starting with "=" (e.g. "= Equity value") were
        # silently stored as formulas and rendered as Err:509. No column-A
        # label cell may be a formula.
        from openpyxl import load_workbook
        _, data = self._build()
        ws = load_workbook(io.BytesIO(data))["DCF Valuation"]
        for i in range(1, ws.max_row + 1):
            c = ws.cell(row=i, column=1)
            self.assertNotEqual(c.data_type, "f",
                                f"row {i} label is a formula: {c.value!r}")

    def test_projection_table_structure(self):
        _, data = self._build()
        ws, _ = self._cells(data)
        col_a = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row + 1)]
        self.assertTrue(all(y in col_a for y in range(1, app._DCF_YEARS + 1)))
        self.assertIn("Terminal", col_a)
        # a projection PV cell (column E) is a formula
        for i in range(1, ws.max_row + 1):
            if ws.cell(row=i, column=1).value == 1:      # year-1 row
                self.assertTrue(str(ws.cell(row=i, column=5).value).startswith("="))
                break

    def test_blanked_archetype_explains_why(self):
        _, data = self._build({"sector": "Financial Services",
                               "industry": "Banks — Diversified"})
        _, cells = self._cells(data)
        self.assertIsNone(cells["DCF Value (USD / share)"])
        self.assertIn("business type", cells["Why N/A"])
        # no projection/bridge built for a blanked archetype
        self.assertNotIn("Enterprise value (DCF)", cells)

    def test_wacc_missing_explains_why(self):
        # No beta -> WACC None -> N/A with the WACC reason
        info = {**_BASE_INFO, "sharesOutstanding": 1e10}
        with _patch_all(), patch("app.get_info", return_value=info):
            data = app.build_dcf_workbook("TST")
        _, cells = self._cells(data)
        self.assertIn("WACC unavailable", cells["Why N/A"])

    def test_error_row_raises(self):
        with patch("app.get_info", return_value={}):
            with self.assertRaises(ValueError):
                app.build_dcf_workbook("BAD")


# ---------------------------------------------------------------------------
# 8. deepdive (panel structure + new fields)
# ---------------------------------------------------------------------------
class TestDeepdive(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def _run(self, ticker="TST"):
        div_info = {
            **_BASE_INFO,
            "exDividendDate": int(datetime.datetime(2026, 5, 11,
                tzinfo=datetime.timezone.utc).timestamp()),
        }
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        app.clear_cache()
        with patch("app.get_info", return_value=div_info), \
             patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.dividend_growth", return_value={
                 "cagr_3y": 5.0, "cagr_5y": 4.0, "annual": []}):
            return app.deepdive(ticker)

    def test_top_level_keys(self):
        d = self._run()
        for k in ("ticker", "name", "price", "panels", "revenue_net_income", "dividend_growth"):
            self.assertIn(k, d)

    def test_panel_names(self):
        d = self._run()
        self.assertEqual(set(d["panels"].keys()),
                         {"valuation", "dividend", "profitability", "health",
                          "risk", "strategies"})

    def test_no_reit_panel_for_non_reit(self):
        # _BASE_INFO is Technology/Software -> no "reit" panel.
        self.assertNotIn("reit", self._run()["panels"])

    def _run_reit(self, industry, cashflow, extra_info):
        info = {**_BASE_INFO, "sector": "Real Estate", "industry": industry,
                **extra_info}
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = cashflow
        mock_tk.balance_sheet = _BAL_DF
        app.clear_cache()
        with patch("app.get_info", return_value=info), \
             patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.dividend_growth", return_value={"cagr_3y": 5.0, "cagr_5y": 4.0, "annual": []}):
            return app.deepdive("REITX")

    def test_equity_reit_gets_ffo_panel(self):
        # Cash flow carries a D&A line -> FFO = NI + D&A is computable.
        cf = _make_df(
            ["Free Cash Flow", "Operating Cash Flow", "Cash Dividends Paid",
             "Depreciation And Amortization"],
            ["2024-12-31", "2023-12-31"],
            {"Free Cash Flow": [40e9, 38e9], "Operating Cash Flow": [48e9, 44e9],
             "Cash Dividends Paid": [-8e9, -7e9],
             "Depreciation And Amortization": [30e9, 28e9]})
        d = self._run_reit("REIT - Retail", cf,
                           {"bookValue": 40.0, "sharesOutstanding": 1e9})
        reit = d["panels"]["reit"]
        self.assertEqual(set(reit), {"FFO", "FFO/Share", "P/FFO", "FFO Payout %",
                                     "FFO Coverage", "Book Value/Share", "Price/Book"})
        # NI (50e9) + D&A (30e9) = 80e9
        self.assertAlmostEqual(reit["FFO"], 80e9)
        self.assertAlmostEqual(reit["FFO/Share"], 80.0)          # 80e9 / 1e9 shares
        self.assertAlmostEqual(reit["P/FFO"], 12.5)              # mktcap 1e12 / 80e9
        self.assertAlmostEqual(reit["FFO Payout %"], 10.0)       # 8e9 / 80e9
        self.assertAlmostEqual(reit["FFO Coverage"], 10.0)       # 80e9 / 8e9
        self.assertAlmostEqual(reit["Book Value/Share"], 40.0)

    def test_ffo_backs_out_property_sale_gains_and_adds_back_impairments(self):
        # NAREIT FFO excludes both, so a REIT can neither flatter its payout by
        # selling buildings nor look stressed for writing one down. Both lines
        # arrive from the cash flow reconciliation already signed to add: a
        # backed-out gain negative, an impairment positive.
        cf = _make_df(
            ["Free Cash Flow", "Operating Cash Flow", "Cash Dividends Paid",
             "Depreciation And Amortization", "Operating Gains Losses",
             "Asset Impairment Charge"],
            ["2024-12-31", "2023-12-31"],
            {"Free Cash Flow": [40e9, 38e9], "Operating Cash Flow": [48e9, 44e9],
             "Cash Dividends Paid": [-8e9, -7e9],
             "Depreciation And Amortization": [30e9, 28e9],
             "Operating Gains Losses": [-12e9, -3e9],
             "Asset Impairment Charge": [2e9, 0.0]})
        d = self._run_reit("REIT - Retail", cf,
                           {"bookValue": 40.0, "sharesOutstanding": 1e9})
        reit = d["panels"]["reit"]
        # NI (50e9) + D&A (30e9) − gains (12e9) + impairment (2e9) = 70e9,
        # against the 80e9 the unadjusted NI + D&A would have claimed.
        self.assertAlmostEqual(reit["FFO"], 70e9)
        self.assertAlmostEqual(reit["P/FFO"], 1e12 / 70e9)
        # The payout the dividend really represents: 11.4%, not 10.0%.
        self.assertAlmostEqual(reit["FFO Payout %"], 8e9 / 70e9 * 100)

    def test_mortgage_reit_gets_book_value_panel(self):
        # No D&A line -> FFO is None -> the book-value branch is used instead.
        cf = _make_df(
            ["Free Cash Flow", "Operating Cash Flow", "Cash Dividends Paid"],
            ["2024-12-31", "2023-12-31"],
            {"Free Cash Flow": [40e9, 38e9], "Operating Cash Flow": [48e9, 44e9],
             "Cash Dividends Paid": [-8e9, -7e9]})
        d = self._run_reit("REIT - Mortgage", cf,
                           {"bookValue": 7.0, "netIncomeToCommon": 4e9})
        reit = d["panels"]["reit"]
        self.assertEqual(set(reit),
                         {"Book Value/Share", "Book Value Trend %", "Price/Book",
                          "Net Income", "Div Coverage (NI)"})
        self.assertNotIn("FFO", reit)                            # no fabricated FFO
        self.assertAlmostEqual(reit["Book Value/Share"], 7.0)
        self.assertAlmostEqual(reit["Net Income"], 4e9)
        self.assertAlmostEqual(reit["Div Coverage (NI)"], 0.5)  # 4e9 / |−8e9|

    def test_mortgage_reit_with_da_line_still_gets_book_value_panel(self):
        # Some mREITs (NLY) do report a D&A line, so FFO must be excluded by
        # business type, not by the line happening to be absent — otherwise they
        # land on the equity branch and show an FFO that means nothing for a
        # securities portfolio.
        cf = _make_df(
            ["Free Cash Flow", "Operating Cash Flow", "Cash Dividends Paid",
             "Depreciation And Amortization", "Operating Gains Losses"],
            ["2024-12-31", "2023-12-31"],
            {"Free Cash Flow": [40e9, 38e9], "Operating Cash Flow": [48e9, 44e9],
             "Cash Dividends Paid": [-8e9, -7e9],
             "Depreciation And Amortization": [1e9, 1e9],
             "Operating Gains Losses": [-5e9, -2e9]})
        d = self._run_reit("REIT - Mortgage", cf,
                           {"bookValue": 7.0, "netIncomeToCommon": 4e9})
        self.assertNotIn("FFO", d["panels"]["reit"])
        self.assertIn("Book Value/Share", d["panels"]["reit"])

    def test_valuation_keys(self):
        d = self._run()
        v = d["panels"]["valuation"]
        for k in ("Market Cap", "Enterprise Value", "Trailing P/E", "Forward P/E",
                  "PEG Ratio", "Price/Book", "Price/Sales", "EV/EBITDA"):
            self.assertIn(k, v, f"Valuation panel missing: {k}")

    def test_dividend_panel_order(self):
        d = self._run()
        keys = list(d["panels"]["dividend"].keys())
        pr_idx = keys.index("Payout Ratio %")
        fcf_idx = keys.index("FCF Coverage")
        # FCF Coverage must appear immediately after Payout Ratio %
        self.assertEqual(fcf_idx, pr_idx + 1,
                         f"FCF Coverage at {fcf_idx}, Payout Ratio at {pr_idx}")

    def test_dividend_fcf_coverage_value(self):
        d = self._run()
        # FCF=40e9, Cash Dividends Paid=-8e9 → coverage = 5.0
        self.assertAlmostEqual(d["panels"]["dividend"]["FCF Coverage"], 5.0)

    def test_health_panel_has_total_equity(self):
        d = self._run()
        h = d["panels"]["health"]
        self.assertIn("Total Equity", h)
        self.assertAlmostEqual(h["Total Equity"], 50e9)

    def test_health_panel_order_equity_after_debt(self):
        d = self._run()
        keys = list(d["panels"]["health"].keys())
        self.assertLess(keys.index("Total Debt"), keys.index("Total Equity"),
                        "Total Equity should appear after Total Debt")
        self.assertLess(keys.index("Total Equity"), keys.index("Debt/Equity"),
                        "Total Equity should appear before Debt/Equity")

    def test_health_panel_has_ocf_and_capex(self):
        d = self._run()
        h = d["panels"]["health"]
        # Operating Cash Flow: direct statement hit (_CF_DF's 2024 value).
        self.assertAlmostEqual(h["Operating Cash Flow"], 48e9)
        # Capital Expenditure: _CF_DF has no such row, so it's derived as
        # FCF − OCF = 40e9 − 48e9 = −8e9 (a cash outflow, correctly negative).
        self.assertAlmostEqual(h["Capital Expenditure"], -8e9)
        # OCF + Capex should reconcile back to Free Cash Flow.
        self.assertAlmostEqual(h["Operating Cash Flow"] + h["Capital Expenditure"],
                               h["Free Cash Flow"])

    def test_health_panel_ocf_capex_order(self):
        d = self._run()
        keys = list(d["panels"]["health"].keys())
        self.assertLess(keys.index("Operating Cash Flow"), keys.index("Capital Expenditure"))
        self.assertLess(keys.index("Capital Expenditure"), keys.index("Free Cash Flow"))

    def test_profitability_panel_has_revenue_and_operating_income(self):
        d = self._run()
        p = d["panels"]["profitability"]
        # Revenue: straight from info.totalRevenue.
        self.assertAlmostEqual(p["Revenue"], 400e9)
        # Operating Income: revenue × operatingMargins = 400e9 × 0.20 = 80e9.
        self.assertAlmostEqual(p["Operating Income"], 80e9)

    def test_profitability_panel_revenue_first_operating_income_before_net(self):
        d = self._run()
        keys = list(d["panels"]["profitability"].keys())
        self.assertEqual(keys[0], "Revenue")
        self.assertLess(keys.index("Operating Income"), keys.index("Net Income"))

    def test_capex_prefers_statement_over_derivation(self):
        # When the cash flow statement reports "Capital Expenditure" directly,
        # that value wins over the FCF−OCF fallback.
        cf_with_capex = _make_df(
            ["Free Cash Flow", "Operating Cash Flow", "Capital Expenditure", "Cash Dividends Paid"],
            ["2024-12-31", "2023-12-31"],
            {
                "Free Cash Flow":      [40e9, 38e9],
                "Operating Cash Flow": [48e9, 44e9],
                "Capital Expenditure": [-9e9, -8e9],   # deliberately != FCF-OCF (-8e9)
                "Cash Dividends Paid": [-8e9, -7e9],
            },
        )
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = cf_with_capex
        mock_tk.balance_sheet = _BAL_DF
        app.clear_cache()
        with patch("app.get_info", return_value=_BASE_INFO), \
             patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.dividend_growth", return_value={"cagr_3y": 5.0, "cagr_5y": 4.0, "annual": []}):
            d = app.deepdive("CAPEX_STMT")
        self.assertAlmostEqual(d["panels"]["health"]["Capital Expenditure"], -9e9)

    def test_roic_history_matches_per_year_formula(self):
        d = self._run()
        hist = {r["period"]: r["roic"] for r in d["roic_history"]}
        self.assertIn("2024", hist)
        self.assertIn("2023", hist)
        # 2024: EBIT=80e9, Pretax=78e9, Tax=16e9, Debt=30e9, Equity=50e9
        tr_2024 = 16e9 / 78e9
        expected_2024 = 80e9 * (1 - tr_2024) / (30e9 + 50e9) * 100
        self.assertAlmostEqual(hist["2024"], expected_2024)
        # 2023: EBIT=75e9, Pretax=73e9, Tax=15e9, Debt=27e9, Equity=45e9
        tr_2023 = 15e9 / 73e9
        expected_2023 = 75e9 * (1 - tr_2023) / (27e9 + 45e9) * 100
        self.assertAlmostEqual(hist["2023"], expected_2023)

    def test_roic_history_latest_year_matches_current_roic(self):
        # The newest year in the history uses the same latest-period statement
        # figures and formula as the single "current" ROIC % in the
        # Profitability panel, so the two must agree exactly.
        d = self._run()
        latest = d["roic_history"][-1]["roic"]
        self.assertAlmostEqual(latest, d["panels"]["profitability"]["ROIC %"])

    def test_wacc_current_none_when_beta_missing(self):
        # _BASE_INFO has no "beta" -> _compute_wacc can't run -> None.
        d = self._run()
        self.assertIsNone(d["wacc_current"])

    def test_wacc_current_matches_panel_when_beta_present(self):
        info = {**_BASE_INFO, "beta": 1.2}
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        app.clear_cache()
        with patch("app.get_info", return_value=info), \
             patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.get_risk_free_rate", return_value=4.0), \
             patch("app.dividend_growth", return_value={"cagr_3y": 5.0, "cagr_5y": 4.0, "annual": []}):
            d = app.deepdive("WACC_TST")
        self.assertIsNotNone(d["wacc_current"])
        self.assertAlmostEqual(d["wacc_current"], d["panels"]["profitability"]["WACC %"])

    def test_revenue_net_income_has_four_series(self):
        d = self._run()
        row = d["revenue_net_income"][0]
        for k in ("revenue", "gross_profit", "net_income", "fcf"):
            self.assertIn(k, row, f"Chart series missing: {k}")

    def test_ex_dividend_date_formatted(self):
        d = self._run()
        ex_date = d["panels"]["dividend"].get("Ex-Dividend Date")
        self.assertEqual(ex_date, "2026-05-11")

    def test_profitability_panel_keys(self):
        d = self._run()
        p = d["panels"]["profitability"]
        for k in ("Gross Margin %", "Operating Margin %", "Profit Margin %", "ROE %", "ROA %"):
            self.assertIn(k, p)

    def test_caching_reuses_result(self):
        # Call twice; second call should return cached dict (same object)
        d1 = self._run("CACHE_TST")
        d2 = app.deepdive("CACHE_TST")   # hits cache
        self.assertIs(d1, d2)

    def test_profitability_has_new_ratios(self):
        d = self._run()
        p = d["panels"]["profitability"]
        for k in ("EBITDA Margin %", "ROIC %", "ROCE %"):
            self.assertIn(k, p, f"Profitability panel missing: {k}")

    def test_revenue_net_income_has_margins(self):
        d = self._run()
        row = d["revenue_net_income"][0]
        for k in ("gross_margin", "operating_margin", "net_margin"):
            self.assertIn(k, row, f"Margin series missing: {k}")
        # Gross margin 2024 = 180e9 / 400e9 * 100 = 45.0
        latest = d["revenue_net_income"][-1]
        self.assertAlmostEqual(latest["gross_margin"], 45.0)

    def test_revenue_net_income_has_operating_income_bar(self):
        # Operating Margin is rendered as a $ bar -> needs an absolute amount.
        d = self._run()
        self.assertIn("operating_income", d["revenue_net_income"][0])

    def test_growth_present_and_is_list(self):
        d = self._run()
        self.assertIn("growth", d)
        self.assertIsInstance(d["growth"], list)

    def test_risk_panel_keys(self):
        d = self._run()
        r = d["panels"]["risk"]
        for k in ("Beta", "Short Interest %", "Days to Cover",
                  "Altman Z-Score", "Piotroski F-Score"):
            self.assertIn(k, r, f"Risk panel missing: {k}")

    def test_beta_moved_from_health_to_risk(self):
        d = self._run()
        self.assertNotIn("Beta", d["panels"]["health"])
        self.assertIn("Beta", d["panels"]["risk"])


# ---------------------------------------------------------------------------
# 8b. deepdive growth panel — YoY bars over the last 3 years
# ---------------------------------------------------------------------------
_GROWTH_INCOME_DF = _make_df(
    ["Total Revenue", "Net Income", "Gross Profit", "Operating Income",
     "EBITDA", "Diluted EPS", "Diluted Average Shares"],
    ["2025-12-31", "2024-12-31", "2023-12-31", "2022-12-31"],
    {
        # newest -> oldest (yfinance column order); _series_from_stmt re-sorts ascending
        "Total Revenue":          [400e9, 360e9, 330e9, 300e9],
        "Net Income":             [100e9,  90e9,  80e9,  70e9],
        "Gross Profit":           [180e9, 160e9, 145e9, 130e9],
        "Operating Income":       [ 90e9,  80e9,  72e9,  64e9],
        "EBITDA":                 [120e9, 110e9, 100e9,  90e9],
        "Diluted EPS":            [  8.0,   7.0,   6.0,   5.0],
        "Diluted Average Shares": [970e6, 980e6, 990e6, 1000e6],
    },
)


class TestDeepdiveGrowth(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def _run(self, ticker="GRW"):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _GROWTH_INCOME_DF
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        with patch("app.get_info", return_value=_BASE_INFO), \
             patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.dividend_growth", return_value={
                 "cagr_3y": 5.0, "cagr_5y": 4.0, "annual": []}):
            return app.deepdive(ticker)

    def test_last_three_years_only(self):
        rows = self._run()["growth"]
        # 4 statement years -> 3 YoY periods (2023, 2024, 2025)
        self.assertEqual([r["period"] for r in rows], ["2023", "2024", "2025"])

    def test_each_row_has_all_series(self):
        rows = self._run()["growth"]
        for r in rows:
            for k in ("revenue_growth", "eps_growth", "ebitda_growth"):
                self.assertIn(k, r, f"Growth row missing: {k}")

    def test_shares_growth_removed_from_growth(self):
        # Shares now live in the Share Dilution panel, not the growth panel.
        for r in self._run()["growth"]:
            self.assertNotIn("shares_growth", r)

    def test_revenue_growth_value(self):
        rows = self._run()["growth"]
        by_year = {r["period"]: r for r in rows}
        # 2025: 400/360 - 1 = 11.11%
        self.assertAlmostEqual(by_year["2025"]["revenue_growth"], (400 / 360 - 1) * 100)
        # 2024: 360/330 - 1 = 9.09%
        self.assertAlmostEqual(by_year["2024"]["revenue_growth"], (360 / 330 - 1) * 100)

    def test_eps_growth_value(self):
        rows = self._run()["growth"]
        by_year = {r["period"]: r for r in rows}
        # EPS 7.0 -> 8.0 in 2025 => 14.29%
        self.assertAlmostEqual(by_year["2025"]["eps_growth"], (8.0 / 7.0 - 1) * 100)

    def test_operating_income_bar_amount_and_margin(self):
        # Operating Margin bar = operating income $; tooltip % from operating_margin.
        latest = self._run()["revenue_net_income"][-1]   # 2025
        self.assertAlmostEqual(latest["operating_income"], 90e9)
        self.assertAlmostEqual(latest["operating_margin"], 90e9 / 400e9 * 100)
        self.assertAlmostEqual(latest["gross_margin"], 180e9 / 400e9 * 100)

    def test_growth_empty_when_single_year(self):
        one_year = _make_df(
            ["Total Revenue"], ["2024-12-31"], {"Total Revenue": [400e9]})
        mock_tk = MagicMock()
        mock_tk.income_stmt = one_year
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        with patch("app.get_info", return_value=_BASE_INFO), \
             patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.dividend_growth", return_value={
                 "cagr_3y": None, "cagr_5y": None, "annual": []}):
            d = app.deepdive("ONEYR")
        # no prior year -> no YoY rows
        self.assertEqual(d["growth"], [])


# ---------------------------------------------------------------------------
# 8c. _dividend_yields & share_dilution
# ---------------------------------------------------------------------------
class TestDividendYields(unittest.TestCase):

    def test_yield_per_year(self):
        divs = pd.Series(
            [0.5, 0.5, 0.6, 0.6],
            index=pd.DatetimeIndex(["2023-03-01", "2023-09-01",
                                    "2024-03-01", "2024-09-01"]),
        )
        prices = pd.Series(
            [90.0, 100.0, 110.0, 120.0],
            index=pd.DatetimeIndex(["2023-06-01", "2023-12-31",
                                    "2024-06-01", "2024-12-31"]),
        )
        out = app._dividend_yields(divs, prices)
        # 2023: 1.0 div / 100 year-end = 1.0% ; 2024: 1.2 / 120 = 1.0%
        self.assertAlmostEqual(out[2023], 1.0)
        self.assertAlmostEqual(out[2024], 1.0)

    def test_empty_inputs(self):
        self.assertEqual(app._dividend_yields(None, None), {})
        self.assertEqual(app._dividend_yields(pd.Series([], dtype=float),
                                              pd.Series([], dtype=float)), {})


_DIL_BAL = _make_df(
    ["Ordinary Shares Number", "Treasury Shares Number", "Total Assets"],
    ["2024-12-31", "2023-12-31"],
    {
        "Ordinary Shares Number":  [1000.0, 1050.0],   # shrinking -> buybacks
        "Treasury Shares Number":  [200.0, 150.0],
        "Total Assets":            [5000.0, 4800.0],
    },
)
_DIL_INC = _make_df(
    ["Net Income", "Diluted EPS"], ["2024-12-31", "2023-12-31"],
    {"Net Income": [500.0, 400.0], "Diluted EPS": [5.0, 4.0]})
_DIL_CF = _make_df(
    ["Cash Dividends Paid"], ["2024-12-31", "2023-12-31"],
    {"Cash Dividends Paid": [-100.0, -80.0]})


class TestSharesOutstandingSeries(unittest.TestCase):
    """The one place share counts are resolved — every caller depends on it
    returning outstanding shares, never issued, whenever that is knowable."""

    def test_prefers_ordinary_shares_row(self):
        series, basis = app._shares_outstanding_series(_DIL_BAL)
        self.assertEqual(basis, "outstanding")
        self.assertEqual([s["value"] for s in series], [1050.0, 1000.0])

    def test_derives_outstanding_from_issued_minus_treasury(self):
        bal = _make_df(["Share Issued", "Treasury Shares Number"],
                       ["2024-12-31", "2023-12-31"],
                       {"Share Issued": [1200.0, 1200.0],
                        "Treasury Shares Number": [200.0, 150.0]})
        series, basis = app._shares_outstanding_series(bal)
        self.assertEqual(basis, "derived")
        self.assertEqual([s["value"] for s in series], [1050.0, 1000.0])

    def test_missing_treasury_value_nets_nothing_for_that_period(self):
        bal = _make_df(["Share Issued", "Treasury Shares Number"],
                       ["2024-12-31", "2023-12-31"],
                       {"Share Issued": [1200.0, 1200.0],
                        "Treasury Shares Number": [200.0, float("nan")]})
        series, basis = app._shares_outstanding_series(bal)
        self.assertEqual(basis, "derived")
        self.assertEqual([s["value"] for s in series], [1200.0, 1000.0])

    def test_issued_without_treasury_row_stays_issued(self):
        bal = _make_df(["Share Issued"], ["2024-12-31"], {"Share Issued": [1200.0]})
        series, basis = app._shares_outstanding_series(bal)
        self.assertEqual(basis, "issued")
        self.assertEqual([s["value"] for s in series], [1200.0])

    def test_no_share_rows(self):
        bal = _make_df(["Total Assets"], ["2024-12-31"], {"Total Assets": [1.0]})
        self.assertEqual(app._shares_outstanding_series(bal), ([], None))


class TestShareDilution(unittest.TestCase):

    def _run(self, bal=None):
        with patch("app.get_dividends", return_value=None), \
             patch("app.get_raw_close", return_value=None):
            rows, _ = app.share_dilution(_DIL_INC, bal if bal is not None else _DIL_BAL,
                                         _DIL_CF, "TST")
            return rows

    def _basis(self, bal):
        with patch("app.get_dividends", return_value=None), \
             patch("app.get_raw_close", return_value=None):
            return app.share_dilution(_DIL_INC, bal, _DIL_CF, "TST")[1]

    def test_returns_rows_with_all_series(self):
        rows = self._run()
        self.assertTrue(rows)
        for k in ("period", "shares_outstanding", "treasury_shares", "eps",
                  "div_yield", "payout_ratio"):
            self.assertIn(k, rows[0])

    def test_eps_per_year(self):
        by_year = {r["period"]: r for r in self._run()}
        self.assertAlmostEqual(by_year["2024"]["eps"], 5.0)
        self.assertAlmostEqual(by_year["2023"]["eps"], 4.0)

    def test_shares_and_treasury_values(self):
        by_year = {r["period"]: r for r in self._run()}
        self.assertAlmostEqual(by_year["2024"]["shares_outstanding"], 1000.0)
        self.assertAlmostEqual(by_year["2024"]["treasury_shares"], 200.0)

    def test_payout_ratio_computed(self):
        by_year = {r["period"]: r for r in self._run()}
        # |−100| / 500 = 20% ; |−80| / 400 = 20%
        self.assertAlmostEqual(by_year["2024"]["payout_ratio"], 20.0)
        self.assertAlmostEqual(by_year["2023"]["payout_ratio"], 20.0)

    def test_basis_is_outstanding_when_ordinary_shares_present(self):
        self.assertEqual(self._basis(_DIL_BAL), "outstanding")

    def test_outstanding_derived_when_only_issued_and_treasury(self):
        # No "Ordinary Shares Number", but treasury is there to net off, so the
        # count is reconstructed exactly (issued − treasury) rather than left as
        # issued. 1200 − 200 = 1000 == the real outstanding count.
        bal = _make_df(
            ["Share Issued", "Treasury Shares Number"], ["2024-12-31", "2023-12-31"],
            {"Share Issued": [1200.0, 1200.0], "Treasury Shares Number": [200.0, 150.0]})
        self.assertEqual(self._basis(bal), "derived")
        rows = self._run(bal)
        self.assertAlmostEqual(rows[-1]["shares_outstanding"], 1000.0)
        self.assertAlmostEqual(rows[0]["shares_outstanding"], 1050.0)

    def test_basis_is_issued_when_treasury_row_absent(self):
        # Only issued shares and nothing to net off: the count may include
        # treasury, so it stays raw and the basis warns rather than pretending.
        bal = _make_df(["Share Issued"], ["2024-12-31", "2023-12-31"],
                       {"Share Issued": [1200.0, 1200.0]})
        self.assertEqual(self._basis(bal), "issued")
        self.assertAlmostEqual(self._run(bal)[-1]["shares_outstanding"], 1200.0)

    def test_basis_reported_even_when_no_share_rows(self):
        bal = _make_df(["Total Assets"], ["2024-12-31"], {"Total Assets": [1.0]})
        with patch("app.get_dividends", return_value=None), \
             patch("app.get_raw_close", return_value=None):
            self.assertEqual(
                app.share_dilution(_DIL_INC, bal, _DIL_CF, "TST"), ([], None))


# ---------------------------------------------------------------------------
# 8d. 5-year scope for revenue_net_income & growth panels
# ---------------------------------------------------------------------------
class TestFiveYearScope(unittest.TestCase):

    # 6 fiscal years so the [-5:] cap is actually exercised.
    _INC6 = _make_df(
        ["Total Revenue", "Net Income", "Gross Profit"],
        ["2025-12-31", "2024-12-31", "2023-12-31",
         "2022-12-31", "2021-12-31", "2020-12-31"],
        {
            "Total Revenue": [600e9, 550e9, 500e9, 450e9, 400e9, 360e9],
            "Net Income":    [150e9, 130e9, 120e9, 100e9, 90e9, 80e9],
            "Gross Profit":  [300e9, 270e9, 250e9, 220e9, 200e9, 180e9],
        },
    )

    def _run(self):
        mock_tk = MagicMock()
        mock_tk.income_stmt = self._INC6
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        with patch("app.get_info", return_value=_BASE_INFO), \
             patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.dividend_growth", return_value={
                 "cagr_3y": None, "cagr_5y": None, "annual": []}):
            return app.deepdive("SIXYR")

    def test_revenue_net_income_capped_at_5(self):
        rows = self._run()["revenue_net_income"]
        self.assertEqual(len(rows), 5)
        self.assertEqual([r["period"] for r in rows],
                         ["2021", "2022", "2023", "2024", "2025"])

    def test_growth_capped_at_5(self):
        rows = self._run()["growth"]
        # 6 years -> 5 YoY periods, all kept by the [-5:] cap
        self.assertEqual(len(rows), 5)
        self.assertEqual([r["period"] for r in rows],
                         ["2021", "2022", "2023", "2024", "2025"])

    def test_share_dilution_key_present(self):
        self.assertIn("share_dilution", self._run())


# ---------------------------------------------------------------------------
# 8b. history() — OHLC candles, per-range intervals, server-side SMAs
# ---------------------------------------------------------------------------
class TestHistory(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    @staticmethod
    def _ohlc_df(periods, freq="B", end="2026-07-15"):
        idx = pd.bdate_range(end=end, periods=periods) if freq == "B" \
            else pd.date_range(end=end, periods=periods, freq=freq)
        close = pd.Series(np.linspace(100, 150, periods), index=idx)
        return pd.DataFrame({
            "Open": close - 0.5, "High": close + 1.0,
            "Low": close - 1.0, "Close": close,
            "Volume": np.full(periods, 1e6),
        })

    def _run(self, rng, df, ticker="HTST"):
        with patch("app.yf.Ticker") as T:
            T.return_value.history.return_value = df
            out = app.history(ticker, rng)
            call = T.return_value.history.call_args
        return out, call.kwargs

    def test_daily_range_fetches_lookback_and_trims(self):
        # 1y visible window: fetched at 2y daily so the SMA-200 is complete
        # from the first visible bar, then trimmed back to ~1y of points.
        df = self._ohlc_df(504)  # ~2 trading years
        out, kw = self._run("1y", df, "HT1")
        self.assertEqual(kw["period"], "2y")
        self.assertEqual(kw["interval"], "1d")
        self.assertEqual(out["interval"], "1d")
        pts = out["points"]
        self.assertLess(len(pts), 504)
        first = datetime.date.fromisoformat(pts[0]["date"])
        last = datetime.date.fromisoformat(pts[-1]["date"])
        self.assertLessEqual((last - first).days, 366)
        # SMA-200 already defined on the very first visible bar
        self.assertIsNotNone(pts[0]["sma200"])

    def test_sma_matches_rolling_mean(self):
        df = self._ohlc_df(504)
        out, _ = self._run("1y", df, "HT2")
        expected = df["Close"].tail(20).mean()
        self.assertAlmostEqual(out["points"][-1]["sma20"], expected, places=6)

    def test_2y_uses_weekly_bars(self):
        df = self._ohlc_df(520, freq="W-FRI")
        out, kw = self._run("2y", df, "HT3")
        self.assertEqual(kw["period"], "10y")
        self.assertEqual(kw["interval"], "1wk")
        self.assertEqual(out["interval"], "1wk")

    def test_max_uses_monthly_bars_untrimmed(self):
        df = self._ohlc_df(240, freq="ME")
        out, kw = self._run("max", df, "HT4")
        self.assertEqual(kw["period"], "max")
        self.assertEqual(kw["interval"], "1mo")
        self.assertEqual(len(out["points"]), 240)

    def test_short_history_smas_none_not_crash(self):
        df = self._ohlc_df(10)
        out, _ = self._run("1mo", df, "HT5")
        self.assertEqual(len(out["points"]), 10)
        self.assertIsNone(out["points"][0]["sma20"])
        self.assertIsNone(out["points"][-1]["sma200"])

    def test_close_only_frame_keeps_close_and_none_ohlc(self):
        # Callers that only read `close` (screener correlation, exports) must
        # keep working even if a data source lacks OHLC columns.
        idx = pd.bdate_range(end="2026-07-15", periods=5)
        df = pd.DataFrame({"Close": np.arange(5, dtype=float) + 100,
                           "Volume": np.full(5, 1e6)}, index=idx)
        out, _ = self._run("1mo", df, "HT6")
        p = out["points"][0]
        self.assertEqual(p["close"], 100.0)
        self.assertIsNone(p["open"])
        self.assertIsNone(p["high"])
        self.assertIsNone(p["low"])


# ---------------------------------------------------------------------------
# 9. financials()
# ---------------------------------------------------------------------------
class TestFinancials(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def test_income_annual_structure(self):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        with patch("app.yf.Ticker", return_value=mock_tk):
            result = app.financials("TST", "income", "annual")
        self.assertEqual(result["stmt"], "income")
        self.assertEqual(result["freq"], "annual")
        # Periods should be newest-first (2024 before 2023)
        self.assertGreater(result["periods"][0], result["periods"][1])
        self.assertGreater(len(result["rows"]), 0)
        # Each row has label + values list
        row = result["rows"][0]
        self.assertIn("label", row)
        self.assertIn("values", row)
        self.assertEqual(len(row["values"]), len(result["periods"]))

    def test_empty_df_returns_empty_rows(self):
        mock_tk = MagicMock()
        mock_tk.balance_sheet = pd.DataFrame()
        with patch("app.yf.Ticker", return_value=mock_tk):
            result = app.financials("TST", "balance", "annual")
        self.assertEqual(result["rows"], [])
        self.assertEqual(result["periods"], [])

    def test_cashflow_quarterly(self):
        mock_tk = MagicMock()
        mock_tk.quarterly_cash_flow = _CF_DF
        with patch("app.yf.Ticker", return_value=mock_tk):
            result = app.financials("TST", "cashflow", "quarterly")
        self.assertEqual(result["freq"], "quarterly")
        self.assertGreater(len(result["rows"]), 0)

    def test_nan_values_become_none(self):
        nan_df = _make_df(["Total Revenue"], ["2024-12-31"], {"Total Revenue": [float("nan")]})
        mock_tk = MagicMock()
        mock_tk.income_stmt = nan_df
        with patch("app.yf.Ticker", return_value=mock_tk):
            result = app.financials("TST", "income", "annual")
        # no quarterly data on the mock -> no TTM column, single annual value
        self.assertIsNone(result["rows"][0]["values"][0])
        self.assertEqual(result["periods"], ["2024-12-31"])


# A 4-quarter income statement so a full TTM can be summed.
_QINC_DF = _make_df(
    ["Total Revenue", "Net Income"],
    ["2024-12-31", "2024-09-30", "2024-06-30", "2024-03-31", "2023-12-31"],
    {
        "Total Revenue": [110e9, 100e9, 95e9, 90e9, 85e9],
        "Net Income":    [ 30e9,  28e9, 26e9, 24e9, 22e9],
    },
)

_QBAL_DF = _make_df(
    ["Stockholders Equity", "Total Debt"],
    ["2025-03-31", "2024-12-31"],
    {
        "Stockholders Equity": [55e9, 50e9],
        "Total Debt":          [33e9, 30e9],
    },
)


class TestTTMColumn(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def _tk(self):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.quarterly_income_stmt = _QINC_DF
        mock_tk.balance_sheet = _BAL_DF
        mock_tk.quarterly_balance_sheet = _QBAL_DF
        return mock_tk

    def test_income_has_ttm_first_column(self):
        with patch("app.yf.Ticker", return_value=self._tk()):
            f = app.financials("TST", "income", "annual")
        self.assertEqual(f["periods"][0], "TTM")
        rev = next(r for r in f["rows"] if r["label"] == "Total Revenue")
        # TTM = sum of the 4 newest quarters: 110+100+95+90 = 395e9
        self.assertAlmostEqual(rev["values"][0], 395e9)
        # remaining columns keep their annual values, newest-first
        self.assertAlmostEqual(rev["values"][1], 400e9)
        self.assertEqual(len(rev["values"]), len(f["periods"]))

    def test_ttm_excludes_oldest_quarter(self):
        # the 5th (oldest) quarter must not be included in the 4-quarter sum
        with patch("app.yf.Ticker", return_value=self._tk()):
            f = app.financials("TST", "income", "annual")
        ni = next(r for r in f["rows"] if r["label"] == "Net Income")
        self.assertAlmostEqual(ni["values"][0], 30e9 + 28e9 + 26e9 + 24e9)

    def test_balance_uses_mrq_snapshot(self):
        with patch("app.yf.Ticker", return_value=self._tk()):
            f = app.financials("TST", "balance", "annual")
        self.assertEqual(f["periods"][0], "MRQ")
        eq = next(r for r in f["rows"] if r["label"] == "Stockholders Equity")
        # MRQ = most-recent quarter only (not a sum): 55e9
        self.assertAlmostEqual(eq["values"][0], 55e9)

    def test_no_ttm_when_under_four_quarters(self):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        # only 2 quarters available -> no TTM column
        mock_tk.quarterly_income_stmt = _make_df(
            ["Total Revenue"], ["2024-12-31", "2024-09-30"],
            {"Total Revenue": [110e9, 100e9]})
        with patch("app.yf.Ticker", return_value=mock_tk):
            f = app.financials("TST", "income", "annual")
        self.assertNotIn("TTM", f["periods"])

    def test_no_ttm_when_no_quarterly_data(self):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.quarterly_income_stmt = pd.DataFrame()
        with patch("app.yf.Ticker", return_value=mock_tk):
            f = app.financials("TST", "income", "annual")
        self.assertNotIn("TTM", f["periods"])
        self.assertEqual(f["periods"][0], "2024-12-31")

    def test_income_adds_yahoo_ebitda_row(self):
        with patch("app.yf.Ticker", return_value=self._tk()), \
             patch("app.get_info", return_value={"ebitda": 165e9}):
            f = app.financials("TST", "income", "annual")
        yrow = next((r for r in f["rows"] if r["label"] == "EBITDA (Yahoo TTM)"), None)
        self.assertIsNotNone(yrow)
        # Yahoo's TTM scalar populates the TTM (first) column only
        self.assertEqual(yrow["values"][0], 165e9)
        self.assertTrue(all(v is None for v in yrow["values"][1:]))

    def test_yahoo_ebitda_row_only_income(self):
        with patch("app.yf.Ticker", return_value=self._tk()), \
             patch("app.get_info", return_value={"ebitda": 165e9}):
            b = app.financials("TST", "balance", "annual")
        self.assertNotIn("EBITDA (Yahoo TTM)", [r["label"] for r in b["rows"]])

    def test_no_yahoo_ebitda_row_without_ttm(self):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.quarterly_income_stmt = pd.DataFrame()   # no TTM column
        with patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.get_info", return_value={"ebitda": 165e9}):
            f = app.financials("TST", "income", "annual")
        self.assertNotIn("EBITDA (Yahoo TTM)", [r["label"] for r in f["rows"]])


# ---------------------------------------------------------------------------
# 10. Excel workbook
# ---------------------------------------------------------------------------
@unittest.skipUnless(app._HAS_OPENPYXL, "openpyxl not installed")
class TestWorkbook(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def _build(self, tickers=("TST", "TST2")):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        close = pd.Series([100.0, 105.0], index=pd.DatetimeIndex(["2025-01-02", "2025-01-03"]))
        hist_df = pd.DataFrame({"Close": close, "Volume": [1e6, 1e6]})
        mock_tk.history.return_value = hist_df
        with patch("app.get_info", return_value=_BASE_INFO), \
             patch("app._get_stmt", side_effect=lambda t, attr: {
                 "income_stmt": _INCOME_DF, "balance_sheet": _BAL_DF,
                 "cash_flow": _CF_DF}.get(attr)), \
             patch("app.get_dividends", return_value=_DIVS), \
             patch("app.get_raw_close", return_value=_CLOSE_SERIES), \
             patch("app.performance", return_value={"ytd": 1.0, "1y": 2.0, "3y": 3.0, "5y": 4.0, "10y": 5.0}), \
             patch("app.yf.Ticker", return_value=mock_tk):
            raw = app.build_workbook(list(tickers))
        from openpyxl import load_workbook
        return load_workbook(io.BytesIO(raw))

    def test_sheet_names(self):
        wb = self._build()
        self.assertEqual(wb.sheetnames, ["Metrics", "Price History", "Financials"])

    def test_metrics_header_count(self):
        wb = self._build()
        ws = wb["Metrics"]
        headers = [c.value for c in ws[1]]
        self.assertEqual(len(headers), len(app._METRIC_COLS))

    def test_metrics_has_data_rows(self):
        wb = self._build(("TST",))
        ws = wb["Metrics"]
        self.assertEqual(ws.max_row, 2)   # 1 header + 1 data row

    def test_metrics_ticker_column(self):
        wb = self._build(("TST",))
        ws = wb["Metrics"]
        self.assertEqual(ws.cell(row=2, column=1).value, "TST")

    def test_price_history_header(self):
        wb = self._build(("TST", "TST2"))
        ws = wb["Price History"]
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers[0], "Date")
        self.assertIn("TST", headers)
        self.assertIn("TST2", headers)

    def test_financials_sheet_has_content(self):
        wb = self._build(("TST",))
        ws = wb["Financials"]
        self.assertGreater(ws.max_row, 2)

    def test_frozen_panes_metrics(self):
        wb = self._build()
        self.assertEqual(wb["Metrics"].freeze_panes, "A2")

    def test_frozen_panes_price_history(self):
        wb = self._build()
        self.assertEqual(wb["Price History"].freeze_panes, "B2")


# ---------------------------------------------------------------------------
# 11. HTTP server integration tests
# ---------------------------------------------------------------------------
def _start_server():
    """Start a real ThreadingHTTPServer on an ephemeral port. Returns (httpd, url)."""
    httpd = ThreadingHTTPServer(("127.0.0.1", 0), app.Handler)
    port = httpd.server_address[1]
    t = threading.Thread(target=httpd.serve_forever, daemon=True)
    t.start()
    return httpd, f"http://127.0.0.1:{port}"


def _get(url):
    with urlopen(url, timeout=10) as r:
        return r.status, r.read(), r.headers.get("Content-Type", "")


def _post(url, body_dict):
    data = json.dumps(body_dict).encode()
    import urllib.request
    req = urllib.request.Request(url, data=data,
                                 headers={"Content-Type": "application/json"})
    with urlopen(req, timeout=10) as r:
        return r.status, r.read(), r.headers.get("Content-Type", "")


class TestHTTPServer(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        app.clear_cache()
        cls.httpd, cls.base = _start_server()
        # Patch all network calls once for the entire server lifecycle
        cls._patches = [
            patch("app.get_info", return_value=_BASE_INFO),
            patch("app._get_stmt", side_effect=lambda t, attr: {
                "income_stmt": _INCOME_DF, "balance_sheet": _BAL_DF,
                "cash_flow": _CF_DF}.get(attr)),
            patch("app.get_dividends", return_value=_DIVS),
            patch("app.get_raw_close", return_value=_CLOSE_SERIES),
            patch("app.performance", return_value={"1m": 0.5, "3m": 1.5, "6m": 3.0, "ytd": 1.0, "1y": 2.0, "3y": 3.0, "5y": 4.0, "10y": 5.0}),
        ]
        for p in cls._patches:
            p.start()
        # Also patch deepdive/history/financials yf.Ticker for those endpoints
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        hist_df = pd.DataFrame(
            {"Open": [99.5, 100.5, 101.5], "High": [100.5, 101.5, 102.5],
             "Low": [99.0, 100.0, 101.0], "Close": [100.0, 101.0, 102.0],
             "Volume": [1e6, 1e6, 1e6]},
            index=pd.DatetimeIndex(["2025-06-01", "2025-06-02", "2025-06-03"]),
        )
        mock_tk.history.return_value = hist_df
        mock_tk.dividends = _DIVS
        mock_tk.quarterly_income_stmt = _INCOME_DF
        mock_tk.quarterly_balance_sheet = _BAL_DF
        mock_tk.quarterly_cash_flow = _CF_DF
        cls._yf_patch = patch("app.yf.Ticker", return_value=mock_tk)
        cls._yf_patch.start()
        cls._ddgrowth_patch = patch("app.dividend_growth", return_value={
            "cagr_3y": 5.0, "cagr_5y": 4.0, "annual": []})
        cls._ddgrowth_patch.start()

    @classmethod
    def tearDownClass(cls):
        for p in cls._patches:
            p.stop()
        cls._yf_patch.stop()
        cls._ddgrowth_patch.stop()
        cls.httpd.shutdown()

    def setUp(self):
        app.clear_cache()

    # -- health -------------------------------------------------------
    def test_health_200(self):
        status, body, ct = _get(f"{self.base}/api/health")
        self.assertEqual(status, 200)
        self.assertIn("application/json", ct)
        self.assertEqual(json.loads(body), {"ok": True})

    # -- calendar endpoints -------------------------------------------
    def test_calendar_endpoint(self):
        fake = {"start": None, "end": None,
                "earnings": [{"ticker": "NKE", "date": "2026-06-30"}],
                "splits": []}
        with patch("app.market_calendar", return_value=fake) as m:
            status, body, ct = _get(f"{self.base}/api/calendar?limit=50")
        self.assertEqual(status, 200)
        self.assertIn("application/json", ct)
        self.assertEqual(json.loads(body)["earnings"][0]["ticker"], "NKE")
        # limit parsed and clamped into range
        self.assertEqual(m.call_args.args[2], 50)

    def test_calendar_bad_limit_defaults(self):
        with patch("app.market_calendar", return_value={}) as m:
            _get(f"{self.base}/api/calendar?limit=notanumber")
        self.assertEqual(m.call_args.args[2], 80)

    def test_stock_calendar_endpoint(self):
        fake = {"ticker": "AAPL", "upcoming": {}, "earnings_history": [], "splits": []}
        with patch("app.stock_calendar", return_value=fake) as m:
            status, body, ct = _get(f"{self.base}/api/stock_calendar?ticker=aapl")
        self.assertEqual(status, 200)
        self.assertEqual(json.loads(body)["ticker"], "AAPL")
        m.assert_called_once_with("AAPL")   # uppercased by _parse_tickers

    def test_stock_calendar_requires_ticker(self):
        try:
            status, _, _ = _get(f"{self.base}/api/stock_calendar")
            self.assertEqual(status, 400)
        except HTTPError as e:
            self.assertEqual(e.code, 400)

    # -- refresh=1 cache eviction ---------------------------------------
    def test_deepdive_refresh_evicts_cached_entry(self):
        app.cached("deepdive:AAPL", 600, lambda: {"ticker": "SENTINEL"})
        status, body, _ = _get(f"{self.base}/api/deepdive?ticker=AAPL")
        self.assertEqual(json.loads(body)["ticker"], "SENTINEL")   # cached hit
        status, body, _ = _get(f"{self.base}/api/deepdive?ticker=AAPL&refresh=1")
        self.assertEqual(status, 200)
        self.assertEqual(json.loads(body)["ticker"], "AAPL")       # rebuilt fresh

    def test_screener_refresh_evicts_ticker_keys(self):
        app.cached("deepdive:AAPL", 600, lambda: {"x": 1})
        app.cached("hist:AAPL:1y", 600, lambda: {"x": 2})
        app.cached("info:MSFT", 600, lambda: {"x": 3})
        _get(f"{self.base}/api/screener?tickers=AAPL&refresh=1")
        with app._CACHE_LOCK:
            keys = set(app._CACHE)
        self.assertNotIn("deepdive:AAPL", keys)
        self.assertNotIn("hist:AAPL:1y", keys)
        self.assertIn("info:MSFT", keys)   # other tickers untouched

    def test_screener_without_refresh_keeps_cache(self):
        app.cached("deepdive:AAPL", 600, lambda: {"x": 1})
        _get(f"{self.base}/api/screener?tickers=AAPL")
        with app._CACHE_LOCK:
            self.assertIn("deepdive:AAPL", app._CACHE)

    def test_calendar_refresh_evicts_mktcal_keys(self):
        app.cached("mktcal:a:b:80", 600, lambda: {})
        app.cached("info:AAPL", 600, lambda: {})
        with patch("app.market_calendar", return_value={}):
            _get(f"{self.base}/api/calendar?refresh=1")
        with app._CACHE_LOCK:
            keys = set(app._CACHE)
        self.assertNotIn("mktcal:a:b:80", keys)
        self.assertIn("info:AAPL", keys)

    # -- static serving -----------------------------------------------
    def test_index_html_served(self):
        status, body, ct = _get(f"{self.base}/")
        self.assertEqual(status, 200)
        self.assertIn("text/html", ct)
        self.assertIn(b"<!DOCTYPE html>", body)

    def test_static_files_not_browser_cached(self):
        # Static responses must carry Cache-Control so browsers revalidate
        # instead of heuristically caching stale JS/CSS across app updates.
        for path in ("/", "/js/views.js", "/css/styles.css"):
            with urlopen(f"{self.base}{path}", timeout=10) as r:
                self.assertEqual(r.headers.get("Cache-Control"), "no-cache",
                                 f"missing no-cache on {path}")

    def test_js_file_served(self):
        status, body, ct = _get(f"{self.base}/js/api.js")
        self.assertEqual(status, 200)
        self.assertIn("javascript", ct)

    def test_css_file_served(self):
        status, body, ct = _get(f"{self.base}/css/styles.css")
        self.assertEqual(status, 200)
        self.assertIn("text/css", ct)

    def test_path_traversal_blocked(self):
        # Attempt to escape STATIC_DIR
        try:
            status, _, _ = _get(f"{self.base}/../app.py")
            self.assertEqual(status, 404)
        except HTTPError as e:
            self.assertEqual(e.code, 404)

    def test_unknown_static_404(self):
        try:
            _get(f"{self.base}/does-not-exist.txt")
            self.fail("Expected 404")
        except HTTPError as e:
            self.assertEqual(e.code, 404)

    # -- screener -----------------------------------------------------
    def test_screener_returns_rows(self):
        status, body, _ = _get(f"{self.base}/api/screener?tickers=TST")
        self.assertEqual(status, 200)
        d = json.loads(body)
        self.assertIn("rows", d)
        self.assertEqual(len(d["rows"]), 1)

    def test_screener_multiple_tickers(self):
        status, body, _ = _get(f"{self.base}/api/screener?tickers=TST,TST2,TST3")
        d = json.loads(body)
        self.assertEqual(len(d["rows"]), 3)

    def test_screener_empty_returns_empty(self):
        status, body, _ = _get(f"{self.base}/api/screener")
        self.assertEqual(status, 200)
        self.assertEqual(json.loads(body), {"rows": []})

    def test_screener_deduplicates_tickers(self):
        status, body, _ = _get(f"{self.base}/api/screener?tickers=TST,TST,TST")
        d = json.loads(body)
        self.assertEqual(len(d["rows"]), 1)

    def test_screener_row_has_all_fields(self):
        _, body, _ = _get(f"{self.base}/api/screener?tickers=TST")
        row = json.loads(body)["rows"][0]
        for field in ("pe", "forward_pe", "peg", "pb", "ps", "pc", "p_fcf", "ev_ebitda",
                      "dcf_value", "dcf_upside",
                      "eps", "income", "profit_margin", "fcf", "roa", "roe", "roic",
                      "debt_to_equity", "lt_debt_to_equity", "div_yield", "five_year_avg_yield",
                      "dividend_estimate", "dividend_ttm", "payout_ratio", "fcf_coverage",
                      "years_div_increase", "perf_ytd", "perf_1y", "perf_3y", "perf_5y", "perf_10y"):
            self.assertIn(field, row, f"Screener row missing: {field}")

    def test_screener_industry_field(self):
        _, body, _ = _get(f"{self.base}/api/screener?tickers=TST")
        row = json.loads(body)["rows"][0]
        self.assertEqual(row["industry"], "Software")

    # -- deepdive -----------------------------------------------------
    def test_deepdive_200(self):
        status, body, _ = _get(f"{self.base}/api/deepdive?ticker=TST")
        self.assertEqual(status, 200)
        d = json.loads(body)
        self.assertIn("panels", d)

    def test_deepdive_no_ticker_400(self):
        try:
            _get(f"{self.base}/api/deepdive")
            self.fail("Expected 400")
        except HTTPError as e:
            self.assertEqual(e.code, 400)

    def test_deepdive_health_panel_has_total_equity(self):
        _, body, _ = _get(f"{self.base}/api/deepdive?ticker=TST")
        d = json.loads(body)
        self.assertIn("Total Equity", d["panels"]["health"])

    def test_deepdive_dividend_fcf_coverage_position(self):
        _, body, _ = _get(f"{self.base}/api/deepdive?ticker=TST")
        d = json.loads(body)
        keys = list(d["panels"]["dividend"].keys())
        self.assertEqual(keys.index("FCF Coverage"), keys.index("Payout Ratio %") + 1)

    def test_deepdive_revenue_net_income_four_series(self):
        _, body, _ = _get(f"{self.base}/api/deepdive?ticker=TST")
        d = json.loads(body)
        row = d["revenue_net_income"][0]
        for k in ("revenue", "gross_profit", "net_income", "fcf"):
            self.assertIn(k, row)

    # -- history ------------------------------------------------------
    def test_history_200(self):
        status, body, _ = _get(f"{self.base}/api/history?ticker=TST&range=1mo")
        self.assertEqual(status, 200)
        d = json.loads(body)
        self.assertIn("points", d)
        self.assertGreater(len(d["points"]), 0)

    def test_history_no_ticker_400(self):
        try:
            _get(f"{self.base}/api/history")
            self.fail("Expected 400")
        except HTTPError as e:
            self.assertEqual(e.code, 400)

    def test_history_point_schema(self):
        _, body, _ = _get(f"{self.base}/api/history?ticker=TST&range=1mo")
        d = json.loads(body)
        self.assertEqual(d["interval"], "1d")
        point = d["points"][0]
        for k in ("date", "open", "high", "low", "close", "volume",
                  "sma20", "sma50", "sma200"):
            self.assertIn(k, point)
        self.assertEqual(point["open"], 99.5)
        self.assertEqual(point["high"], 100.5)
        self.assertEqual(point["low"], 99.0)
        self.assertEqual(point["close"], 100.0)

    # -- financials ---------------------------------------------------
    def test_financials_income_200(self):
        status, body, _ = _get(f"{self.base}/api/financials?ticker=TST&stmt=income&freq=annual")
        self.assertEqual(status, 200)
        d = json.loads(body)
        self.assertGreater(len(d["rows"]), 0)

    def test_financials_no_ticker_400(self):
        try:
            _get(f"{self.base}/api/financials?stmt=income")
            self.fail("Expected 400")
        except HTTPError as e:
            self.assertEqual(e.code, 400)

    def test_financials_balance_quarterly(self):
        _, body, _ = _get(f"{self.base}/api/financials?ticker=TST&stmt=balance&freq=quarterly")
        d = json.loads(body)
        self.assertEqual(d["stmt"], "balance")
        self.assertEqual(d["freq"], "quarterly")

    def test_financials_periods_newest_first(self):
        _, body, _ = _get(f"{self.base}/api/financials?ticker=TST&stmt=income&freq=annual")
        d = json.loads(body)
        self.assertGreater(d["periods"][0], d["periods"][1])

    # -- cache clear --------------------------------------------------
    def test_cache_clear_post(self):
        status, body, _ = _post(f"{self.base}/api/cache/clear", {})
        self.assertEqual(status, 200)
        self.assertEqual(json.loads(body), {"ok": True})

    # -- export -------------------------------------------------------
    @unittest.skipUnless(app._HAS_OPENPYXL, "openpyxl not installed")
    def test_export_returns_xlsx(self):
        status, body, ct = _post(f"{self.base}/api/export", {"tickers": ["TST"]})
        self.assertEqual(status, 200)
        self.assertIn("spreadsheetml", ct)
        self.assertTrue(len(body) > 1000)

    @unittest.skipUnless(app._HAS_OPENPYXL, "openpyxl not installed")
    def test_export_xlsx_three_sheets(self):
        from openpyxl import load_workbook
        _, body, _ = _post(f"{self.base}/api/export", {"tickers": ["TST"]})
        wb = load_workbook(io.BytesIO(body))
        self.assertEqual(wb.sheetnames, ["Metrics", "Price History", "Financials"])

    def test_export_empty_tickers_400(self):
        try:
            _post(f"{self.base}/api/export", {"tickers": []})
            self.fail("Expected 400")
        except HTTPError as e:
            self.assertEqual(e.code, 400)

    @unittest.skipUnless(app._HAS_OPENPYXL, "openpyxl not installed")
    def test_export_dcf_returns_xlsx(self):
        from openpyxl import load_workbook
        status, body, ct = _post(f"{self.base}/api/export_dcf", {"ticker": "TST"})
        self.assertEqual(status, 200)
        self.assertIn("spreadsheetml", ct)
        wb = load_workbook(io.BytesIO(body))
        self.assertEqual(wb.sheetnames, ["DCF Valuation"])

    def test_export_dcf_no_ticker_400(self):
        try:
            _post(f"{self.base}/api/export_dcf", {})
            self.fail("Expected 400")
        except HTTPError as e:
            self.assertEqual(e.code, 400)

    def test_unknown_post_endpoint_404(self):
        try:
            _post(f"{self.base}/api/unknown", {})
            self.fail("Expected 404")
        except HTTPError as e:
            self.assertEqual(e.code, 404)


# ---------------------------------------------------------------------------
# 12. Calendars — earnings & stock splits
# ---------------------------------------------------------------------------
class TestTsToIso(unittest.TestCase):
    def test_none(self):
        self.assertIsNone(app._ts_to_iso(None))

    def test_nat(self):
        self.assertIsNone(app._ts_to_iso(pd.NaT))

    def test_timestamp(self):
        self.assertEqual(app._ts_to_iso(pd.Timestamp("2026-07-30 20:00:00+0000")),
                         "2026-07-30")

    def test_date(self):
        self.assertEqual(app._ts_to_iso(datetime.date(2026, 5, 11)), "2026-05-11")


_EARN_CAL_DF = pd.DataFrame(
    {
        "Company": ["NIKE, Inc."],
        "Marketcap": [6.03e10],
        "Event Name": ["Q4 2026 Earnings Announcement"],
        "Event Start Date": [pd.Timestamp("2026-06-30 20:00:00+0000")],
        "Timing": ["AMC"],
        "EPS Estimate": [0.12],
        "Reported EPS": [np.nan],
        "Surprise(%)": [np.nan],
    },
    index=pd.Index(["NKE"], name="Symbol"),
)

_SPLIT_CAL_DF = pd.DataFrame(
    {
        "Company": ["Dhautoware Co Ltd"],
        "Payable On": [pd.Timestamp("2026-07-03 04:00:00+0000")],
        "Optionable": [False],
        "Old Share Worth": [5],
        "Share Worth": [1],
    },
    index=pd.Index(["025440.KQ"], name="Symbol"),
)


class TestMarketCalendar(unittest.TestCase):
    def setUp(self):
        app.clear_cache()

    def _mock_calendars(self):
        m = MagicMock()
        m.get_earnings_calendar.return_value = _EARN_CAL_DF
        m.get_splits_calendar.return_value = _SPLIT_CAL_DF
        return m

    def test_shapes_and_values(self):
        with patch("app.yf.Calendars", return_value=self._mock_calendars()):
            out = app.market_calendar("2026-06-26", "2026-07-10", 80)
        self.assertEqual(len(out["earnings"]), 1)
        self.assertEqual(len(out["splits"]), 1)
        e = out["earnings"][0]
        self.assertEqual(e["ticker"], "NKE")
        self.assertEqual(e["date"], "2026-06-30")
        self.assertEqual(e["timing"], "AMC")
        self.assertEqual(e["eps_estimate"], 0.12)
        self.assertIsNone(e["eps_actual"])   # NaN -> None
        s = out["splits"][0]
        self.assertEqual(s["ticker"], "025440.KQ")
        self.assertEqual(s["date"], "2026-07-03")
        self.assertEqual(s["old_share_worth"], 5)
        self.assertEqual(s["share_worth"], 1)
        self.assertFalse(s["optionable"])

    def test_json_serializable(self):
        with patch("app.yf.Calendars", return_value=self._mock_calendars()):
            out = app.market_calendar(None, None, 80)
        json.dumps(out)  # must not raise

    def test_earnings_error_isolated(self):
        m = MagicMock()
        m.get_earnings_calendar.side_effect = RuntimeError("boom")
        m.get_splits_calendar.return_value = _SPLIT_CAL_DF
        with patch("app.yf.Calendars", return_value=m):
            out = app.market_calendar(None, None, 80)
        self.assertIn("earnings_error", out)
        self.assertEqual(out["earnings"], [])
        self.assertEqual(len(out["splits"]), 1)   # splits still returned

    def test_constructor_error(self):
        with patch("app.yf.Calendars", side_effect=RuntimeError("no net")):
            out = app.market_calendar(None, None, 80)
        self.assertIn("error", out)
        self.assertEqual(out["earnings"], [])


class TestStockCalendar(unittest.TestCase):
    def setUp(self):
        app.clear_cache()

    def _mock_ticker(self, with_history=True):
        tk = MagicMock()
        tk.calendar = {
            "Dividend Date": datetime.date(2026, 5, 14),
            "Ex-Dividend Date": datetime.date(2026, 5, 11),
            "Earnings Date": [datetime.date(2026, 7, 30)],
            "Earnings High": 1.99, "Earnings Low": 1.83, "Earnings Average": 1.89,
            "Revenue High": 1.12e11, "Revenue Low": 1.07e11, "Revenue Average": 1.08e11,
        }
        if with_history:
            tk.get_earnings_dates.return_value = pd.DataFrame(
                {"EPS Estimate": [2.97, 1.73], "Reported EPS": [np.nan, 1.54],
                 "Surprise(%)": [np.nan, -10.88]},
                index=pd.DatetimeIndex(["2026-10-30", "2025-07-22"], name="Date"),
            )
        else:
            tk.get_earnings_dates.side_effect = ImportError("lxml missing")
        tk.splits = pd.Series(
            [7.0, 4.0],
            index=pd.DatetimeIndex(["2014-06-09", "2020-08-31"]),
            name="Stock Splits",
        )
        return tk

    def test_upcoming_and_splits(self):
        with patch("app.yf.Ticker", return_value=self._mock_ticker()):
            out = app.stock_calendar("AAPL")
        u = out["upcoming"]
        self.assertEqual(u["earnings_dates"], ["2026-07-30"])
        self.assertEqual(u["eps_avg"], 1.89)
        self.assertEqual(u["ex_dividend_date"], "2026-05-11")
        self.assertEqual(u["dividend_date"], "2026-05-14")
        # splits newest-first
        self.assertEqual([s["date"] for s in out["splits"]],
                         ["2020-08-31", "2014-06-09"])
        self.assertEqual(out["splits"][0]["ratio"], 4.0)

    def test_earnings_history(self):
        with patch("app.yf.Ticker", return_value=self._mock_ticker()):
            out = app.stock_calendar("AAPL")
        self.assertEqual(len(out["earnings_history"]), 2)
        self.assertEqual(out["earnings_history"][0]["eps_estimate"], 2.97)
        self.assertIsNone(out["earnings_history"][0]["eps_actual"])

    def test_earnings_history_degrades(self):
        with patch("app.yf.Ticker", return_value=self._mock_ticker(with_history=False)):
            out = app.stock_calendar("AAPL")
        self.assertEqual(out["earnings_history"], [])
        self.assertIn("earnings_history_error", out)
        # rest of payload still present
        self.assertEqual(out["upcoming"]["earnings_dates"], ["2026-07-30"])

    def test_json_serializable(self):
        with patch("app.yf.Ticker", return_value=self._mock_ticker()):
            out = app.stock_calendar("AAPL")
        json.dumps(out)


# ---------------------------------------------------------------------------
# 16. Expanded screener_row fields
# ---------------------------------------------------------------------------
class TestScreenerRowNewFields(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def _row(self):
        with _patch_all():
            return app.screener_row("TST")

    def test_new_keys_present(self):
        row = self._row()
        for k in ("enterprise_value", "gross_margin", "operating_margin",
                  "ebitda_margin", "roce", "revenue_per_share", "current_ratio",
                  "quick_ratio", "total_cash", "total_debt", "total_equity",
                  "div_growth_3y", "div_growth_5y", "ex_dividend_date",
                  "debt_to_equity_mrq", "debt_ebitda", "ebitda_fcf",
                  "short_interest", "days_to_cover", "altman_z", "piotroski_f"):
            self.assertIn(k, row)

    def test_totals_from_info_and_statements(self):
        row = self._row()
        self.assertEqual(row["total_cash"], 10e9)     # _BASE_INFO totalCash
        self.assertEqual(row["total_debt"], 30e9)     # _BASE_INFO totalDebt
        self.assertEqual(row["total_equity"], 50e9)   # Stockholders Equity (newest)

    def test_profit_margin_is_fraction(self):
        # screener_row keeps yfinance's native fraction for margins
        self.assertAlmostEqual(self._row()["profit_margin"], 0.25)


# ---------------------------------------------------------------------------
# 17. Export unit normalization → decimal fractions
# ---------------------------------------------------------------------------
class TestExportValHelper(unittest.TestCase):

    def test_pct_keys_divided_by_100(self):
        self.assertAlmostEqual(app._export_val("roic", 79.5), 0.795)
        self.assertAlmostEqual(app._export_val("perf_1y", 12.0), 0.12)
        self.assertAlmostEqual(app._export_val("div_yield", 2.0), 0.02)
        self.assertAlmostEqual(app._export_val("debt_to_equity", 50.0), 0.50)

    def test_wacc_normalized_like_roic_roce(self):
        # WACC is stored in percentage points like ROIC/ROCE, so it must export
        # as a decimal fraction too — not a raw percentage.
        self.assertAlmostEqual(app._export_val("wacc", 8.5), 0.085)
        self.assertAlmostEqual(app._export_val("roce", 15.0), 0.15)

    def test_dcf_upside_fraction_but_dcf_value_currency(self):
        # DCF Upside is percentage points -> fraction; DCF Value is a per-share
        # currency amount and must pass through untouched.
        self.assertAlmostEqual(app._export_val("dcf_upside", 25.0), 0.25)
        self.assertEqual(app._export_val("dcf_value", 150.0), 150.0)

    def test_fraction_keys_pass_through(self):
        self.assertEqual(app._export_val("profit_margin", 0.25), 0.25)
        self.assertEqual(app._export_val("roe", 0.20), 0.20)
        self.assertEqual(app._export_val("payout_ratio", 0.40), 0.40)
        self.assertEqual(app._export_val("short_interest", 0.01), 0.01)

    def test_currency_and_multiples_pass_through(self):
        self.assertEqual(app._export_val("market_cap", 1e12), 1e12)
        self.assertEqual(app._export_val("pe", 20.0), 20.0)

    def test_none_stays_none(self):
        self.assertIsNone(app._export_val("roic", None))
        self.assertIsNone(app._export_val("market_cap", None))


class TestDdMetricHelper(unittest.TestCase):

    def test_percent_suffix_label_to_fraction(self):
        self.assertEqual(app._dd_metric("ROE %", 20.0), ("ROE", 0.20))
        self.assertEqual(app._dd_metric("Gross Margin %", 45.0), ("Gross Margin", 0.45))

    def test_debt_equity_special_cased(self):
        self.assertEqual(app._dd_metric("Debt/Equity", 50.0), ("Debt/Equity", 0.50))

    def test_non_rate_pass_through(self):
        self.assertEqual(app._dd_metric("Market Cap", 1e12), ("Market Cap", 1e12))
        self.assertEqual(app._dd_metric("FCF Coverage", 5.0), ("FCF Coverage", 5.0))
        self.assertEqual(app._dd_metric("Ex-Dividend Date", "2026-05-11"),
                         ("Ex-Dividend Date", "2026-05-11"))

    def test_none_value_keeps_relabel(self):
        self.assertEqual(app._dd_metric("ROE %", None), ("ROE", None))


class TestMetricColsExpanded(unittest.TestCase):

    def test_new_keys_present(self):
        keys = {k for k, _ in app._METRIC_COLS}
        for k in ("enterprise_value", "gross_margin", "operating_margin",
                  "ebitda_margin", "roce", "revenue_per_share", "current_ratio",
                  "quick_ratio", "total_cash", "total_debt", "total_equity",
                  "div_growth_3y", "div_growth_5y", "ex_dividend_date",
                  "debt_to_equity_mrq", "debt_ebitda", "ebitda_fcf",
                  "short_interest", "days_to_cover", "altman_z", "piotroski_f",
                  "dcf_value", "dcf_upside"):
            self.assertIn(k, keys)

    def test_ticker_is_first_column(self):
        self.assertEqual(app._METRIC_COLS[0][0], "ticker")

    def test_no_percent_in_any_label(self):
        for _, label in app._METRIC_COLS:
            self.assertFalse(label.endswith("%"), f"{label!r} carries a % suffix")


@unittest.skipUnless(app._HAS_OPENPYXL, "openpyxl not installed")
class TestScreenerExportFractions(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def _build_and_row(self):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        close = pd.Series([100.0, 105.0],
                          index=pd.DatetimeIndex(["2025-01-02", "2025-01-03"]))
        mock_tk.history.return_value = pd.DataFrame({"Close": close, "Volume": [1e6, 1e6]})
        with patch("app.get_info", return_value=_BASE_INFO), \
             patch("app._get_stmt", side_effect=lambda t, attr: {
                 "income_stmt": _INCOME_DF, "balance_sheet": _BAL_DF,
                 "cash_flow": _CF_DF}.get(attr)), \
             patch("app.get_dividends", return_value=_DIVS), \
             patch("app.get_raw_close", return_value=_CLOSE_SERIES), \
             patch("app.performance", return_value={
                 "ytd": 5.0, "1y": 12.0, "3y": 30.0, "5y": 60.0, "10y": 100.0}), \
             patch("app.yf.Ticker", return_value=mock_tk):
            raw = app.build_workbook(["TST"])
            row = app.screener_row("TST")
        from openpyxl import load_workbook
        return load_workbook(io.BytesIO(raw)), row

    def _cells(self, wb):
        ws = wb["Metrics"]
        headers = [c.value for c in ws[1]]
        values = [c.value for c in ws[2]]
        label_for = {k: lab for k, lab in app._METRIC_COLS}
        return dict(zip(headers, values)), label_for

    def test_pct_columns_become_fractions(self):
        wb, row = self._build_and_row()
        cell, label_for = self._cells(wb)
        self.assertAlmostEqual(cell[label_for["perf_1y"]], row["perf_1y"] / 100.0)
        self.assertAlmostEqual(cell[label_for["perf_1y"]], 0.12)
        self.assertAlmostEqual(cell[label_for["roic"]], row["roic"] / 100.0)
        # debt_to_equity is now annual (totalDebt 30e9 / equity 50e9 = 60% -> 0.60)
        self.assertAlmostEqual(cell[label_for["debt_to_equity"]], 0.60)
        # debt_to_equity_mrq is Yahoo's value (50.0% -> 0.50)
        self.assertAlmostEqual(cell[label_for["debt_to_equity_mrq"]], 0.50)

    def test_fraction_columns_unchanged(self):
        wb, row = self._build_and_row()
        cell, label_for = self._cells(wb)
        self.assertEqual(cell[label_for["profit_margin"]], 0.25)
        self.assertEqual(cell[label_for["payout_ratio"]], 0.40)

    def test_currency_columns_unchanged(self):
        wb, _ = self._build_and_row()
        cell, label_for = self._cells(wb)
        self.assertEqual(cell[label_for["market_cap"]], 1e12)

    def test_no_percent_headers(self):
        wb, _ = self._build_and_row()
        headers = [c.value for c in wb["Metrics"][1]]
        self.assertFalse(any(str(h).endswith("%") for h in headers))


@unittest.skipUnless(app._HAS_OPENPYXL, "openpyxl not installed")
class TestDeepdiveExportFractions(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def _build(self):
        div_info = {**_BASE_INFO, "exDividendDate": int(datetime.datetime(
            2026, 5, 11, tzinfo=datetime.timezone.utc).timestamp())}
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        with patch("app.get_info", return_value=div_info), \
             patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.dividend_growth", return_value={
                 "cagr_3y": 5.0, "cagr_5y": 4.0, "annual": []}):
            raw = app.build_deepdive_workbook("TST")
        from openpyxl import load_workbook
        return load_workbook(io.BytesIO(raw))

    def _overview(self, wb):
        out = {}
        for row in wb["Overview"].iter_rows(values_only=True):
            if row[0] and row[1] is not None:
                out[row[0]] = row[1]
        return out

    def test_sheet_names(self):
        wb = self._build()
        self.assertEqual(wb.sheetnames, ["Overview", "Charts Data",
            "Income Statement", "Balance Sheet", "Cash Flow"])

    def test_overview_rates_are_fractions(self):
        ov = self._overview(self._build())
        self.assertAlmostEqual(ov["Profit Margin"], 0.25)
        self.assertAlmostEqual(ov["ROE"], 0.20)
        self.assertAlmostEqual(ov["ROA"], 0.10)
        # annual Debt/Equity (totalDebt 30e9 / equity 50e9 = 60% -> 0.60)
        self.assertAlmostEqual(ov["Debt/Equity"], 0.60)
        # Yahoo's MRQ measure, surfaced separately (50.0% -> 0.50)
        self.assertAlmostEqual(ov["Debt/Equity (MRQ)"], 0.50)

    def test_overview_currency_unchanged(self):
        ov = self._overview(self._build())
        self.assertEqual(ov["Market Cap"], 1e12)

    def test_overview_no_percent_labels(self):
        for row in self._build()["Overview"].iter_rows(values_only=True):
            if isinstance(row[0], str):
                self.assertFalse(row[0].endswith("%"), f"{row[0]!r}")

    def test_charts_margins_are_fractions(self):
        rows = list(self._build()["Charts Data"].iter_rows(values_only=True))
        hi = next(i for i, r in enumerate(rows) if r and r[0] == "Year")
        headers = [c for c in rows[hi] if c]
        self.assertIn("Gross Margin", headers)
        self.assertNotIn("Gross Margin %", headers)
        gm = rows[hi].index("Gross Margin")
        vals = [r[gm] for r in rows[hi + 1:hi + 3]
                if isinstance(r[gm], (int, float))]
        self.assertTrue(vals)
        # 2024 gross margin = 180e9 / 400e9 = 0.45 (fraction, not 45%)
        self.assertTrue(any(abs(v - 0.45) < 1e-6 for v in vals))
        self.assertTrue(all(abs(v) < 1.5 for v in vals))

    def test_consistent_with_screener_export(self):
        # Net/Profit margin reads identically (0.25) in both workbooks.
        ov = self._overview(self._build())
        self.assertAlmostEqual(ov["Profit Margin"], 0.25)


class TestDebtEquityReconciliation(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def test_screener_debt_equity_reconciles(self):
        with _patch_all():
            row = app.screener_row("TST")
        # annual basis: totalDebt 30e9 / equity 50e9 * 100 = 60.0
        self.assertAlmostEqual(row["debt_to_equity"], 60.0)
        self.assertAlmostEqual(row["debt_to_equity"],
                               row["total_debt"] / row["total_equity"] * 100)

    def test_screener_debt_equity_mrq_is_yahoo_value(self):
        with _patch_all():
            row = app.screener_row("TST")
        self.assertAlmostEqual(row["debt_to_equity_mrq"], 50.0)  # _BASE_INFO debtToEquity

    def test_deepdive_debt_equity_reconciles(self):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        with patch("app.get_info", return_value=_BASE_INFO), \
             patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.dividend_growth", return_value={
                 "cagr_3y": None, "cagr_5y": None, "annual": []}):
            d = app.deepdive("TST")
        h = d["panels"]["health"]
        self.assertAlmostEqual(h["Debt/Equity"],
                               h["Total Debt"] / h["Total Equity"] * 100)
        self.assertAlmostEqual(h["Debt/Equity"], 60.0)
        self.assertAlmostEqual(h["Debt/Equity (MRQ)"], 50.0)


class TestEbitdaRatios(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def test_screener_debt_ebitda(self):
        with _patch_all():
            row = app.screener_row("TST")
        # totalDebt 30e9 / ebitda 120e9 = 0.25
        self.assertAlmostEqual(row["debt_ebitda"], 0.25)

    def test_screener_ebitda_fcf(self):
        with _patch_all():
            row = app.screener_row("TST")
        # ebitda 120e9 / fcf 40e9 = 3.0
        self.assertAlmostEqual(row["ebitda_fcf"], 3.0)

    def test_screener_ebitda_value(self):
        with _patch_all():
            row = app.screener_row("TST")
        self.assertAlmostEqual(row["ebitda"], 120e9)

    def test_screener_ebitda_none_when_missing(self):
        info = {k: v for k, v in _BASE_INFO.items() if k != "ebitda"}
        with _patch_all(), patch("app.get_info", return_value=info):
            row = app.screener_row("TST")
        self.assertIsNone(row["ebitda"])

    def test_screener_ratios_none_without_ebitda(self):
        info = {k: v for k, v in _BASE_INFO.items() if k != "ebitda"}
        with patch("app.get_info", return_value=info), \
             patch("app._get_stmt", side_effect=lambda t, attr: {
                 "income_stmt": _INCOME_DF, "balance_sheet": _BAL_DF,
                 "cash_flow": _CF_DF}.get(attr)), \
             patch("app.get_dividends", return_value=_DIVS), \
             patch("app.get_raw_close", return_value=_CLOSE_SERIES), \
             patch("app.performance", return_value={
                 "ytd": 1.0, "1y": 1.0, "3y": 1.0, "5y": 1.0, "10y": 1.0}):
            row = app.screener_row("TST")
        self.assertIsNone(row["debt_ebitda"])
        self.assertIsNone(row["ebitda_fcf"])

    def test_deepdive_health_has_ebitda_ratios(self):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        with patch("app.get_info", return_value=_BASE_INFO), \
             patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.dividend_growth", return_value={
                 "cagr_3y": None, "cagr_5y": None, "annual": []}):
            d = app.deepdive("TST")
        h = d["panels"]["health"]
        self.assertAlmostEqual(h["Debt/EBITDA"], 0.25)
        self.assertAlmostEqual(h["EBITDA/FCF"], 3.0)   # ebitda 120e9 / fcf 40e9

    def test_growth_rows_carry_ebitda_margin(self):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        with patch("app.get_info", return_value=_BASE_INFO), \
             patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.dividend_growth", return_value={
                 "cagr_3y": None, "cagr_5y": None, "annual": []}):
            d = app.deepdive("TST")
        self.assertTrue(d["growth"])
        for g in d["growth"]:
            self.assertIn("ebitda_margin", g)

    def test_export_ratios_pass_through_not_fraction(self):
        # Debt/EBITDA and EBITDA/FCF are multiples, not percentages
        self.assertEqual(app._export_val("debt_ebitda", 0.25), 0.25)
        self.assertEqual(app._export_val("ebitda_fcf", 3.0), 3.0)
        keys = {k for k, _ in app._METRIC_COLS}
        self.assertIn("debt_ebitda", keys)
        self.assertIn("ebitda", keys)   # raw EBITDA exported alongside its ratios
        self.assertIn("ebitda_fcf", keys)


class TestDeepdivePanelCoverage(unittest.TestCase):
    """Screener metrics that were previously absent from the deep-dive panels."""

    def setUp(self):
        app.clear_cache()

    def _panels(self):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        with patch("app.get_info", return_value=_BASE_INFO), \
             patch("app.yf.Ticker", return_value=mock_tk), \
             patch("app.get_dividends", return_value=_DIVS), \
             patch("app.dividend_growth", return_value={
                 "cagr_3y": None, "cagr_5y": None, "annual": []}):
            return app.deepdive("TST")["panels"]

    def test_valuation_has_pc_pfcf_eps(self):
        v = self._panels()["valuation"]
        for k in ("Price/Cash", "Price/FCF", "Diluted EPS", "Basic EPS"):
            self.assertIn(k, v)
        # P/C = marketCap 1e12 / totalCash 10e9 = 100
        self.assertAlmostEqual(v["Price/Cash"], 100.0)

    def test_profitability_has_net_income(self):
        self.assertIn("Net Income", self._panels()["profitability"])

    def test_health_has_lt_debt_equity(self):
        # LT debt 25e9 / equity 50e9 * 100 = 50
        self.assertAlmostEqual(self._panels()["health"]["LT Debt/Equity"], 50.0)

    def test_dividend_has_ttm_and_years(self):
        d = self._panels()["dividend"]
        self.assertIn("Dividend TTM", d)
        self.assertIn("Years ▲ Dividend", d)


class TestEpochToIsoNoDeprecation(unittest.TestCase):

    def test_no_deprecation_warning(self):
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("error", DeprecationWarning)
            self.assertEqual(app._epoch_to_iso(1747000000), "2025-05-11")


# ---------------------------------------------------------------------------
# 18. Sparklines — downsampling + per-window price series
# ---------------------------------------------------------------------------
class TestDownsample(unittest.TestCase):

    def test_passthrough_when_at_or_under_target(self):
        self.assertEqual(app._downsample([1, 2, 3]), [1, 2, 3])
        self.assertEqual(app._downsample(list(range(50)), target=50), list(range(50)))

    def test_caps_at_target(self):
        self.assertEqual(len(app._downsample(list(range(1000)), target=50)), 50)

    def test_includes_first_and_last(self):
        out = app._downsample(list(range(1000)), target=50)
        self.assertEqual(out[0], 0)
        self.assertEqual(out[-1], 999)

    def test_custom_target(self):
        self.assertEqual(len(app._downsample(list(range(100)), target=10)), 10)

    def test_indices_monotonic(self):
        out = app._downsample(list(range(500)), target=50)
        self.assertEqual(out, sorted(out))


class TestSparklines(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def test_windows_present_and_capped(self):
        with patch("app.get_raw_close", return_value=_CLOSE_SERIES):
            s = app.sparklines("TST")
        for k in ("6mo", "1y", "5y"):
            self.assertIn(k, s)
            self.assertGreaterEqual(len(s[k]), 2)
            self.assertLessEqual(len(s[k]), 50)

    def test_empty_when_no_price_data(self):
        with patch("app.get_raw_close", return_value=None):
            self.assertEqual(app.sparklines("TST"),
                             {"6mo": [], "1y": [], "5y": []})

    def test_values_are_json_floats(self):
        with patch("app.get_raw_close", return_value=_CLOSE_SERIES):
            s = app.sparklines("TST")
        self.assertTrue(all(isinstance(v, float) for v in s["1y"]))
        json.dumps(s)  # must not raise

    def test_longer_window_starts_earlier(self):
        # The 5y window's first close should be <= the 6mo window's first close
        # index-wise (earlier start on a monotonically rising series → lower val).
        with patch("app.get_raw_close", return_value=_CLOSE_SERIES):
            s = app.sparklines("TST")
        self.assertLessEqual(s["5y"][0], s["6mo"][0])


class TestScreenerRowSparklines(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def test_row_carries_spark_series(self):
        with _patch_all():
            row = app.screener_row("TST")
        for k in ("spark_6mo", "spark_1y", "spark_5y"):
            self.assertIn(k, row)
            self.assertIsInstance(row[k], list)
            self.assertTrue(row[k])   # non-empty for the mocked price series


# ---------------------------------------------------------------------------
# 19. Performance columns — export order + percent number format
# ---------------------------------------------------------------------------
class TestPerfColumnOrder(unittest.TestCase):

    def test_perf_block_follows_industry_longest_first(self):
        # After Industry come the strategy grades and flags, then the
        # performance block (longest window first).
        keys = [k for k, _ in app._METRIC_COLS]
        i = keys.index("industry")
        self.assertEqual(
            keys[i + 1:i + 6],
            ["strategy_1", "strategy_2", "strategy_3", "strategy_min",
             "strategy_1_flags"],
        )
        self.assertEqual(
            keys[i + 6:i + 11],
            ["perf_10y", "perf_5y", "perf_3y", "perf_1y", "perf_ytd"],
        )

    def test_perf_columns_appear_once(self):
        keys = [k for k, _ in app._METRIC_COLS]
        for k in ("perf_ytd", "perf_1y", "perf_3y", "perf_5y", "perf_10y"):
            self.assertEqual(keys.count(k), 1)


@unittest.skipUnless(app._HAS_OPENPYXL, "openpyxl not installed")
class TestPerfColumnFormat(unittest.TestCase):

    def setUp(self):
        app.clear_cache()

    def _build(self):
        mock_tk = MagicMock()
        mock_tk.income_stmt = _INCOME_DF
        mock_tk.cash_flow = _CF_DF
        mock_tk.balance_sheet = _BAL_DF
        close = pd.Series([100.0, 105.0],
                          index=pd.DatetimeIndex(["2025-01-02", "2025-01-03"]))
        mock_tk.history.return_value = pd.DataFrame({"Close": close, "Volume": [1e6, 1e6]})
        with patch("app.get_info", return_value=_BASE_INFO), \
             patch("app._get_stmt", side_effect=lambda t, attr: {
                 "income_stmt": _INCOME_DF, "balance_sheet": _BAL_DF,
                 "cash_flow": _CF_DF}.get(attr)), \
             patch("app.get_dividends", return_value=_DIVS), \
             patch("app.get_raw_close", return_value=_CLOSE_SERIES), \
             patch("app.performance", return_value={
                 "ytd": 5.0, "1y": 12.0, "3y": 30.0, "5y": 60.0, "10y": 100.0}), \
             patch("app.yf.Ticker", return_value=mock_tk):
            raw = app.build_workbook(["TST"])
        from openpyxl import load_workbook
        return load_workbook(io.BytesIO(raw))["Metrics"]

    def test_perf_cells_have_percent_format(self):
        ws = self._build()
        headers = [c.value for c in ws[1]]
        label_for = {k: lab for k, lab in app._METRIC_COLS}
        for key in ("perf_ytd", "perf_1y", "perf_5y", "perf_10y"):
            col = headers.index(label_for[key]) + 1
            cell = ws.cell(row=2, column=col)
            self.assertEqual(cell.number_format, "0.00%",
                             f"{key} not percent-formatted")

    def test_perf_value_is_fraction(self):
        ws = self._build()
        headers = [c.value for c in ws[1]]
        label_for = {k: lab for k, lab in app._METRIC_COLS}
        # perf_1y = 12.0 (pct points) stored as the fraction 0.12; the % format
        # renders it as 12.00%.
        col = headers.index(label_for["perf_1y"]) + 1
        self.assertAlmostEqual(ws.cell(row=2, column=col).value, 0.12)


# ---------------------------------------------------------------------------
# 20. Strategy graders (strategies.py)
# ---------------------------------------------------------------------------
def _strategy_row(**over):
    """A healthy capital-intensive (manufacturer-like) screener row — the
    default archetype, on which the classic Altman/ROIC-WACC rubric applies
    unchanged. Units match screener_row: margins/roe/roa/payout are fractions;
    roic/roce/wacc/div_yield/debt_to_equity/perf_* are percentage points.
    Override `sector` to exercise the other archetypes (see _business_type)."""
    row = {
        "ticker": "TST", "name": "Test Co", "sector": "Industrials",
        "industry": "Specialty Industrial Machinery", "currency": "USD",
        "price": 100.0, "market_cap": 1e12, "enterprise_value": 1e12,
        "pe": 20.0, "forward_pe": 18.0, "peg": 1.2, "pb": 5.0, "ps": 5.0,
        "pc": 50.0, "p_fcf": 20.0, "ev_ebitda": 12.0, "eps": 5.0, "eps_basic": 5.0,
        "income": 10e9, "profit_margin": 0.20, "gross_margin": 0.55,
        "operating_margin": 0.25, "ebitda_margin": 0.30, "fcf": 9e9,
        "roa": 0.12, "roe": 0.25, "roic": 18.0, "roce": 20.0, "wacc": 9.0,
        "revenue_per_share": 40.0,
        "debt_to_equity": 40.0, "debt_to_equity_mrq": 40.0, "debt_ebitda": 1.0,
        "lt_debt_to_equity": 30.0, "current_ratio": 2.0, "quick_ratio": 1.5,
        "total_cash": 20e9, "total_debt": 10e9, "total_equity": 50e9,
        "ebitda": 15e9, "ebitda_fcf": 1.67,
        # total_cash > total_debt, so this row is net cash — S3 Pillar C's
        # FCF/net-debt leg takes its full points without touching fcf.
        "interest_expense": 0.5e9, "interest_coverage": 15.0,
        "div_yield": 2.0, "five_year_avg_yield": 2.0, "payout_ratio": 0.35,
        "div_growth_3y": 8.0, "div_growth_5y": 8.0, "dividend_estimate": 2.0,
        "dividend_ttm": 2.0, "fcf_coverage": 3.0, "years_div_increase": 12,
        "ex_dividend_date": "2026-05-11",
        "beta": 1.0, "short_interest": 0.02, "days_to_cover": 2.0,
        # altman_z carries a market-cap term, altman_z_prime doesn't, and they
        # sit on different bands (3.0/1.81 vs 2.9/1.23) — both healthy here.
        "altman_z": 5.0, "altman_z_prime": 3.5, "piotroski_f": 8,
        "perf_ytd": 5.0, "perf_1y": 15.0, "perf_3y": 60.0,
        "perf_5y": 120.0, "perf_10y": 300.0,
    }
    row.update(over)
    return row


class TestStrategyTriage(unittest.TestCase):

    def test_strong_row_scores_100_advance(self):
        score, verdict = strategies.grade_triage(_strategy_row())
        self.assertEqual(score, 100)
        self.assertEqual(verdict, "Advance")

    def test_missing_equity_quarantines(self):
        score, verdict = strategies.grade_triage(_strategy_row(total_equity=None))
        self.assertIsNone(score)
        self.assertIn("Quarantine", verdict)
        self.assertIn("total_equity", verdict)

    def test_blank_debt_with_cash_treated_as_zero_not_quarantined(self):
        # The blank-but-cash-rich carve-out: no Total Debt but cash on hand ->
        # zero debt (best case), so leverage metrics take full points.
        row = _strategy_row(total_debt=None, debt_ebitda=None, debt_to_equity=None)
        score, verdict = strategies.grade_triage(row)
        self.assertEqual(score, 100)
        self.assertEqual(verdict, "Advance")

    def test_blank_debt_without_cash_quarantines(self):
        row = _strategy_row(total_debt=None, total_cash=None,
                            debt_ebitda=None, debt_to_equity=None)
        score, verdict = strategies.grade_triage(row)
        self.assertIsNone(score)
        self.assertIn("total_debt", verdict)

    def test_high_ev_market_cap_ratio_is_not_quarantined(self):
        # There is no currency-mixing quarantine on EV/Market Cap: Yahoo
        # always returns both in the same currency, so a large ratio here
        # means genuine leverage, not a data artifact — it must score
        # normally (and can still legitimately fail Stage 1/2 on its own
        # merits, just not via a Stage 0 quarantine).
        row = _strategy_row(market_cap=1e9, enterprise_value=5e9)
        score, verdict = strategies.grade_triage(row)
        self.assertIsNotNone(score)
        self.assertNotIn("Quarantine", verdict)

    def test_twin_negative_kill_switch(self):
        score, verdict = strategies.grade_triage(
            _strategy_row(income=-1e9, fcf=-1e9))
        self.assertEqual(score, 0)
        self.assertIn("Discard", verdict)
        self.assertIn("twin-negative", verdict)

    def test_single_negative_is_not_killed(self):
        score, _ = strategies.grade_triage(_strategy_row(fcf=-1e9))
        self.assertGreater(score, 0)

    def test_liquidity_kill_requires_both_conditions(self):
        # CR < 1 but QR at 0.6 (not < 0.5): compound condition must NOT fire.
        score, verdict = strategies.grade_triage(
            _strategy_row(current_ratio=0.96, quick_ratio=0.6))
        self.assertGreater(score, 0)
        self.assertNotIn("Discard", verdict)
        # Both legs breached -> kill.
        score2, verdict2 = strategies.grade_triage(
            _strategy_row(current_ratio=0.96, quick_ratio=0.4))
        self.assertEqual(score2, 0)
        self.assertIn("liquidity", verdict2)

    def test_altman_distress_kill(self):
        score, verdict = strategies.grade_triage(_strategy_row(altman_z=1.2))
        self.assertEqual(score, 0)
        self.assertIn("Altman", verdict)

    def test_roe_leverage_cap(self):
        # Same ROE but D/E > 1.5 (150 pct points): ROE points halved (10 -> 5).
        base, _ = strategies.grade_triage(_strategy_row(debt_to_equity=140.0))
        capped, _ = strategies.grade_triage(_strategy_row(debt_to_equity=160.0))
        # d_eq 140 already loses the Pillar-C debt/equity points vs 160 equally
        # (both > 100), so the only difference is the ROE cap.
        self.assertEqual(base - capped, 5)

    def test_financials_scored_on_roe_margin_piotroski(self):
        row = _strategy_row(sector="Financial Services",
                            industry="Banks—Diversified", altman_z=1.0,
                            debt_ebitda=8.0, current_ratio=0.5, quick_ratio=0.4,
                            roe=0.18, profit_margin=0.25, piotroski_f=8)
        score, verdict = strategies.grade_triage(row)
        # Altman / leverage / liquidity kill-switches must not fire; full marks
        # on the financial rubric: ROE 40 + net margin 30 + Piotroski 30.
        self.assertEqual(score, 100)
        self.assertEqual(verdict, "Advance")

    def test_negative_spread_caps_score_at_55(self):
        # roic=18, wacc=9 (default) -> spread +9 -> no gate, full 100.
        # Push wacc above roic for a negative spread; everything else stays
        # strong, so without the gate this would still score 100/Advance.
        row = _strategy_row(wacc=20.0)
        score, verdict = strategies.grade_triage(row)
        self.assertEqual(score, 55)
        self.assertEqual(verdict, "Watchlist")

    def test_thin_positive_spread_is_not_gated(self):
        # roic=9.5, wacc=9 -> spread +0.5 (thin, but not negative): the gate
        # must not fire on estimation noise around a small positive spread.
        row = _strategy_row(roic=9.5, wacc=9.0)
        score, verdict = strategies.grade_triage(row)
        self.assertGreater(score, 55)

    def test_negative_equity_scores_worst_leverage_band(self):
        # Negative equity flips D/E negative, which would otherwise beat every
        # "low leverage" band; ROE on negative equity is arithmetic noise.
        # Both must score 0: -10 (ROE) and -8 (Pillar C D/E) vs the base row.
        row = _strategy_row(total_equity=-5e9, debt_to_equity=-20.0)
        score, verdict = strategies.grade_triage(row)
        self.assertEqual(score, 82)
        self.assertEqual(verdict, "Advance")

    def test_negative_ebitda_with_net_debt_kills(self):
        # Debt/EBITDA is undefined (None) when EBITDA <= 0 — without the
        # companion switch this row would sail past the leverage kill.
        row = _strategy_row(ebitda=-2e9, debt_ebitda=None, ebitda_fcf=None,
                            total_cash=1e9)
        score, verdict = strategies.grade_triage(row)
        self.assertEqual(score, 0)
        self.assertIn("negative EBITDA", verdict)

    def test_negative_ebitda_with_net_cash_not_killed(self):
        # Same negative EBITDA but cash exceeds debt: no leverage problem.
        row = _strategy_row(ebitda=-2e9, debt_ebitda=None, ebitda_fcf=None)
        score, verdict = strategies.grade_triage(row)
        self.assertGreater(score, 0)
        self.assertNotIn("Discard", verdict)

    def test_growth_phase_routed_to_watchlist(self):
        # A 40-44 scorer failing purely on margins, with strong gross margin
        # (>60%) and positive FCF, is routed to Watchlist instead of Discard.
        row = _strategy_row(gross_margin=0.65, operating_margin=0.02,
                            profit_margin=0.01, roe=0.05, roic=2.0,
                            income=0.5e9, fcf=0.45e9, ebitda_fcf=5.0)
        score, verdict = strategies.grade_triage(row)
        self.assertEqual(score, 42)
        self.assertEqual(verdict, "Watchlist")
        # Same shape but a thin gross margin: genuinely weak -> Discard.
        weak = _strategy_row(gross_margin=0.30, operating_margin=0.02,
                             profit_margin=0.01, roe=0.05, roic=2.0,
                             income=0.5e9, fcf=0.45e9, ebitda_fcf=5.0)
        score2, verdict2 = strategies.grade_triage(weak)
        self.assertLess(score2, 45)
        self.assertEqual(verdict2, "Discard")

    def test_negative_spread_gate_does_not_apply_to_financials(self):
        # Financials don't use ROIC/WACC in Stage 2 at all, so a negative
        # spread on a financial-sector row must not trigger the gate.
        row = _strategy_row(sector="Financial Services",
                            industry="Banks—Diversified", altman_z=1.0,
                            debt_ebitda=8.0, current_ratio=0.5, quick_ratio=0.4,
                            roe=0.18, profit_margin=0.25, piotroski_f=8,
                            roic=-5.0, wacc=9.0)
        score, verdict = strategies.grade_triage(row)
        self.assertEqual(score, 100)
        self.assertEqual(verdict, "Advance")

    def test_thin_margins_with_strong_returns_floor_profitability(self):
        # High-turnover model (grocer/distributor): every margin misses its
        # band, but the fixture's ROIC 18 vouches for the model — the pillar is
        # floored at half (12.5/25) instead of 0.
        thin = dict(profit_margin=0.02, operating_margin=0.04, gross_margin=0.20)
        _, _, pillars = strategies._grade_triage(_strategy_row(**thin))
        pill = next(p for p in pillars if p["k"] == "Profitability")
        self.assertEqual(pill["p"], 12.5)
        self.assertIn("floored", pill["d"])
        # weak returns: same margins score what they score — nothing vouches
        _, _, p2 = strategies._grade_triage(
            _strategy_row(roic=5.0, roe=0.10, **thin))
        self.assertEqual(next(p for p in p2 if p["k"] == "Profitability")["p"], 0.0)

    def test_margin_floor_roe_leg_requires_modest_leverage(self):
        # Without ROIC, the ROE ≥ 15 leg only vouches when the ROE isn't a
        # leverage artifact: D/E ≤ 100% qualifies, 200% doesn't.
        thin = dict(profit_margin=0.02, operating_margin=0.04, gross_margin=0.20,
                    roic=5.0, roe=0.20)
        _, _, ok = strategies._grade_triage(
            _strategy_row(debt_to_equity=80.0, **thin))
        self.assertEqual(next(p for p in ok if p["k"] == "Profitability")["p"], 12.5)
        _, _, lev = strategies._grade_triage(
            _strategy_row(debt_to_equity=200.0, **thin))
        self.assertEqual(next(p for p in lev if p["k"] == "Profitability")["p"], 0.0)


class TestFinancialClassification(unittest.TestCase):
    """_is_financial matches balance-sheet businesses by industry — Yahoo's
    'Financial Services' sector alone also sweeps in fee businesses."""

    def _fin(self, industry, sector="Financial Services"):
        return strategies._is_financial({"sector": sector, "industry": industry})

    def test_balance_sheet_industries_match(self):
        for ind in ("Banks—Diversified", "Banks—Regional",
                    "Insurance—Property & Casualty", "Capital Markets",
                    "Credit Services", "Mortgage Finance",
                    "Financial Conglomerates"):
            self.assertTrue(self._fin(ind), ind)

    def test_fee_businesses_score_on_standard_rubric(self):
        for ind in ("Insurance Brokers", "Financial Data & Stock Exchanges",
                    "Asset Management"):
            self.assertFalse(self._fin(ind), ind)

    def test_non_financial_industry_wins_over_financial_sector(self):
        self.assertFalse(self._fin("Software—Infrastructure"))

    def test_sector_fallback_when_industry_missing(self):
        self.assertTrue(self._fin(None))
        self.assertFalse(strategies._is_financial(
            {"sector": "Technology", "industry": None}))


class TestBusinessType(unittest.TestCase):
    """_business_type routes each sector to the right archetype rubric."""

    def _bt(self, sector, industry=""):
        return strategies._business_type({"sector": sector, "industry": industry})

    def test_archetype_by_sector(self):
        self.assertEqual(self._bt("Industrials", "Aerospace & Defense"),
                         "capital_intensive")
        self.assertEqual(self._bt("Utilities", "Utilities—Regulated Electric"),
                         "capital_intensive")
        self.assertEqual(self._bt("Consumer Cyclical", "Auto Manufacturers"),
                         "capital_intensive")
        self.assertEqual(self._bt("Technology", "Software—Infrastructure"),
                         "asset_light")
        self.assertEqual(self._bt("Communication Services", "Internet Content"),
                         "asset_light")
        self.assertEqual(self._bt("Healthcare", "Biotechnology"), "asset_light")
        self.assertEqual(self._bt("Energy", "Oil & Gas E&P"), "cyclical")
        self.assertEqual(self._bt("Basic Materials", "Copper"), "cyclical")
        # industry overrides: sectors that misroute these capital-heavy models
        self.assertEqual(self._bt("Technology", "Semiconductors"),
                         "capital_intensive")
        self.assertEqual(self._bt("Technology",
                                  "Semiconductor Equipment & Materials"),
                         "capital_intensive")
        self.assertEqual(self._bt("Communication Services", "Telecom Services"),
                         "capital_intensive")
        self.assertEqual(self._bt("Real Estate", "REIT—Retail"), "reit")
        # a mortgage REIT is its own archetype (leveraged securities portfolio)
        self.assertEqual(self._bt("Real Estate", "REIT—Mortgage"), "mreit")
        self.assertEqual(self._bt("Real Estate", "REIT - Mortgage"), "mreit")

    def test_financial_takes_precedence(self):
        # a balance-sheet business is 'financial' even if sector says otherwise
        self.assertEqual(self._bt("Financial Services", "Banks—Regional"),
                         "financial")
        # a mortgage *originator* (no "reit" in the industry) stays a financial —
        # only mortgage REITs (industry has both "reit" and "mortgage") route to mreit
        self.assertEqual(self._bt("Financial Services", "Mortgage Finance"),
                         "financial")

    def test_unknown_sector_defaults_to_capital_intensive(self):
        self.assertEqual(self._bt("", ""), "capital_intensive")
        self.assertEqual(self._bt("Consumer Defensive", "Grocery Stores"),
                         "capital_intensive")


class TestBusinessTypeGrading(unittest.TestCase):
    """Altman-Z and ROIC-WACC are only decisive for the archetypes they fit."""

    def _distressed(self, **over):
        # low Altman-Z + weak everything, but NOT twin-negative / Piotroski-killed
        return _strategy_row(altman_z=1.2, piotroski_f=6, **over)

    def test_altman_kill_only_for_capital_intensive(self):
        cap = strategies.grade_triage(self._distressed(sector="Industrials"))
        self.assertEqual(cap[0], 0)
        self.assertIn("Altman Z distress", cap[1])
        # asset-light / cyclical / REIT are not killed by a low Altman-Z
        for sector in ("Technology", "Energy", "Real Estate"):
            score, verdict = strategies.grade_triage(self._distressed(sector=sector))
            self.assertNotIn("Altman", verdict, sector)
            self.assertNotEqual(score, 0, sector)

    def test_low_altman_flag_for_asset_light_and_cyclical(self):
        for sector in ("Technology", "Energy"):
            row = self._distressed(sector=sector)
            flags = strategies.triage_flags(row, 70)
            self.assertTrue(any("Low Altman-Z" in f for f in flags), sector)
        # capital-intensive is killed (not flagged); REIT ignores Altman entirely
        self.assertFalse(any("Low Altman-Z" in f for f in
                             strategies.triage_flags(self._distressed(sector="Real Estate"), 70)))

    def test_utility_not_killed_by_low_altman(self):
        # Altman excluded utilities from his sample; a healthy regulated
        # utility's Z sits below 1.8 by construction — soft flag, not a kill.
        row = self._distressed(sector="Utilities",
                               industry="Utilities—Regulated Electric")
        score, verdict = strategies.grade_triage(row)
        self.assertNotIn("Altman", verdict)
        self.assertNotEqual(score, 0)
        flags = strategies.triage_flags(row, score)
        self.assertTrue(any("Low Altman-Z" in f for f in flags))

    def test_utility_debt_ebitda_kill_at_7(self):
        # Regulated utilities run 4.5–5.5× in the normal course — the kill
        # line moves from 6× to 7× for them, and only for them.
        util = dict(sector="Utilities", industry="Utilities—Regulated Electric")
        ok, v_ok = strategies.grade_triage(_strategy_row(debt_ebitda=6.5, **util))
        self.assertNotIn("Debt/EBITDA", v_ok)
        killed, v_kill = strategies.grade_triage(
            _strategy_row(debt_ebitda=7.5, **util))
        self.assertEqual(killed, 0)
        self.assertIn("Debt/EBITDA > 7", v_kill)
        # an industrial at 6.5× is still killed on the generic 6× line
        ind_killed, v_ind = strategies.grade_triage(_strategy_row(debt_ebitda=6.5))
        self.assertEqual(ind_killed, 0)
        self.assertIn("Debt/EBITDA > 6", v_ind)

    def test_utility_low_altman_spared_by_s2_solvency_guard(self):
        # The S2 solvency guard's Altman leg is also a manufacturer test.
        row = _strategy_row(sector="Utilities",
                            industry="Utilities—Regulated Electric",
                            altman_z=1.2)
        score, _ = strategies.grade_compounder(row)
        self.assertGreater(score, 35)
        ind, _ = strategies.grade_compounder(_strategy_row(altman_z=1.2))
        self.assertLessEqual(ind, 35)

    def test_reit_not_killed_by_leverage_or_liquidity(self):
        # a REIT with heavy debt and thin liquidity that would sink an industrial
        row = _strategy_row(sector="Real Estate", industry="REIT—Residential",
                            debt_ebitda=9.0, current_ratio=0.6, quick_ratio=0.3,
                            debt_to_equity=140.0)
        score, verdict = strategies.grade_triage(row)
        self.assertNotEqual(score, 0)
        self.assertNotIn("Debt/EBITDA", verdict)
        self.assertNotIn("liquidity", verdict)

    def test_value_destruction_kill_exempts_cyclical_and_reit(self):
        vd = dict(roic=-5.0, operating_margin=-0.10)
        # capital-intensive & asset-light: killed
        for sector in ("Industrials", "Technology"):
            score, verdict = strategies.grade_triage(
                _strategy_row(sector=sector, **vd))
            self.assertIn("value destruction", verdict, sector)
        # cyclical (trough) & REIT: not killed on value destruction
        for sector in ("Energy", "Real Estate"):
            score, verdict = strategies.grade_triage(
                _strategy_row(sector=sector, **vd))
            self.assertNotIn("value destruction", verdict, sector)

    def test_negative_spread_cap_needs_deep_spread_off_manufacturing(self):
        # spread = roic - wacc = 8 - 12 = -4  (shallow negative)
        shallow = dict(roic=8.0, wacc=12.0)
        cap = strategies.grade_triage(_strategy_row(sector="Industrials", **shallow))
        self.assertLessEqual(cap[0], 55)          # capital-intensive: capped at <0
        light = strategies.grade_triage(_strategy_row(sector="Technology", **shallow))
        self.assertGreater(light[0], 55)          # asset-light: -4 is within noise
        # a deep negative spread (-10) caps even asset-light
        deep = strategies.grade_triage(
            _strategy_row(sector="Technology", roic=2.0, wacc=12.0))
        self.assertLessEqual(deep[0], 55)

    def test_s2_solvency_guard_altman_only_capital_intensive(self):
        capped = strategies.grade_compounder(
            _strategy_row(sector="Industrials", altman_z=1.2))
        self.assertLessEqual(capped[0], 35)
        # asset-light with the same low Z is not solvency-capped by Altman
        light = strategies.grade_compounder(
            _strategy_row(sector="Technology", altman_z=1.2))
        self.assertGreater(light[0], 35)

    def test_reit_grades_are_reasonable_and_bounded(self):
        row = _strategy_row(sector="Real Estate", industry="REIT—Residential")
        for fn in (strategies._grade_triage, strategies._grade_compounder,
                   strategies._grade_defensive):
            score, verdict, pillars = fn(dict(row))
            self.assertIsNotNone(score)
            self.assertTrue(0 <= score <= 100, f"{verdict}: {score}")
            self.assertEqual(strategies._round_score(sum(p["p"] for p in pillars)),
                             score)

    def test_reit_uses_ffo_when_available(self):
        # ffo/p_ffo/ffo_payout/ffo_coverage are populated the way app.py would
        # (from Net Income + D&A) — the REIT branch should switch onto them.
        row = _strategy_row(sector="Real Estate", industry="REIT—Residential",
                            ffo=5e9, p_ffo=10.0, ffo_payout=0.75, ffo_coverage=1.3)
        score, verdict, pillars = strategies._grade_triage(row)
        by_k = {p["k"]: p for p in pillars}
        self.assertIn("FFO", by_k["Cash generation"]["d"])
        self.assertIn("FFO payout", by_k["Distribution"]["d"])
        self.assertIn("P/FFO", by_k["Valuation"]["d"])
        self.assertTrue(0 <= score <= 100, verdict)

    def test_reit_falls_back_to_fcf_without_ffo_data(self):
        # no ffo/p_ffo override -> the base fixture has neither field at all,
        # simulating a REIT whose cash flow statement lacks a D&A line.
        row = _strategy_row(sector="Real Estate", industry="REIT—Residential")
        self.assertNotIn("ffo", row)
        _score, _verdict, pillars = strategies._grade_triage(row)
        by_k = {p["k"]: p for p in pillars}
        self.assertIn("FCF", by_k["Cash generation"]["d"])
        self.assertIn("no FFO data", by_k["Valuation"]["d"])

    def _reit(self, **over):
        base = dict(sector="Real Estate", industry="REIT—Residential",
                    ffo=5e9, p_ffo=14.0, ffo_payout=0.7, ffo_coverage=1.4,
                    total_debt=40e9, total_equity=50e9)
        base.update(over)
        return _strategy_row(**base)

    def test_s2_reit_branch_uses_ffo_not_industrial_bands(self):
        _score, _v, pillars = strategies._grade_compounder(self._reit())
        by_k = {p["k"]: p for p in pillars}
        # returns measured on FFO/IC, discipline on FFO payout, valuation on P/FFO
        self.assertIn("FFO/IC", by_k["Returns on capital"]["d"])
        self.assertIn("FFO payout", by_k["Capital discipline"]["d"])
        self.assertIn("REIT bands", by_k["Capital discipline"]["d"])
        self.assertIn("P/FFO", by_k["Valuation sanity"]["d"])

    def test_s2_reit_not_zeroed_by_high_debt_ebitda(self):
        # Debt/EBITDA of 8x would zero an industrial's capital-discipline pillar;
        # a REIT with conservative D/E should still earn its leverage points.
        row = self._reit(debt_ebitda=8.0, debt_to_equity=80.0)
        _score, _v, pillars = strategies._grade_compounder(row)
        c = next(p for p in pillars if p["k"] == "Capital discipline")
        self.assertGreater(c["p"], 10)          # leverage + FFO-payout points earned

    def test_s3_reit_branch_uses_ffo_and_pb(self):
        _score, _v, pillars = strategies._grade_defensive(self._reit())
        by_k = {p["k"]: p for p in pillars}
        self.assertIn("P/FFO", by_k["Earnings/cash yield"]["d"])
        self.assertIn("discount to assets", by_k["Asset backing"]["d"])
        self.assertIn("FFO", by_k["Earnings quality"]["d"])

    def test_s3_reit_cheap_on_ffo_scores_as_value(self):
        # a REIT trading at a low P/FFO and below book should read as value,
        # where the old P/E-based pillar A would have scored it near zero
        # (REIT P/E is depreciation-inflated).
        cheap = strategies.grade_defensive(self._reit(p_ffo=8.0, pb=0.9))
        rich = strategies.grade_defensive(self._reit(p_ffo=30.0, pb=3.0))
        self.assertGreater(cheap[0], rich[0])
        self.assertGreaterEqual(cheap[0], 60)

    def test_s3_reit_strength_scores_coverage_on_ffo_not_fcf(self):
        # Pillar C must judge the distribution on FFO like S1/S2 do, not on the
        # GAAP FCF the rest of the REIT rubric deliberately avoids. fcf_coverage
        # is set far above every band, so a pillar reading it would score full.
        _s, _v, pillars = strategies._grade_defensive(
            self._reit(ffo_coverage=1.3, fcf_coverage=99.0))
        c = next(p for p in pillars if p["k"] == "Financial strength")
        self.assertIn("FFO cover", c["d"])
        self.assertNotIn("FCF cover", c["d"])
        self.assertEqual(c["p"], 21.0)          # 12 D/E + 4 half-band + 5 FFO>0

    def test_s3_reit_in_development_not_punished_for_negative_gaap_fcf(self):
        # The bug this guards: a REIT building properties has deeply negative
        # GAAP FCF (development capex is indistinguishable from maintenance) and
        # was losing all 13 FCF-driven points of Pillar C while healthy on FFO.
        dev = self._reit(fcf=-2e9, fcf_coverage=-0.5, ffo_coverage=1.8)
        _s, _v, pillars = strategies._grade_defensive(dev)
        c = next(p for p in pillars if p["k"] == "Financial strength")
        self.assertEqual(c["p"], 25.0)          # 12 D/E + 8 cover + 5 FFO>0

    def test_s3_reit_without_ffo_falls_back_to_fcf_bands(self):
        # No D&A line -> no FFO -> the lenient GAAP-FCF bands still apply, so a
        # REIT is never left unscored on strength.
        row = self._reit()
        del row["ffo"], row["ffo_coverage"]
        _s, _v, pillars = strategies._grade_defensive(row)
        c = next(p for p in pillars if p["k"] == "Financial strength")
        self.assertIn("FCF cover", c["d"])
        self.assertEqual(c["p"], 25.0)          # 12 D/E + 8 (fcf_cov 3.0) + 5 FCF>0

    def test_reit_pillars_resum_across_all_three_strategies(self):
        for fn in (strategies._grade_triage, strategies._grade_compounder,
                   strategies._grade_defensive):
            score, _v, pillars = fn(self._reit())
            self.assertEqual(
                strategies._round_score(sum(p["p"] for p in pillars)), score)
            for p in pillars:
                if p["m"] > 0:
                    self.assertLessEqual(p["p"], p["m"])
                    self.assertGreaterEqual(p["p"], 0)

    # ----- mortgage REITs (mreit) -----
    def _mreit(self, **over):
        base = dict(sector="Real Estate", industry="REIT - Mortgage",
                    piotroski_f=5, altman_z=1.0)   # low Altman must NOT kill mREITs
        base.update(over)
        return _strategy_row(**base)

    def test_mreit_uses_dedicated_rubric_not_bank_rubric(self):
        # S1 pillars are the mREIT ones (coverage / P-B / leverage / BVPS),
        # NOT the financial rubric's ROE / net margin / Piotroski.
        _s, _v, pillars = strategies._grade_triage(self._mreit())
        labels = {p["k"] for p in pillars}
        self.assertEqual(labels, {"Dividend coverage", "Price vs book",
                                  "Leverage (mREIT)", "Book value trend"})
        self.assertNotIn("Net margin", labels)

    def test_mreit_uncovered_dividend_and_eroding_book_scores_discard(self):
        # ORC-shaped: 152% payout (uncovered), near book, ~8x leverage, book
        # value eroding — should land in Discard, not the old inflated Advance.
        row = self._mreit(payout_ratio=1.52, pb=0.97, debt_to_equity=790.0,
                          bvps_growth=-14.0)
        score, verdict = strategies.grade_triage(row)
        self.assertLess(score, 45)
        self.assertEqual(verdict, "Discard")

    def test_mreit_covered_dividend_and_stable_book_scores_well(self):
        # NLY-shaped: 90% covered payout, mild BVPS slide, agency leverage.
        row = self._mreit(payout_ratio=0.90, pb=0.98, debt_to_equity=740.0,
                          bvps_growth=-1.8)
        score, _verdict = strategies.grade_triage(row)
        self.assertGreaterEqual(score, 55)   # covered + holding book -> respectable

    def test_mreit_leverage_bands_tolerate_agency_levels(self):
        # 8x (800%) is normal agency leverage -> full points; >1000% -> zero.
        base = dict(payout_ratio=0.9, pb=0.9, bvps_growth=1.0)
        normal = strategies._grade_triage(self._mreit(debt_to_equity=800.0, **base))
        excessive = strategies._grade_triage(self._mreit(debt_to_equity=1200.0, **base))
        lev_n = next(p for p in normal[2] if p["k"] == "Leverage (mREIT)")["p"]
        lev_x = next(p for p in excessive[2] if p["k"] == "Leverage (mREIT)")["p"]
        self.assertEqual(lev_n, 20)
        self.assertEqual(lev_x, 0)

    def test_mreit_low_altman_does_not_kill(self):
        # altman_z=1.0 would kill a capital-intensive name; a mREIT ignores it.
        score, verdict = strategies.grade_triage(self._mreit())
        self.assertNotEqual(score, 0)
        self.assertNotIn("Altman", verdict)

    def test_mreit_pillars_resum_across_all_three_strategies(self):
        for fn in (strategies._grade_triage, strategies._grade_compounder,
                   strategies._grade_defensive):
            score, _v, pillars = fn(self._mreit(payout_ratio=1.1, pb=1.0,
                                                debt_to_equity=700.0, bvps_growth=-3.0))
            self.assertEqual(
                strategies._round_score(sum(p["p"] for p in pillars)), score)
            for p in pillars:
                if p["m"] > 0:
                    self.assertLessEqual(p["p"], p["m"])
                    self.assertGreaterEqual(p["p"], 0)

    def test_piotroski_kill_exempts_financial_and_reit(self):
        weak_f = {"piotroski_f": 2}
        # capital-intensive, asset-light and cyclical are still killed
        for sector in ("Industrials", "Technology", "Energy"):
            score, verdict = strategies.grade_triage(_strategy_row(sector=sector, **weak_f))
            self.assertEqual(score, 0, sector)
            self.assertIn("Piotroski", verdict, sector)
        # financial and REIT: not killed, flagged instead
        for sector, industry in (("Financial Services", "Banks—Regional"),
                                 ("Real Estate", "REIT—Residential")):
            row = _strategy_row(sector=sector, industry=industry, **weak_f)
            score, verdict = strategies.grade_triage(row)
            self.assertNotIn("Piotroski", verdict, sector)
            self.assertNotEqual(score, 0, sector)
            flags = strategies.triage_flags(row, score)
            self.assertTrue(any("Low Piotroski" in f for f in flags), sector)

    def test_low_piotroski_flag_absent_for_healthy_bank_and_reit(self):
        for sector, industry in (("Financial Services", "Banks—Regional"),
                                 ("Real Estate", "REIT—Residential")):
            row = _strategy_row(sector=sector, industry=industry)  # piotroski_f=8
            score, _verdict = strategies.grade_triage(row)
            flags = strategies.triage_flags(row, score)
            self.assertFalse(any("Low Piotroski" in f for f in flags), sector)


class TestRoundScore(unittest.TestCase):

    def test_rounds_half_up_not_bankers(self):
        # Python's round() is banker's rounding (64.5 -> 64); band edges must
        # round predictably upward instead.
        self.assertEqual(strategies._round_score(64.5), 65)
        self.assertEqual(strategies._round_score(63.5), 64)
        self.assertEqual(strategies._round_score(64.4), 64)


class TestTriageFlags(unittest.TestCase):
    """Stage 0 sanity + Stage 3 valuation-context flags (never disqualifying)."""

    def test_healthy_row_has_no_flags(self):
        self.assertEqual(strategies.triage_flags(_strategy_row(), 80), [])

    def test_quarantined_row_gets_no_flags(self):
        self.assertEqual(
            strategies.triage_flags(_strategy_row(pb=45.0), None), [])

    def test_priced_for_perfection_triggers(self):
        for over in ({"peg": 3.5}, {"p_fcf": 45.0}, {"ev_ebitda": 35.0}):
            flags = strategies.triage_flags(_strategy_row(**over), 80)
            self.assertTrue(any("Priced for perfection" in f for f in flags), over)

    def test_suspiciously_cheap_needs_mediocre_score(self):
        cheap = _strategy_row(pe=7.0)
        self.assertTrue(any("Suspiciously cheap" in f
                            for f in strategies.triage_flags(cheap, 55)))
        # Same multiples on a strong scorer: cheapness alone is not suspicious.
        self.assertFalse(any("Suspiciously cheap" in f
                             for f in strategies.triage_flags(cheap, 65)))

    def test_divergent_multiples(self):
        flags = strategies.triage_flags(_strategy_row(pe=20.0, forward_pe=9.0), 80)
        self.assertTrue(any("Divergent multiples" in f for f in flags))

    def test_payout_stress(self):
        for over in ({"payout_ratio": 0.65}, {"fcf_coverage": 1.0}):
            flags = strategies.triage_flags(_strategy_row(**over), 80)
            self.assertTrue(any("Payout stress" in f for f in flags), over)

    def _stressed(self, **over):
        return any("Payout stress" in f
                   for f in strategies.triage_flags(_strategy_row(**over), 80))

    def test_payout_stress_judges_reits_on_ffo_not_earnings(self):
        # A healthy equity REIT: earnings payout is huge (depreciation crushes
        # the denominator) and GAAP FCF barely covers the dividend, but FFO —
        # what it actually distributes out of — is comfortable. The generic
        # earnings test fired on every REIT alive; the FFO test must not.
        reit = dict(sector="Real Estate", industry="REIT—Residential",
                    payout_ratio=2.65, fcf_coverage=1.0)
        self.assertFalse(self._stressed(**reit, ffo_payout=0.74, ffo_coverage=1.35))
        # Above the comfortable band and nearing the 100% line: real stress.
        self.assertTrue(self._stressed(**reit, ffo_payout=0.95, ffo_coverage=1.05))
        # Negative FFO leaves the payout undefined — coverage still catches it.
        self.assertTrue(self._stressed(**reit, ffo_payout=None, ffo_coverage=-1.2))
        # No FFO line at all: fall back to FCF coverage, as the graders do.
        self.assertTrue(self._stressed(**reit, ffo_payout=None, ffo_coverage=None))
        self.assertFalse(self._stressed(sector="Real Estate",
                                        industry="REIT—Residential",
                                        payout_ratio=2.65, fcf_coverage=3.0,
                                        ffo_payout=None, ffo_coverage=None))

    def test_payout_stress_ignores_bank_fcf(self):
        # A bank's lending runs through operating cash flow, so a year of loan
        # growth prints a deeply negative FCF (JPM: −$148B) that says nothing
        # about the dividend. Earnings payout is the only leg that applies.
        bank = dict(sector="Financial Services", industry="Banks - Diversified")
        self.assertFalse(self._stressed(**bank, payout_ratio=0.26, fcf_coverage=-8.9))
        self.assertTrue(self._stressed(**bank, payout_ratio=0.65, fcf_coverage=3.0))

    def test_payout_stress_judges_mreits_on_earnings_at_100pct(self):
        # Mortgage REITs hold securities, not buildings: earnings are the right
        # denominator, but they distribute nearly all of them by design, so only
        # a payout above 100% — the dividend eating book value — is stress.
        mreit = dict(sector="Real Estate", industry="REIT — Mortgage")
        self.assertFalse(self._stressed(**mreit, payout_ratio=0.95))
        self.assertTrue(self._stressed(**mreit, payout_ratio=1.10))
        # FCF is an income-statement artifact for a securities portfolio and
        # must not fire the flag on its own.
        self.assertFalse(self._stressed(**mreit, payout_ratio=0.95, fcf_coverage=0.1))

    def test_crowded_short_and_high_beta(self):
        flags = strategies.triage_flags(
            _strategy_row(short_interest=0.20, beta=2.0), 80)
        self.assertTrue(any("Crowded short" in f for f in flags))
        self.assertTrue(any("High beta" in f for f in flags))

    def test_data_sanity_flags(self):
        flags = strategies.triage_flags(
            _strategy_row(pb=45.0, enterprise_value=-1e9), 80)
        self.assertTrue(any("P/B > 40" in f for f in flags))
        self.assertTrue(any("negative EV" in f for f in flags))

    def test_grade_row_joins_flags_into_string(self):
        g = strategies.grade_row(_strategy_row(peg=3.5, beta=2.0))
        self.assertIn("Priced for perfection", g["strategy_1_flags"])
        self.assertIn("High beta", g["strategy_1_flags"])
        self.assertIn(" · ", g["strategy_1_flags"])


class TestStrategyCompounder(unittest.TestCase):

    def test_strong_row_is_compounder(self):
        score, verdict = strategies.grade_compounder(_strategy_row())
        self.assertEqual(score, 100)
        self.assertEqual(verdict, "Compounder")

    def test_solvency_guard_caps_at_35(self):
        score, verdict = strategies.grade_compounder(_strategy_row(altman_z=1.2))
        self.assertLessEqual(score, 35)
        self.assertEqual(verdict, "Pass")

    def test_no_10y_history_loses_track_record_points(self):
        score, _ = strategies.grade_compounder(_strategy_row(perf_10y=None))
        self.assertEqual(score, 90)

    def test_roe_leverage_cap_de_above_1(self):
        full, _ = strategies.grade_compounder(_strategy_row(debt_to_equity=90.0))
        capped, _ = strategies.grade_compounder(_strategy_row(debt_to_equity=110.0))
        self.assertEqual(full - capped, 5)

    def test_negative_equity_zeroes_roe_points(self):
        # ROE on negative equity is noise: full 10 ROE points lost vs base.
        score, _ = strategies.grade_compounder(
            _strategy_row(total_equity=-5e9, debt_to_equity=-20.0))
        self.assertEqual(score, 90)

    def test_financial_rubric_full_marks(self):
        # Bank rubric: ROE 30 + net margin 20 + (payout 10 + Piotroski 10)
        # + track record 20 + valuation 10 = 100.
        row = _strategy_row(sector="Financial Services",
                            industry="Banks—Diversified",
                            roe=0.18, profit_margin=0.25,
                            payout_ratio=0.35, piotroski_f=8)
        score, verdict = strategies.grade_compounder(row)
        self.assertEqual(score, 100)
        self.assertEqual(verdict, "Compounder")

    def test_valuation_pillar_tiers(self):
        # Full tier (10): default row (peg 1.2) scores 100. Second tier (5):
        # peg unavailable, P/FCF 30 (< 40). Zero: both unavailable.
        mid, _ = strategies.grade_compounder(_strategy_row(peg=None, p_fcf=30.0))
        self.assertEqual(mid, 95)
        none, _ = strategies.grade_compounder(_strategy_row(peg=None, p_fcf=None))
        self.assertEqual(none, 90)

    def test_cagr_pct_edges(self):
        self.assertIsNone(strategies._cagr_pct(None, 5))
        self.assertEqual(strategies._cagr_pct(-100.0, 5), -100.0)
        # +300% over 10y -> 4^(1/10) - 1 ~ 14.87%/yr
        self.assertAlmostEqual(strategies._cagr_pct(300.0, 10), 14.87, places=2)

    def test_thin_margins_with_strong_returns_floor_moat(self):
        # Same escape hatch as S1's profitability pillar: the fixture's ROIC 18
        # floors the moat pillar at half (10/20) for a thin-margin model.
        thin = dict(profit_margin=0.02, operating_margin=0.04, gross_margin=0.20)
        _, _, pillars = strategies._grade_compounder(_strategy_row(**thin))
        pill = next(p for p in pillars if p["k"] == "Margin moat")
        self.assertEqual(pill["p"], 10)
        self.assertIn("floored", pill["d"])
        _, _, p2 = strategies._grade_compounder(
            _strategy_row(roic=5.0, roe=0.10, **thin))
        self.assertEqual(next(p for p in p2 if p["k"] == "Margin moat")["p"], 0.0)


class TestStrategyDefensive(unittest.TestCase):

    def _value_row(self):
        return _strategy_row(pe=10.0, pb=1.2, p_fcf=12.0, ev_ebitda=8.0,
                             current_ratio=2.5, debt_to_equity=30.0,
                             altman_z=4.0, div_yield=3.0, payout_ratio=0.4,
                             fcf_coverage=2.0)

    def test_value_row_scores_100(self):
        score, verdict = strategies.grade_defensive(self._value_row())
        self.assertEqual(score, 100)
        self.assertEqual(verdict, "Value candidate")

    def test_growth_stock_fails_asset_backing(self):
        # The default row (P/B 5, P/E 20): pillar B = 0, verdict below Value.
        score, verdict = strategies.grade_defensive(_strategy_row())
        self.assertLess(score, 80)
        self.assertGreaterEqual(score, 50)   # still healthy, just not cheap

    def test_graham_multiplier(self):
        # P/E 14 × P/B 1.6 = 22.4 <= 22.5 -> +7 vs the same row at P/B 1.7.
        # ev_ebitda=8 keeps both totals integral (no half-point rounding).
        ok, _ = strategies.grade_defensive(_strategy_row(pe=14.0, pb=1.6, ev_ebitda=8.0))
        over, _ = strategies.grade_defensive(_strategy_row(pe=14.0, pb=1.7, ev_ebitda=8.0))
        self.assertEqual(ok - over, 7)

    def test_loss_maker_gets_no_pe_points(self):
        score_neg, _ = strategies.grade_defensive(_strategy_row(pe=None, income=-1e9))
        score_pos, _ = strategies.grade_defensive(self._value_row())
        self.assertLess(score_neg, score_pos)

    def _hist(self, ni_pos, ni_yrs, fcf_pos, fcf_yrs, **over):
        return self._value_row() | {
            "ni_positive_years": ni_pos, "ni_years": ni_yrs,
            "fcf_positive_years": fcf_pos, "fcf_years": fcf_yrs} | over

    def test_s3_quality_scores_consistency_not_latest_year_sign(self):
        # The point of the stability legs: a business profitable in 3 of 4 years
        # that took one loss last year is not the same as a chronic loss-maker,
        # though the old NI>0 sign test scored them identically (both zero).
        one_bad = strategies.grade_defensive(self._hist(3, 4, 4, 4, income=-1e9))
        chronic = strategies.grade_defensive(self._hist(0, 4, 0, 4, income=-1e9))
        self.assertGreater(one_bad[0], chronic[0])

    def test_s3_quality_one_bad_year_takes_half_not_zero(self):
        clean, _v, pillars = strategies._grade_defensive(self._hist(4, 4, 4, 4))
        bad, _v2, pillars_bad = strategies._grade_defensive(self._hist(3, 4, 4, 4))
        d = next(p for p in pillars if p["k"] == "Earnings quality")
        d_bad = next(p for p in pillars_bad if p["k"] == "Earnings quality")
        self.assertEqual(d["p"] - d_bad["p"], 2.5)      # half of the 5-pt NI leg
        self.assertIn("3/4 yrs", d_bad["d"])

    def test_s3_quality_two_bad_years_zeroes_the_leg(self):
        _s, _v, pillars = strategies._grade_defensive(self._hist(2, 4, 4, 4))
        d = next(p for p in pillars if p["k"] == "Earnings quality")
        self.assertEqual(d["p"], 15.0)                  # 0 NI + 5 FCF + 10 F-score

    def test_s3_quality_falls_back_to_sign_without_history(self):
        # A young company (one usable period) has no consistency to judge and
        # must not be zeroed for a track record it cannot yet have.
        _s, _v, pillars = strategies._grade_defensive(self._hist(1, 1, 1, 1))
        d = next(p for p in pillars if p["k"] == "Earnings quality")
        self.assertEqual(d["p"], 20.0)
        self.assertIn("NI +", d["d"])                   # sign text, not "1/1 yrs"

    def test_s3_quality_history_keys_absent_preserves_sign_behaviour(self):
        # Rows built before the counts existed (and any row whose statements
        # lack the line) still score exactly as the sign test scored them.
        self.assertEqual(strategies.grade_defensive(self._value_row())[0], 100)

    def test_s3_pillar_c_legs_and_weights(self):
        _s, _v, pillars = strategies._grade_defensive(self._value_row())
        c = next(p for p in pillars if p["k"] == "Financial strength")
        self.assertEqual(c["p"], 25.0)          # 8 cover + 7 net cash + 5 CR + 5 Z'
        for frag in ("int cover", "FCF/net debt", "CR", "Altman Z′"):
            self.assertIn(frag, c["d"])

    def test_s3_interest_coverage_bands(self):
        def cov_pts(ic):
            _s, _v, pillars = strategies._grade_defensive(
                self._value_row() | {"interest_coverage": ic})
            return next(p for p in pillars if p["k"] == "Financial strength")["p"]
        self.assertEqual(cov_pts(12.0), 25.0)   # > 8x  -> full 8
        self.assertEqual(cov_pts(5.0), 21.0)    # > 4x  -> half 4
        self.assertEqual(cov_pts(1.5), 17.0)    # thin  -> 0
        self.assertEqual(cov_pts(None), 17.0)   # missing earns 0, per the doc

    def test_s3_current_ratio_bands_are_modern_not_graham(self):
        # Graham's 2.0/1.5 zeroed 80% of a 118-name universe — a leg the whole
        # population fails measures nothing. Banded at 1.5/1.0 the median large
        # cap (~1.1) now scores, and only a genuinely tight balance sheet zeroes.
        def cr_pts(cr):
            _s, _v, pillars = strategies._grade_defensive(
                self._value_row() | {"current_ratio": cr})
            return next(p for p in pillars
                        if p["k"] == "Financial strength")["p"]
        self.assertEqual(cr_pts(2.5), 25.0)     # > 1.5 -> full 5
        self.assertEqual(cr_pts(1.6), 25.0)     # > 1.5 -> full 5
        self.assertEqual(cr_pts(1.1), 22.5)     # > 1.0 -> half 2.5 (the median)
        self.assertEqual(cr_pts(0.9), 20.0)     # <= 1.0 -> 0
        self.assertEqual(cr_pts(None), 20.0)    # missing earns 0

    def test_s3_debt_free_takes_full_coverage_points(self):
        # Nothing to cover -> the leg cannot be failed for a blank coupon.
        row = self._value_row() | {"total_debt": 0.0, "interest_coverage": None}
        _s, _v, pillars = strategies._grade_defensive(row)
        c = next(p for p in pillars if p["k"] == "Financial strength")
        self.assertEqual(c["p"], 25.0)
        self.assertIn("debt-free", c["d"])

    def test_s3_fcf_net_debt_is_sign_safe(self):
        # The trap this leg is shaped to avoid: net-debt/FCF would flip a
        # cash-burner's ratio negative and read it as unlevered. FCF/net debt
        # fails it instead. (Net debt 30bn against negative FCF.)
        row = self._value_row() | {"total_debt": 50e9, "total_cash": 20e9,
                                   "fcf": -5e9}
        _s, _v, pillars = strategies._grade_defensive(row)
        c = next(p for p in pillars if p["k"] == "Financial strength")
        self.assertEqual(c["p"], 18.0)          # 8 + 0 + 5 + 5, the leg zeroed

    def test_s3_net_cash_takes_full_repayment_points(self):
        rich = self._value_row() | {"total_debt": 10e9, "total_cash": 20e9}
        poor = self._value_row() | {"total_debt": 50e9, "total_cash": 1e9,
                                    "fcf": 2e9}     # 2/49 = 0.04x -> 0
        r = strategies._grade_defensive(rich)[2]
        p = strategies._grade_defensive(poor)[2]
        self.assertIn("net cash",
                      next(x for x in r if x["k"] == "Financial strength")["d"])
        self.assertEqual(
            next(x for x in r if x["k"] == "Financial strength")["p"]
            - next(x for x in p if x["k"] == "Financial strength")["p"], 7.0)

    def test_s3_strength_no_longer_penalises_cheapness(self):
        # The bug that motivated Z': under the classic Z a cheaper stock scored
        # *lower* financial strength, because Z's X4 is market-cap/liabilities.
        # Pillar C must now be identical for two rows differing only on price.
        cheap = self._value_row() | {"pe": 8.0, "pb": 0.8}
        rich = self._value_row() | {"pe": 24.0, "pb": 2.4}
        c_cheap = next(p for p in strategies._grade_defensive(cheap)[2]
                       if p["k"] == "Financial strength")
        c_rich = next(p for p in strategies._grade_defensive(rich)[2]
                      if p["k"] == "Financial strength")
        self.assertEqual(c_cheap["p"], c_rich["p"])

    def test_negative_equity_priced_through_altman_z_prime(self):
        # Pillar C no longer carries a D/E leg, so the sign-flip carve-out that
        # guarded it is gone too — nothing here divides by book equity any more.
        # Negative equity is priced through Altman Z' instead, whose X4 term is
        # equity/liabilities and goes negative with it (see TestAltmanZPrime,
        # which covers the computation this leg reads).
        strong = strategies.grade_defensive(self._value_row())[0]
        weak = strategies.grade_defensive(
            self._value_row() | {"total_equity": -5e9, "debt_to_equity": -20.0,
                                 "altman_z_prime": -0.4})[0]
        self.assertEqual(strong - weak, 5)      # the Z' leg, in full

    def test_buyback_negative_equity_still_scores_if_cash_generative(self):
        # The reason to price negative equity continuously rather than with a
        # carve-out: a buyback-driven negative book value on a net-cash balance
        # sheet with 15x interest cover is not fragile, and the old rubric's
        # flat 9-point strike could not tell it from a distressed one.
        row = self._value_row() | {"total_equity": -5e9, "debt_to_equity": -20.0,
                                   "altman_z_prime": 1.4}
        _s, _v, pillars = strategies._grade_defensive(row)
        c = next(p for p in pillars if p["k"] == "Financial strength")
        self.assertEqual(c["p"], 22.5)          # 8 cover + 7 net cash + 5 CR + 2.5 Z'

    def test_non_payer_capped_at_85(self):
        row = self._value_row() | {"div_yield": None, "years_div_increase": None,
                                   "payout_ratio": None, "fcf_coverage": None}
        score, verdict = strategies.grade_defensive(row)
        self.assertEqual(score, 85)
        self.assertEqual(verdict, "Value candidate")

    def test_dividend_half_tiers(self):
        # Yield 1% (2.5) + 5 yrs of increases (2.5) + payout 0.70/cov 1.2
        # (2.5) = 7.5 -> 85 + 7.5 = 92.5, rounded half-up to 93.
        row = self._value_row() | {"div_yield": 1.0, "years_div_increase": 5,
                                   "payout_ratio": 0.70, "fcf_coverage": 1.2}
        score, _ = strategies.grade_defensive(row)
        self.assertEqual(score, 93)

    def test_financials_use_roa_for_strength(self):
        row = _strategy_row(sector="Financial Services",
                            industry="Insurance—Diversified", current_ratio=None,
                            debt_to_equity=None, altman_z=None, roa=0.02,
                            pe=10.0, pb=1.2, p_fcf=12.0, ev_ebitda=8.0,
                            div_yield=3.0, payout_ratio=0.4, fcf_coverage=2.0)
        score, _ = strategies.grade_defensive(row)
        self.assertEqual(score, 100)


class TestGradeRowComposite(unittest.TestCase):

    def test_grade_row_keys_and_min(self):
        g = strategies.grade_row(_strategy_row())
        for k in ("strategy_1", "strategy_2", "strategy_3", "strategy_min",
                  "strategy_1_verdict", "strategy_2_verdict", "strategy_3_verdict",
                  "strategy_1_flags"):
            self.assertIn(k, g)
        self.assertEqual(g["strategy_1_flags"], "")   # healthy row: no flags
        self.assertEqual(g["strategy_min"],
                         min(g["strategy_1"], g["strategy_2"], g["strategy_3"]))

    def test_quarantine_makes_min_none(self):
        g = strategies.grade_row(_strategy_row(total_equity=None))
        self.assertIsNone(g["strategy_1"])
        self.assertIsNone(g["strategy_min"])

    def test_grade_row_attaches_pillar_detail(self):
        g = strategies.grade_row(_strategy_row())
        for k in ("strategy_1_detail", "strategy_2_detail", "strategy_3_detail"):
            self.assertIsInstance(g[k], list)
            self.assertTrue(g[k])                    # scored row has pillars
            for p in g[k]:
                self.assertEqual(set(p), {"k", "p", "m", "d"})

    def test_quarantine_has_empty_detail(self):
        # Missing critical data -> nothing computable -> empty breakdown.
        quar = strategies.grade_row(_strategy_row(total_equity=None))
        self.assertEqual(quar["strategy_1_detail"], [])

    def test_kill_keeps_breakdown_with_disqualified_row(self):
        # A Stage 1 kill forces the score to 0 but retains the Stage 2 pillar
        # breakdown so the tooltip can show the underlying values, ending with a
        # signed 'Disqualified' adjustment row.
        kill = strategies.grade_row(_strategy_row(income=-1e9, fcf=-1e9))
        self.assertEqual(kill["strategy_1"], 0)
        self.assertTrue(kill["strategy_1_verdict"].startswith("Discard · "))
        detail = kill["strategy_1_detail"]
        self.assertTrue(detail)                            # values retained
        self.assertEqual(detail[-1]["k"], "Disqualified")
        self.assertEqual(detail[-1]["m"], 0)               # signed adjustment
        # Pillars must still re-sum to the reported (zero) score.
        self.assertEqual(
            strategies._round_score(sum(p["p"] for p in detail)), 0)


class TestPillarBreakdownInvariant(unittest.TestCase):
    """The derivation is a by-product of the real scoring, so a scored row's
    pillar points must always re-sum to its score — this guards against the
    breakdown drifting away from the number it explains."""

    def _rows(self):
        return [
            _strategy_row(),                                 # capital-intensive
            _strategy_row(total_equity=-5e9, debt_to_equity=-20.0),
            _strategy_row(sector="Financial Services",
                          industry="Banks—Diversified"),
            _strategy_row(sector="Real Estate", industry="REIT—Retail"),
            _strategy_row(sector="Real Estate", industry="REIT—Retail",
                          ffo=5e9, p_ffo=10.0, ffo_payout=0.75, ffo_coverage=1.3),
            _strategy_row(sector="Real Estate", industry="REIT—Retail",
                          piotroski_f=2),                        # low-Piotroski flag path
            _strategy_row(sector="Real Estate", industry="REIT - Mortgage",
                          payout_ratio=1.4, pb=0.9, debt_to_equity=760.0,
                          bvps_growth=-6.0),                     # mreit
            _strategy_row(sector="Real Estate", industry="REIT - Mortgage",
                          total_equity=-1e9, debt_to_equity=-20.0),  # mreit + neg equity
            _strategy_row(sector="Technology", industry="Software—Infrastructure"),
            _strategy_row(sector="Energy", industry="Oil & Gas E&P"),
            _strategy_row(sector="Technology", industry="Software", wacc=25.0),
            _strategy_row(wacc=20.0),                        # neg-spread cap
            _strategy_row(altman_z=1.2),                     # solvency guard
            _strategy_row(pe=10.0, pb=1.2, p_fcf=12.0, ev_ebitda=8.0),
            _strategy_row(perf_10y=None, peg=None, p_fcf=None),
        ]

    def test_pillars_resum_to_score(self):
        graders = (("s1", strategies._grade_triage),
                   ("s2", strategies._grade_compounder),
                   ("s3", strategies._grade_defensive))
        for i, row in enumerate(self._rows()):
            for name, fn in graders:
                score, _verdict, pillars = fn(dict(row))
                if score is None or not pillars:
                    continue
                total = strategies._round_score(sum(p["p"] for p in pillars))
                self.assertEqual(total, score, f"row {i} {name}")

    def test_pillar_points_never_exceed_their_max(self):
        # (adjustment rows carry m=0 and a signed delta — skip those)
        for row in self._rows():
            for fn in (strategies._grade_triage, strategies._grade_compounder,
                       strategies._grade_defensive):
                _score, _v, pillars = fn(dict(row))
                for p in pillars:
                    if p["m"] > 0:
                        self.assertLessEqual(p["p"], p["m"])
                        self.assertGreaterEqual(p["p"], 0)

    def test_empty_row_grades_without_crashing(self):
        g = strategies.grade_row({"ticker": "X"})
        self.assertIsNone(g["strategy_1"])       # quarantined
        self.assertIsInstance(g["strategy_2"], int)
        self.assertIsInstance(g["strategy_3"], int)


class TestStrategyRobustness(unittest.TestCase):
    """Every grader stays in [0, 100] (or None) and never raises, whatever
    shape the row arrives in."""

    def test_scores_bounded_on_adversarial_rows(self):
        rows = [
            {"ticker": "X"},
            {k: None for k in _strategy_row()},
            _strategy_row(income=-1e12, fcf=-1e12, total_debt=1e12,
                          total_equity=-1e12, ebitda=-1e12),
            _strategy_row(roe=-50.0, roic=-500.0, wacc=50.0,
                          debt_to_equity=-999.0, pe=-5.0, pb=-2.0, peg=-1.0,
                          altman_z=-10.0, piotroski_f=0, current_ratio=0.01,
                          quick_ratio=0.0, ebitda=-1e12),
            _strategy_row(pe=1e9, pb=1e9, peg=1e9, p_fcf=1e9, ev_ebitda=1e9,
                          perf_5y=1e9, perf_10y=1e9),
        ]
        graders = (strategies.grade_triage, strategies.grade_compounder,
                   strategies.grade_defensive)
        for i, row in enumerate(rows):
            for fn in graders:
                score, verdict = fn(dict(row))
                if score is not None:
                    self.assertGreaterEqual(score, 0, (i, fn.__name__))
                    self.assertLessEqual(score, 100, (i, fn.__name__))
                self.assertIsInstance(verdict, str)
            g = strategies.grade_row(dict(row))
            self.assertIsInstance(g["strategy_1_flags"], str)


class TestStrategyIntegration(unittest.TestCase):

    def test_metric_cols_include_strategy_columns(self):
        keys = [k for k, _ in app._METRIC_COLS]
        for k in ("strategy_1", "strategy_2", "strategy_3", "strategy_min"):
            self.assertIn(k, keys)
            self.assertNotIn(k, app._PCT_KEYS)   # raw 0-100, not a fraction

    def test_panel_order_includes_strategies(self):
        self.assertIn(("strategies", "Strategy Ratings"), app._PANEL_ORDER)

    def test_strategy_panel_formats_scores(self):
        p = app._strategy_panel({
            "strategy_1": 72, "strategy_1_verdict": "Advance",
            "strategy_1_flags": "🔺 Priced for perfection",
            "strategy_2": 55, "strategy_2_verdict": "Quality watch",
            "strategy_3": 40, "strategy_3_verdict": "Expensive/weak",
            "strategy_min": 40,
        })
        self.assertEqual(p["S1 · Triage"], "72 / 100 — Advance")
        self.assertEqual(p["S1 · Flags"], "🔺 Priced for perfection")
        self.assertEqual(p["Min · All Strategies"], "40 / 100")

    def test_strategy_panel_flags_empty_string_reads_na(self):
        p = app._strategy_panel({"strategy_1": 90, "strategy_1_verdict": "Advance",
                                 "strategy_1_flags": ""})
        self.assertIsNone(p["S1 · Flags"])   # clean row: N/A, not ""

    def test_strategy_panel_handles_error_row(self):
        p = app._strategy_panel({"ticker": "X", "error": "No data"})
        self.assertIsNone(p["S2 · Compounder"])
        self.assertIsNone(p["Min · All Strategies"])


# ---------------------------------------------------------------------------
# 21. Analyst chat (chat.py + /api/chat)
# ---------------------------------------------------------------------------
import chat  # noqa: E402  (chat agent module)


class TestChatCompactRows(unittest.TestCase):

    def test_drops_sparks_details_and_missing_values(self):
        rows = [{"ticker": "AAPL", "pe": 30.123456789, "pb": None, "name": "",
                 "spark_1y": [1, 2, 3], "strategy_1_detail": [{"k": "x"}],
                 "strategy_1": 89, "strategy_1_flags": "🔺 Priced for perfection"}]
        out = chat.compact_rows(rows)
        self.assertEqual(out, [{"ticker": "AAPL", "pe": 30.1235,
                                "strategy_1": 89,
                                "strategy_1_flags": "🔺 Priced for perfection"}])

    def test_caps_at_max_rows_and_skips_junk(self):
        rows = [{"ticker": f"T{i}"} for i in range(chat.MAX_ROWS + 20)] + ["junk"]
        self.assertEqual(len(chat.compact_rows(rows)), chat.MAX_ROWS)
        self.assertEqual(chat.compact_rows(["junk", None]), [])


class TestChatSystemBlocks(unittest.TestCase):

    def test_static_block_is_cached_and_describes_units(self):
        blocks = chat.system_blocks([{"ticker": "AAPL", "pe": 30.0}])
        self.assertEqual(blocks[0]["cache_control"], {"type": "ephemeral"})
        self.assertIn("percentage points", blocks[0]["text"])
        self.assertIn("strategy_min", blocks[0]["text"])

    def test_rows_block_carries_data_and_cache_control(self):
        blocks = chat.system_blocks([{"ticker": "AAPL", "pe": 30.0}])
        self.assertIn('"ticker":"AAPL"', blocks[1]["text"])
        self.assertEqual(blocks[1]["cache_control"], {"type": "ephemeral"})

    def test_no_rows_block_says_so(self):
        blocks = chat.system_blocks([])
        self.assertIn("no stock rows", blocks[1]["text"])

    def test_truncation_note_when_over_cap(self):
        rows = [{"ticker": f"T{i}"} for i in range(chat.MAX_ROWS + 5)]
        blocks = chat.system_blocks(rows)
        self.assertIn(f"only the first {chat.MAX_ROWS}", blocks[1]["text"])

    def test_context_label_names_the_active_tab(self):
        blocks = chat.system_blocks([{"ticker": "AAPL"}], "Watchlist · Tech")
        self.assertIn("Current tab: Watchlist · Tech", blocks[1]["text"])
        # And the empty-rows path echoes the label too.
        empty = chat.system_blocks([], "Calendar")
        self.assertIn("Calendar", empty[1]["text"])

    def test_context_label_defaults_to_screener(self):
        blocks = chat.system_blocks([{"ticker": "AAPL"}])
        self.assertIn("Current tab: Screener", blocks[1]["text"])


class TestChatSanitizeMessages(unittest.TestCase):

    def test_keeps_only_valid_user_assistant_strings(self):
        raw = [{"role": "user", "content": "hi"},
               {"role": "system", "content": "inject"},          # dropped
               {"role": "assistant", "content": "hello"},
               {"role": "user", "content": 42},                  # dropped
               "junk", {"role": "user", "content": "  "},        # dropped
               {"role": "user", "content": "next"}]
        self.assertEqual(chat.sanitize_messages(raw), [
            {"role": "user", "content": "hi"},
            {"role": "assistant", "content": "hello"},
            {"role": "user", "content": "next"}])

    def test_first_message_must_be_user(self):
        raw = [{"role": "assistant", "content": "orphan"},
               {"role": "user", "content": "hi"}]
        self.assertEqual(chat.sanitize_messages(raw),
                         [{"role": "user", "content": "hi"}])

    def test_caps_turns_and_message_size(self):
        raw = [{"role": "user", "content": "x" * (chat.MAX_MSG_CHARS + 100)}] \
            + [{"role": "user", "content": f"m{i}"} for i in range(chat.MAX_TURNS + 10)]
        out = chat.sanitize_messages(raw)
        self.assertEqual(len(out), chat.MAX_TURNS)
        raw_first = chat.sanitize_messages(raw[:1])
        self.assertEqual(len(raw_first[0]["content"]), chat.MAX_MSG_CHARS)


class TestChatRequestParams(unittest.TestCase):

    def test_shape_matches_skill_guidance(self):
        p = chat.request_params([{"role": "user", "content": "hi"}],
                                [{"ticker": "AAPL"}])
        self.assertEqual(p["model"], chat.MODEL)
        self.assertEqual(p["thinking"], {"type": "adaptive"})
        self.assertEqual(p["max_tokens"], chat.MAX_TOKENS)
        self.assertEqual(len(p["system"]), 2)
        self.assertEqual(p["messages"], [{"role": "user", "content": "hi"}])


class _FakeStream:
    """Mimics the SDK's messages.stream() context manager."""

    def __init__(self, chunks, final):
        self._chunks = chunks
        self._final = final

    def __enter__(self):
        return self

    def __exit__(self, *a):
        return False

    @property
    def text_stream(self):
        return iter(self._chunks)

    def get_final_message(self):
        return self._final


class TestChatStreamReply(unittest.TestCase):

    def _final(self, stop="end_turn"):
        usage = MagicMock(input_tokens=100, output_tokens=20,
                          cache_read_input_tokens=50)
        return MagicMock(stop_reason=stop, usage=usage)

    def test_streams_text_then_done(self):
        client = MagicMock()
        client.messages.stream.return_value = _FakeStream(
            ["Hello ", "world"], self._final())
        with patch("anthropic.Anthropic", return_value=client):
            events = list(chat.stream_reply(
                [{"role": "user", "content": "hi"}], [{"ticker": "AAPL"}]))
        self.assertEqual(events[0], {"text": "Hello "})
        self.assertEqual(events[1], {"text": "world"})
        self.assertTrue(events[2]["done"])
        self.assertEqual(events[2]["output_tokens"], 20)
        # The request was built by request_params (model + adaptive thinking).
        kwargs = client.messages.stream.call_args.kwargs
        self.assertEqual(kwargs["model"], chat.MODEL)
        self.assertEqual(kwargs["thinking"], {"type": "adaptive"})

    def test_refusal_stop_reason_becomes_error_event(self):
        client = MagicMock()
        client.messages.stream.return_value = _FakeStream([], self._final("refusal"))
        with patch("anthropic.Anthropic", return_value=client):
            events = list(chat.stream_reply(
                [{"role": "user", "content": "hi"}], []))
        self.assertIn("declined", events[-1]["error"])

    def test_empty_history_is_an_error_not_an_api_call(self):
        events = list(chat.stream_reply([], []))
        self.assertEqual(len(events), 1)
        self.assertIn("Empty", events[0]["error"])

    def test_auth_error_yields_actionable_message(self):
        import anthropic as _an
        err = _an.AuthenticationError.__new__(_an.AuthenticationError)
        with patch("anthropic.Anthropic", side_effect=err):
            events = list(chat.stream_reply(
                [{"role": "user", "content": "hi"}], []))
        self.assertIn("ANTHROPIC_API_KEY", events[0]["error"])

    def test_generic_exception_never_raises(self):
        with patch("anthropic.Anthropic", side_effect=RuntimeError("boom")):
            events = list(chat.stream_reply(
                [{"role": "user", "content": "hi"}], []))
        self.assertEqual(events, [{"error": "boom"}])

    def test_missing_credentials_yield_actionable_message(self):
        # No credentials at all: the SDK's client constructor raises TypeError
        # ("Could not resolve authentication method…"), not AuthenticationError.
        with patch("anthropic.Anthropic",
                   side_effect=TypeError("Could not resolve authentication method")):
            events = list(chat.stream_reply(
                [{"role": "user", "content": "hi"}], []))
        self.assertIn("ANTHROPIC_API_KEY", events[0]["error"])


class TestChatEndpoint(unittest.TestCase):
    """/api/chat over a real server, with the agent stubbed out."""

    @classmethod
    def setUpClass(cls):
        cls.httpd, cls.url = _start_server()

    @classmethod
    def tearDownClass(cls):
        cls.httpd.shutdown()

    def _post_chat(self, body):
        import urllib.request
        req = urllib.request.Request(
            self.url + "/api/chat", data=json.dumps(body).encode(),
            headers={"Content-Type": "application/json"})
        with urlopen(req, timeout=10) as r:
            return r.status, r.read().decode(), r.headers.get("Content-Type", "")

    def test_streams_sse_events(self):
        def fake(messages, rows, context_label=None):
            yield {"text": "Hi "}
            yield {"text": "there"}
            yield {"done": True}
        with patch.object(app.chat_agent, "stream_reply", side_effect=fake):
            status, body, ctype = self._post_chat(
                {"messages": [{"role": "user", "content": "hi"}], "rows": []})
        self.assertEqual(status, 200)
        self.assertIn("text/event-stream", ctype)
        frames = [json.loads(f[6:]) for f in body.strip().split("\n\n")]
        self.assertEqual(frames[0], {"text": "Hi "})
        self.assertEqual(frames[-1], {"done": True})

    def test_error_events_pass_through(self):
        with patch.object(app.chat_agent, "stream_reply",
                          side_effect=lambda m, r, c=None: iter([{"error": "nope"}])):
            _status, body, _ = self._post_chat({"messages": [], "rows": []})
        self.assertIn('{"error": "nope"}', body)

    def test_payload_including_context_label_reaches_the_agent(self):
        seen = {}
        def fake(messages, rows, context_label=None):
            seen["messages"], seen["rows"], seen["label"] = messages, rows, context_label
            yield {"done": True}
        with patch.object(app.chat_agent, "stream_reply", side_effect=fake):
            self._post_chat({"messages": [{"role": "user", "content": "q"}],
                             "rows": [{"ticker": "AAPL"}],
                             "context_label": "Dashboard"})
        self.assertEqual(seen["messages"], [{"role": "user", "content": "q"}])
        self.assertEqual(seen["rows"], [{"ticker": "AAPL"}])
        self.assertEqual(seen["label"], "Dashboard")


import wiki  # noqa: E402  (Wikipedia company-context module)


class TestBriefSummary(unittest.TestCase):
    def test_none_and_short_pass_through(self):
        self.assertIsNone(app._brief_summary(None))
        self.assertIsNone(app._brief_summary(""))
        self.assertEqual(app._brief_summary("Makes shoes."), "Makes shoes.")

    def test_long_text_cut_at_sentence_boundary(self):
        text = ("Acme designs widgets for industry. " * 3
                + "It also does many other things " * 20)
        out = app._brief_summary(text, limit=120)
        self.assertLessEqual(len(out), 120)
        self.assertTrue(out.endswith("."))          # ended on a full sentence

    def test_no_sentence_boundary_falls_back_to_word_cut(self):
        text = "word " * 100                        # no '. ' anywhere
        out = app._brief_summary(text, limit=50)
        self.assertLessEqual(len(out), 51)          # cut + ellipsis
        self.assertTrue(out.endswith("…"))


class TestWikiPureFunctions(unittest.TestCase):
    def test_clean_company_name(self):
        self.assertEqual(wiki.clean_company_name("NIKE, Inc."), "NIKE")
        self.assertEqual(wiki.clean_company_name("Toyota Motor Corp."), "Toyota Motor")
        self.assertEqual(wiki.clean_company_name("Apple Inc"), "Apple")
        # double suffix sheds both layers
        self.assertEqual(wiki.clean_company_name("Zoom Video Holdings, Inc."),
                         "Zoom Video")
        # suffix must be its own word — 'Visa' keeps its 'sa', 'Cisco' its 'co'
        self.assertEqual(wiki.clean_company_name("Visa Inc."), "Visa")
        self.assertEqual(wiki.clean_company_name("Visa"), "Visa")
        self.assertEqual(wiki.clean_company_name("Cisco"), "Cisco")
        # never empties the name
        self.assertEqual(wiki.clean_company_name("Inc."), "Inc.")
        self.assertEqual(wiki.clean_company_name(None), "")

    def test_pick_title_skips_disambiguation_and_lists(self):
        titles = ["Apple (disambiguation)", "List of Apple products", "Apple Inc."]
        self.assertEqual(wiki.pick_title(titles, "Apple"), "Apple Inc.")
        self.assertIsNone(wiki.pick_title([], "X"))
        self.assertIsNone(wiki.pick_title(["Foo (disambiguation)"], "Foo"))

    def test_pick_ethics_sections_matches_and_dedupes_children(self):
        sections = [
            {"index": "1", "number": "1", "line": "History"},
            {"index": "8", "number": "8", "line": "Controversies"},
            {"index": "9", "number": "8.1", "line": "Child labour"},
            {"index": "10", "number": "8.2", "line": "Environmental record"},
            {"index": "12", "number": "9", "line": "Lawsuits"},
            {"index": "14", "number": "10", "line": "See also"},
        ]
        picked = wiki.pick_ethics_sections(sections)
        # children 8.1/8.2 folded into parent 8; 'See also'/'History' ignored
        self.assertEqual([s["index"] for s in picked], ["8", "12"])
        self.assertEqual(picked[0]["line"], "Controversies")

    def test_pick_ethics_sections_matches_orphan_subsection(self):
        # an ethics-flavoured subsection under an innocuous parent still matches
        sections = [
            {"index": "2", "number": "2", "line": "Operations"},
            {"index": "3", "number": "2.1", "line": "Labor practices"},
        ]
        picked = wiki.pick_ethics_sections(sections)
        self.assertEqual([s["index"] for s in picked], ["3"])

    def test_pick_ethics_sections_skips_unfetchable_and_caps(self):
        sections = [{"index": "T-1", "number": "5", "line": "Controversies"}]
        self.assertEqual(wiki.pick_ethics_sections(sections), [])  # transcluded
        many = [{"index": str(i), "number": str(i), "line": f"Lawsuit {i}"}
                for i in range(1, 12)]
        self.assertEqual(len(wiki.pick_ethics_sections(many)), wiki.MAX_SECTIONS)

    def test_section_blocks_structure_and_cleanup(self):
        # mirrors real MediaWiki output: h2 section heading, an h3 sub-heading,
        # a dropped style block + hatnote, a paragraph and a two-item list.
        html = (
            '<div class="mw-heading mw-heading2"><h2>Controversies</h2>'
            '<span class="mw-editsection">[edit]</span></div>'
            '<div class="mw-heading mw-heading3"><h3>Ad dispute</h3></div>'
            '<style>.hatnote{}</style>'
            '<div role="note" class="hatnote navigation-not-searchable">'
            'Main article: <a href="/x">Somewhere</a></div>'
            '<p>Acme was fined<sup>[9]</sup> &amp; sued.</p>'
            '<ul><li>case one</li><li>case two</li></ul>')
        blocks = wiki.section_blocks(html)
        self.assertEqual([b["t"] for b in blocks], ["h", "p", "li", "li"])
        self.assertEqual(blocks[0]["s"], "Ad dispute")       # h2 skipped, h3 kept
        self.assertEqual(blocks[1]["s"], "Acme was fined & sued.")
        self.assertEqual(blocks[2]["s"], "case one")
        # references, [edit] links and the "Main article" hatnote are gone
        joined = " ".join(b["s"] for b in blocks)
        self.assertNotIn("[9]", joined)
        self.assertNotIn("[edit]", joined)
        self.assertNotIn("Main article", joined)
        self.assertNotIn("<", joined)

    def test_section_blocks_drops_dangling_trailing_heading(self):
        html = "<p>Body.</p><h3>Empty tail heading</h3>"
        blocks = wiki.section_blocks(html)
        self.assertEqual([b["t"] for b in blocks], ["p"])

    def test_section_blocks_respects_char_cap(self):
        html = "".join(f"<p>{'x' * 100}</p>" for _ in range(10))
        blocks = wiki.section_blocks(html, max_chars=250)
        self.assertEqual(len(blocks), 3)                     # stops once over cap

    def test_ethics_heading_regex_spot_checks(self):
        yes = ["Controversies", "Criticism", "Human rights concerns",
               "Labour disputes", "Environmental record", "Antitrust cases",
               "Sexual harassment allegations", "Tax avoidance"]
        no = ["History", "Products", "Finances", "See also", "References",
              "Corporate affairs", "Sponsorships"]
        for h in yes:
            self.assertTrue(wiki.ETHICS_HEADING.search(h), h)
        for h in no:
            self.assertFalse(wiki.ETHICS_HEADING.search(h), h)


class TestWikiCompanyContext(unittest.TestCase):
    """company_context with the HTTP layer mocked out."""

    def _run(self, responses):
        calls = []

        def fake_get(params):
            calls.append(params)
            return responses[params["action"], params.get("prop") or params.get("list")]
        with patch.object(wiki, "_get", side_effect=fake_get):
            return wiki.company_context("NIKE, Inc."), calls

    def test_happy_path(self):
        responses = {
            ("query", "search"): {"query": {"search": [{"title": "Nike, Inc."}]}},
            ("parse", "sections"): {"parse": {"title": "Nike, Inc.", "sections": [
                {"index": "5", "number": "5", "line": "Controversies"}]}},
            ("parse", "text"): {"parse": {"text":
                "<h2>Controversies</h2><p>Sweatshop allegations.</p>"}},
        }
        out, _calls = self._run(responses)
        self.assertEqual(out["title"], "Nike, Inc.")
        self.assertEqual(out["url"], wiki.PAGE_URL + "Nike,_Inc.")
        self.assertEqual(len(out["sections"]), 1)
        self.assertEqual(out["sections"][0]["heading"], "Controversies")
        blocks = out["sections"][0]["blocks"]
        self.assertEqual(blocks, [{"t": "p", "s": "Sweatshop allegations."}])
        self.assertNotIn("error", out)

    def test_no_page_found(self):
        out, _ = self._run({("query", "search"): {"query": {"search": []}}})
        self.assertIsNone(out["title"])
        self.assertEqual(out["sections"], [])
        self.assertEqual(out["error"], "no Wikipedia page found")

    def test_network_error_is_reported_not_raised(self):
        import requests as _rq
        with patch.object(wiki, "_get",
                          side_effect=_rq.ConnectionError("boom")):
            out = wiki.company_context("Acme")
        self.assertIn("Wikipedia unreachable", out["error"])
        self.assertEqual(out["sections"], [])


class TestCompanyProfileEndpoint(unittest.TestCase):
    def test_company_profile_caches_and_tags_ticker(self):
        app.clear_cache()
        fake = {"query": "Acme", "title": "Acme", "url": "u", "sections": []}
        with patch.object(app, "get_info",
                          return_value={"shortName": "Acme Inc."}), \
             patch.object(app.wiki, "company_context",
                          return_value=dict(fake)) as cc:
            out1 = app.company_profile("ACME")
            out2 = app.company_profile("ACME")   # served from cache
        self.assertEqual(out1["ticker"], "ACME")
        self.assertEqual(out2, out1)
        self.assertEqual(cc.call_count, 1)
        self.assertEqual(cc.call_args[0][0], "Acme Inc.")
        app.clear_cache()


if __name__ == "__main__":
    unittest.main(verbosity=2)
